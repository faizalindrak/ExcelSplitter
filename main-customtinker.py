# split_gui.py
# GUI untuk split Excel per nilai unik dengan template-based rendering.
# Build exe: pyinstaller split_gui.spec
# Dependencies: customtkinter, pandas, openpyxl, reportlab (opsional untuk PDF Engine "reportlab")

import os
import re
import shutil
import subprocess
import threading
from pathlib import Path
import configparser
import tkinter as tk
import customtkinter as ctk
from tkinter import filedialog, messagebox

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ==== (Opsional) ReportLab untuk PDF pure-Python ====
try:
    from reportlab.lib.pagesizes import A4, LETTER, landscape, portrait
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
    REPORTLAB_AVAILABLE = True
except Exception:
    REPORTLAB_AVAILABLE = False


# ----------------- Helpers -----------------

def safe_file_part(s: str) -> str:
    s = "" if s is None else str(s)
    return re.sub(r'[:\\/\?\*\[\]<>|"]', "_", s).strip() or "Key"

def set_print_titles_and_area(ws, header_rows: int, last_col_idx: int, last_data_row: int):
    ws.print_title_rows = f"1:{header_rows}"
    last_col_letter = get_column_letter(last_col_idx if last_col_idx > 0 else 1)
    last_row = last_data_row if last_data_row >= (header_rows + 1) else (header_rows + 1)
    ws.print_area = f"A1:{last_col_letter}{last_row}"

def find_soffice(explicit_path: str | None = None) -> str | None:
    """Cari soffice.exe dari explicit, env, PATH, atau lokasi umum."""
    def _normalize(p):
        if not p:
            return None
        p = str(Path(p))
        if p.lower().endswith("soffice.exe") and Path(p).exists():
            return p
        prog = Path(p) / "soffice.exe"
        if prog.exists():
            return str(prog)
        prog = Path(p) / "program" / "soffice.exe"
        if prog.exists():
            return str(prog)
        return None

    # explicit
    s = _normalize(explicit_path)
    if s: return s
    # env
    s = _normalize(os.environ.get("LIBREOFFICE_PATH"))
    if s: return s
    # PATH
    s = shutil.which("soffice")
    if s: return s
    # lokasi umum
    common = [
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
        r"C:\PortableApps\LibreOfficePortable\App\libreoffice\program\soffice.exe",
        r"D:\PortableApps\LibreOfficePortable\App\libreoffice\program\soffice.exe",
    ]
    for c in common:
        if Path(c).exists():
            return c
    return None

def export_pdf_via_lo(xlsx_path: Path, soffice_path: str | None = None):
    exe = soffice_path or "soffice"
    cmd = [exe, "--headless", "--convert-to", "pdf", "--outdir", str(xlsx_path.parent), str(xlsx_path)]
    subprocess.run(cmd, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)


# ---------- ReportLab (pure-Python) PDF ----------
PAPER_MAP = {9: A4, 1: LETTER}

def _excel_col_width_to_points(w):
    return float(w) * 7.0 * 0.75 if w else 50.0

def _page_setup_from_template(ws):
    ps = ws.page_setup
    paper = PAPER_MAP.get(ps.paperSize, A4)
    orientation = str(getattr(ps, "orientation", "portrait")).lower()
    paper = landscape(paper) if orientation == "landscape" else portrait(paper)
    pm = ws.page_margins
    return paper, {
        "left": (pm.left or 0.7) * inch,
        "right": (pm.right or 0.7) * inch,
        "top": (pm.top or 0.75) * inch,
        "bottom": (pm.bottom or 0.75) * inch,
    }

def export_pdf_pure(group_df: pd.DataFrame, template_path: Path, header_rows: int, pdf_out: Path):
    if not REPORTLAB_AVAILABLE:
        raise RuntimeError("ReportLab belum terpasang. Jalankan: pip install reportlab")

    wb = load_workbook(template_path, read_only=True, data_only=True)
    ws = wb.active

    templ_headers, hidden_cols_idx, col_width_pts = [], set(), []
    empty_streak, c = 0, 1
    while c <= 500 and empty_streak < 5:
        val = ws.cell(row=header_rows, column=c).value
        col_dim = ws.column_dimensions.get(ws.cell(row=1, column=c).column_letter)
        hidden = bool(getattr(col_dim, "hidden", False))
        width = getattr(col_dim, "width", None)
        if val is None and width is None:
            empty_streak += 1
        else:
            empty_streak = 0
        templ_headers.append("" if val is None else str(val))
        if hidden:
            hidden_cols_idx.add(c - 1)
        col_width_pts.append(_excel_col_width_to_points(width))
        c += 1
    wb.close()

    # Map urutan kolom mengikuti header template bila cocok
    if templ_headers and any(h.strip() for h in templ_headers):
        ordered = [h for h in templ_headers if h and h in group_df.columns]
        if ordered:
            df = group_df[ordered].copy()
            keep_idx = [templ_headers.index(h) for h in ordered]
            col_width_pts = [col_width_pts[idx] for idx in keep_idx]
            hidden_cols_idx = {i for i in range(len(keep_idx)) if keep_idx[i] in hidden_cols_idx}
        else:
            df = group_df.copy()
    else:
        df = group_df.copy()

    # Drop kolom hidden
    keep_cols = [i for i in range(len(df.columns)) if i not in hidden_cols_idx] or list(range(len(df.columns)))
    df = df.iloc[:, keep_cols]
    col_width_pts = [col_width_pts[i] for i in keep_cols]

    # Susun data tabel (header + isi)
    table_data = [list(df.columns.astype(str))] + df.fillna("").astype(str).values.tolist()

    # Page setup
    pagesize, margins = _page_setup_from_template(ws)
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
    from reportlab.lib import colors
    doc = SimpleDocTemplate(
        str(pdf_out),
        pagesize=pagesize,
        leftMargin=margins["left"],
        rightMargin=margins["right"],
        topMargin=margins["top"],
        bottomMargin=margins["bottom"],
        title=pdf_out.stem,
    )

    tbl = Table(table_data, colWidths=col_width_pts or None, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTSIZE", (0, 1), (-1, -1), 9),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.black),
        ("LINEABOVE", (0, 0), (-1, 0), 0.75, colors.black),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F0F0F0")),
    ]))
    doc.build([tbl])


# ----------------- Split Logic -----------------

def split_excel_with_template(
    source_path: Path, sheet_name: str, key_col, template_path: Path, out_dir: Path,
    header_rows: int, pdf_engine: str = "reportlab", soffice_path: str | None = None,
    prefix: str = "", suffix: str = "", status_cb=None, progress_cb=None
):
    if status_cb is None: status_cb = lambda msg: None
    if progress_cb is None: progress_cb = lambda t, c: None

    if not source_path.exists():
        raise FileNotFoundError(f"Sumber tidak ditemukan: {source_path}")
    if not template_path.exists():
        raise FileNotFoundError(f"Template tidak ditemukan: {template_path}")
    out_dir.mkdir(parents=True, exist_ok=True)

    status_cb("Membaca sumber...")

    # Diagnostic logging
    try:
        file_size = source_path.stat().st_size if source_path.exists() else 0
        status_cb(f"Debug: File path: {source_path}")
        status_cb(f"Debug: File exists: {source_path.exists()}")
        status_cb(f"Debug: File size: {file_size:,} bytes ({file_size/1024/1024:.2f} MB)")
        status_cb(f"Debug: Sheet name: '{sheet_name}'")
        status_cb("Debug: Starting pd.read_excel...")

        # Try reading with timeout and error handling
        import time
        start_time = time.time()

        # First try to read just the header to test file accessibility
        status_cb("Debug: Testing file accessibility...")
        try:
            # Try to read just first few rows to test
            df_test = pd.read_excel(source_path, sheet_name=sheet_name, nrows=5, dtype=object)
            status_cb(f"Debug: Successfully read {len(df_test)} rows for testing")
        except Exception as test_e:
            status_cb(f"Debug: Test read failed: {str(test_e)}")
            raise test_e

        # Now read the full file
        df = pd.read_excel(source_path, sheet_name=sheet_name, dtype=object)

        elapsed = time.time() - start_time
        status_cb(f"Debug: Successfully read {len(df)} rows in {elapsed:.2f} seconds")

    except FileNotFoundError as e:
        status_cb(f"Debug: File not found: {e}")
        raise FileNotFoundError(f"File tidak ditemukan: {source_path}")
    except PermissionError as e:
        status_cb(f"Debug: Permission denied: {e}")
        raise PermissionError(f"Tidak ada akses ke file: {source_path}")
    except Exception as e:
        status_cb(f"Debug: Error reading Excel: {type(e).__name__}: {str(e)}")
        raise e

    # Tentukan kolom kunci
    if isinstance(key_col, int):
        if key_col < 1 or key_col > df.shape[1]:
            raise ValueError("Index kolom kunci di luar jangkauan DataFrame.")
        key_series = df.iloc[:, key_col - 1]
    else:
        if key_col not in df.columns:
            raise ValueError(f"Header kolom kunci '{key_col}' tidak ditemukan.")
        key_series = df[key_col]

    # Debug: Check for categorical data issues
    status_cb(f"Debug: Key column '{key_col}' data type: {key_series.dtype}")
    status_cb(f"Debug: Key column unique values: {len(key_series.unique())}")
    status_cb(f"Debug: Key column has null values: {key_series.isnull().sum()}")

    # Check for categorical columns in the entire DataFrame
    categorical_cols = []
    for col in df.columns:
        if df[col].dtype.name == 'category':
            categorical_cols.append(col)
            status_cb(f"Debug: Found categorical column: {col}")

    if categorical_cols:
        status_cb(f"Debug: Converting {len(categorical_cols)} categorical columns to string...")
        for col in categorical_cols:
            try:
                df[col] = df[col].astype(str)
                status_cb(f"Debug: Converted {col} to string")
            except Exception as conv_e:
                status_cb(f"Debug: Failed to convert {col}: {conv_e}")

    # Selaraskan urutan kolom ke header template jika cocok
    templ_cols = None
    try:
        wb_probe = load_workbook(template_path, read_only=True, data_only=True)
        ws_probe = wb_probe.active
        tmp, c, empty_streak = [], 1, 0
        while c <= 500 and empty_streak < 5:
            val = ws_probe.cell(row=header_rows, column=c).value
            if val is None or str(val).strip() == "":
                empty_streak += 1
            else:
                tmp.append(str(val).strip()); empty_streak = 0
            c += 1
        wb_probe.close()
        if tmp:
            templ_cols = tmp
    except Exception:
        templ_cols = None

    if templ_cols:
        exist = [col for col in templ_cols if col in df.columns]
        if exist:
            df = df[exist]

    # Debug: Check groupby operation
    status_cb("Debug: Starting groupby operation...")
    try:
        groups = df.groupby(key_series, dropna=False, sort=False)
        total, current = len(groups), 0
        status_cb(f"Debug: Groupby successful, found {total} groups")
        progress_cb(total, 0)
    except Exception as groupby_e:
        status_cb(f"Debug: Groupby error: {type(groupby_e).__name__}: {str(groupby_e)}")

        # Try multiple approaches to fix the issue
        if "categorical" in str(groupby_e).lower():
            status_cb("Debug: Attempting to fix categorical issue...")

            # Method 1: Try converting all categorical columns to object
            try:
                status_cb("Debug: Method 1 - Converting all categorical columns to object...")
                df_no_cat = df.copy()
                for col in df_no_cat.columns:
                    if df_no_cat[col].dtype.name == 'category':
                        df_no_cat[col] = df_no_cat[col].astype('object')
                groups = df_no_cat.groupby(key_series, dropna=False, sort=False)
                total, current = len(groups), 0
                status_cb(f"Debug: Method 1 successful, found {total} groups")
                progress_cb(total, 0)
                # Update df to use the fixed version
                df = df_no_cat
            except Exception as method1_e:
                status_cb(f"Debug: Method 1 failed: {method1_e}")

                # Method 2: Try using string conversion for groupby
                try:
                    status_cb("Debug: Method 2 - Using string keys for groupby...")
                    string_keys = key_series.astype(str)
                    groups = df.groupby(string_keys, dropna=False, sort=False)
                    total, current = len(groups), 0
                    status_cb(f"Debug: Method 2 successful, found {total} groups")
                    progress_cb(total, 0)
                except Exception as method2_e:
                    status_cb(f"Debug: Method 2 failed: {method2_e}")
                    raise groupby_e
        else:
            raise groupby_e

    for key_val, group in groups:
        current += 1
        status_cb(f"Proses [{current}/{total}] key={key_val}")
        progress_cb(total, current)

        # 1) Tulis XLSX dari template
        wb = load_workbook(template_path)
        ws = wb.active
        start_row = header_rows + 1

        # Comprehensive cleanup to eliminate all Excel repair warnings
        try:
            # 1. Remove all named ranges (common cause of repair warnings)
            if hasattr(wb, 'defined_names'):
                wb.defined_names = {}

            # 2. Remove external links
            if hasattr(wb, '_external_links') and wb._external_links:
                wb._external_links.clear()

            # 3. Remove external link relationships
            if hasattr(wb, '_external_link_rels'):
                wb._external_link_rels.clear()

            # 4. Remove all drawings completely (they often cause repair issues)
            if hasattr(ws, '_drawings'):
                ws._drawings = []

            # 5. Clean up worksheet-level named ranges
            if hasattr(ws, 'defined_names'):
                ws.defined_names = []

            # 6. Remove external link parts from the package
            try:
                from openpyxl.packaging.relationship import Relationship
                if hasattr(wb, '_rels'):
                    # Remove external link relationships
                    rels_to_remove = []
                    for rel in wb._rels:
                        if ('externalLink' in str(rel.target) or
                            'externalLinks' in str(rel.target) or
                            'drawing' in str(rel.target).lower()):
                            rels_to_remove.append(rel)
                    for rel in rels_to_remove:
                        wb._rels.remove(rel)
            except:
                pass

            # 7. Clean up any hyperlinks that might reference external data
            for row in ws.iter_rows():
                for cell in row:
                    if cell.hyperlink:
                        cell.hyperlink = None

        except Exception as cleanup_e:
            # If cleanup fails, continue anyway
            pass

        values = group.fillna("").values.tolist()

        # Copy formatting from template row to data rows
        # Use the first data row from template as formatting template
        template_row = start_row  # This should be the first data row in template

        for r_off, row_vals in enumerate(values, start=0):
            row_idx = start_row + r_off

            # Copy values
            for c_idx, v in enumerate(row_vals, start=1):
                ws.cell(row=row_idx, column=c_idx, value=v)

            # Copy formatting from template row to current row
            if r_off > 0:  # Don't copy formatting to the template row itself
                try:
                    # Copy formatting from template row to current row
                    for col_idx in range(1, len(row_vals) + 1):
                        template_cell = ws.cell(row=template_row, column=col_idx)
                        current_cell = ws.cell(row=row_idx, column=col_idx)

                        # Copy all formatting properties
                        if template_cell.has_style:
                            current_cell.font = template_cell.font.copy()
                            current_cell.fill = template_cell.fill.copy()
                            current_cell.border = template_cell.border.copy()
                            current_cell.alignment = template_cell.alignment.copy()
                            current_cell.number_format = template_cell.number_format
                            current_cell.protection = template_cell.protection.copy()

                except Exception as format_e:
                    # If formatting copy fails, continue without it
                    pass

        last_data_row = start_row + len(values) - 1
        set_print_titles_and_area(ws, header_rows, max(1, group.shape[1]), last_data_row)

        # Build filename with prefix and suffix
        key_part = safe_file_part(key_val)
        parts = []
        if prefix:
            parts.append(prefix)
        parts.append(key_part)
        if suffix:
            parts.append(suffix)

        out_name = " ".join(parts)
        xlsx_out = out_dir / f"{out_name}.xlsx"
        wb.save(xlsx_out)

        # 2) PDF (opsional)
        eng = (pdf_engine or "none").lower()
        if eng != "none":
            if eng == "reportlab":
                export_pdf_pure(group, template_path, header_rows, xlsx_out.with_suffix(".pdf"))
            elif eng == "libreoffice":
                export_pdf_via_lo(xlsx_out, soffice_path=soffice_path)

    status_cb("Selesai.")
    progress_cb(total, total)


# ----------------- GUI -----------------

class SplitApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Excel Splitter (Template-based)")
        self.geometry("980x700")
        ctk.set_appearance_mode("System")
        ctk.set_default_color_theme("blue")

        self.var_source = tk.StringVar()
        self.var_sheet = tk.StringVar()
        self.var_keycol = tk.StringVar()
        self.var_template = tk.StringVar()
        self.var_outdir = tk.StringVar()
        self.var_header_rows = tk.IntVar(value=5)
        self.var_pdf_engine = tk.StringVar(value="reportlab")
        self.var_lo_path = tk.StringVar()
        self.var_prefix = tk.StringVar(value="")
        self.var_suffix = tk.StringVar(value="")
        self.var_theme = tk.StringVar(value="blue")
        self.var_loaded_ini = tk.StringVar(value="")

        self.is_running = False
        self.worker_thread = None

        # Add validation traces
        self.var_source.trace_add("write", self.validate_source)
        self.var_template.trace_add("write", self.validate_template)
        self.var_outdir.trace_add("write", self.validate_outdir)
        self.var_lo_path.trace_add("write", self.validate_lo_path)

        self._build_ui()

    def change_theme(self, theme):
        ctk.set_default_color_theme(theme)
        # Destroy current UI and rebuild with new theme
        for widget in self.winfo_children():
            if widget != self:  # Don't destroy the root window
                widget.destroy()
        self._build_ui()
        messagebox.showinfo("Theme Changed", f"Theme changed to {theme} successfully!")

    def validate_source(self, *args):
        path = self.var_source.get().strip()
        if path and Path(path).exists() and Path(path).is_file():
            # Valid file
            if hasattr(self, 'entry_source'):
                self.entry_source.configure(border_color="#00FF00")  # Green
        else:
            if hasattr(self, 'entry_source'):
                self.entry_source.configure(border_color="#FF0000")  # Red

    def validate_template(self, *args):
        path = self.var_template.get().strip()
        if path and Path(path).exists() and Path(path).is_file():
            if hasattr(self, 'entry_template'):
                self.entry_template.configure(border_color="#00FF00")
        else:
            if hasattr(self, 'entry_template'):
                self.entry_template.configure(border_color="#FF0000")

    def validate_outdir(self, *args):
        path = self.var_outdir.get().strip()
        if path and Path(path).exists() and Path(path).is_dir():
            if hasattr(self, 'entry_outdir'):
                self.entry_outdir.configure(border_color="#00FF00")
        else:
            if hasattr(self, 'entry_outdir'):
                self.entry_outdir.configure(border_color="#FF0000")

    def validate_lo_path(self, *args):
        path = self.var_lo_path.get().strip()
        if not path or (Path(path).exists() and Path(path).is_file()):
            if hasattr(self, 'entry_lo_path'):
                self.entry_lo_path.configure(border_color="#00FF00")
        else:
            if hasattr(self, 'entry_lo_path'):
                self.entry_lo_path.configure(border_color="#FF0000")

    def _build_ui(self):
        pad = {"padx": 12, "pady": 10}

        # Main container
        main_frame = ctk.CTkFrame(self)
        main_frame.pack(fill="both", expand=True, padx=16, pady=16)

        # Top bar with left and right sections
        top_frame = ctk.CTkFrame(main_frame)
        top_frame.pack(fill="x", padx=10, pady=(10, 0))

        # Left section: Save/Load buttons
        left_frame = ctk.CTkFrame(top_frame, fg_color="transparent")
        left_frame.pack(side="left")
        self.btn_save_top = ctk.CTkButton(left_frame, text="Save .ini", command=self.save_ini, width=80)
        self.btn_save_top.pack(side="left", padx=(0, 8))
        self.btn_load_top = ctk.CTkButton(left_frame, text="Load .ini", command=self.load_ini, width=80)
        self.btn_load_top.pack(side="left", padx=(0, 8))
        self.lbl_loaded_ini = ctk.CTkLabel(left_frame, textvariable=self.var_loaded_ini, fg_color="transparent")
        self.lbl_loaded_ini.pack(side="left")

        # Right section: Theme selector
        right_frame = ctk.CTkFrame(top_frame, fg_color="transparent")
        right_frame.pack(side="right")
        ctk.CTkLabel(right_frame, text="Theme:").pack(side="left", padx=(0, 5))
        theme_combo = ctk.CTkComboBox(right_frame, values=["blue", "green", "dark-blue"], variable=self.var_theme, width=120, command=self.change_theme)
        theme_combo.pack(side="left")

        # Tabview
        self.tabview = ctk.CTkTabview(main_frame, width=800, height=600)
        self.tabview.pack(fill="both", expand=True, padx=10, pady=10)

        # Create tabs
        self.tabview.add("Input")
        self.tabview.add("Template")
        self.tabview.add("Output")
        self.tabview.add("Actions")

        # INPUT TAB
        input_tab = self.tabview.tab("Input")
        ctk.CTkLabel(input_tab, text="Source Excel").grid(row=0, column=0, sticky="e", **pad)
        self.entry_source = ctk.CTkEntry(input_tab, textvariable=self.var_source, width=520)
        self.entry_source.grid(row=0, column=1, sticky="we", **pad)
        self.btn_browse_source = ctk.CTkButton(input_tab, text="Browse...", command=self.browse_source)
        self.btn_browse_source.grid(row=0, column=2, **pad)

        ctk.CTkLabel(input_tab, text="Sheet Name").grid(row=1, column=0, sticky="e", **pad)
        self.cmb_sheet = ctk.CTkComboBox(input_tab, values=[], variable=self.var_sheet, width=240)
        self.cmb_sheet.grid(row=1, column=1, sticky="w", **pad)
        self.btn_load_sheets = ctk.CTkButton(input_tab, text="Load Sheets", command=self.load_sheets)
        self.btn_load_sheets.grid(row=1, column=2, **pad)

        ctk.CTkLabel(input_tab, text="Key Column (header or index)").grid(row=2, column=0, sticky="e", **pad)
        self.cmb_key = ctk.CTkComboBox(input_tab, values=[], variable=self.var_keycol, width=240)
        self.cmb_key.grid(row=2, column=1, sticky="w", **pad)
        self.btn_load_headers = ctk.CTkButton(input_tab, text="Load Headers", command=self.load_headers)
        self.btn_load_headers.grid(row=2, column=2, **pad)

        input_tab.grid_columnconfigure(1, weight=1)

        # TEMPLATE TAB
        template_tab = self.tabview.tab("Template")
        ctk.CTkLabel(template_tab, text="Template Excel").grid(row=0, column=0, sticky="e", **pad)
        self.entry_template = ctk.CTkEntry(template_tab, textvariable=self.var_template, width=520)
        self.entry_template.grid(row=0, column=1, sticky="we", **pad)
        self.btn_browse_template = ctk.CTkButton(template_tab, text="Browse...", command=self.browse_template)
        self.btn_browse_template.grid(row=0, column=2, **pad)

        ctk.CTkLabel(template_tab, text="HEADER_ROWS").grid(row=1, column=0, sticky="e", **pad)
        header_entry = ctk.CTkEntry(template_tab, textvariable=self.var_header_rows, width=80)
        header_entry.grid(row=1, column=1, sticky="w", **pad)

        template_tab.grid_columnconfigure(1, weight=1)

        # OUTPUT TAB
        output_tab = self.tabview.tab("Output")
        ctk.CTkLabel(output_tab, text="Output Folder").grid(row=0, column=0, sticky="e", **pad)
        self.entry_outdir = ctk.CTkEntry(output_tab, textvariable=self.var_outdir, width=520)
        self.entry_outdir.grid(row=0, column=1, sticky="we", **pad)
        self.btn_browse_outdir = ctk.CTkButton(output_tab, text="Browse...", command=self.browse_outdir)
        self.btn_browse_outdir.grid(row=0, column=2, **pad)

        ctk.CTkLabel(output_tab, text="PDF Engine").grid(row=1, column=0, sticky="e", **pad)
        pdf_combo = ctk.CTkComboBox(output_tab, values=["reportlab", "libreoffice", "none"], variable=self.var_pdf_engine, width=200)
        pdf_combo.grid(row=1, column=1, sticky="w", **pad)

        ctk.CTkLabel(output_tab, text="LibreOffice (soffice.exe)").grid(row=2, column=0, sticky="e", **pad)
        self.entry_lo_path = ctk.CTkEntry(output_tab, textvariable=self.var_lo_path, width=520)
        self.entry_lo_path.grid(row=2, column=1, sticky="we", **pad)
        self.btn_browse_soffice = ctk.CTkButton(output_tab, text="Browse...", command=self.browse_soffice)
        self.btn_browse_soffice.grid(row=2, column=2, **pad)

        ctk.CTkLabel(output_tab, text="Prefix").grid(row=3, column=0, sticky="e", **pad)
        prefix_entry = ctk.CTkEntry(output_tab, textvariable=self.var_prefix, width=240)
        prefix_entry.grid(row=3, column=1, sticky="w", **pad)

        ctk.CTkLabel(output_tab, text="Suffix").grid(row=4, column=0, sticky="e", **pad)
        suffix_entry = ctk.CTkEntry(output_tab, textvariable=self.var_suffix, width=240)
        suffix_entry.grid(row=4, column=1, sticky="w", **pad)

        output_tab.grid_columnconfigure(1, weight=1)

        # ACTIONS TAB
        actions_tab = self.tabview.tab("Actions")

        # Configuration Summary
        summary_frame = ctk.CTkFrame(actions_tab)
        summary_frame.pack(fill="x", padx=12, pady=(10, 10))

        ctk.CTkLabel(summary_frame, text="Configuration Summary", font=ctk.CTkFont(size=14, weight="bold")).pack(pady=(10, 5))

        # Create summary labels
        self.summary_labels = {}
        config_items = [
            ("Source File", self.var_source),
            ("Sheet Name", self.var_sheet),
            ("Key Column", self.var_keycol),
            ("Template File", self.var_template),
            ("Header Rows", self.var_header_rows),
            ("Output Folder", self.var_outdir),
            ("PDF Engine", self.var_pdf_engine),
            ("LibreOffice Path", self.var_lo_path),
            ("Prefix", self.var_prefix),
            ("Suffix", self.var_suffix),
        ]

        for label_text, var in config_items:
            row_frame = ctk.CTkFrame(summary_frame, fg_color="transparent")
            row_frame.pack(fill="x", padx=10, pady=0)
            ctk.CTkLabel(row_frame, text=f"{label_text}:", width=120, anchor="w").pack(side="left")
            value_label = ctk.CTkLabel(row_frame, textvariable=var, anchor="w")
            value_label.pack(side="left", fill="x", expand=True)
            self.summary_labels[label_text] = value_label

        # Buttons
        btns = ctk.CTkFrame(actions_tab)
        btns.pack(fill="x", padx=12, pady=(10, 4))
        self.btn_run = ctk.CTkButton(btns, text="Generate", height=44, command=self.on_run_clicked)
        self.btn_run.pack(side="left", padx=(0, 8))

        # PROGRESS + LOG
        self.pbar = ctk.CTkProgressBar(actions_tab, mode="determinate", width=640)
        self.pbar.pack(fill="x", padx=12, pady=(14, 4))
        self.pbar.set(0.0)
        self.txt_status = ctk.CTkTextbox(actions_tab, height=300)
        self.txt_status.pack(fill="both", expand=True, padx=12, pady=(10, 14))

    # ------------- UI Helpers -------------

    def log(self, msg: str):
        self.after(0, self._append_log, msg)

    def _append_log(self, msg: str):
        self.txt_status.insert("end", msg + "\n")
        self.txt_status.see("end")

    def set_progress(self, total: int, current: int):
        self.after(0, self._set_progress_impl, total, current)

    def _set_progress_impl(self, total: int, current: int):
        ratio = 0.0 if total <= 0 else max(0.0, min(1.0, current / total))
        self.pbar.set(ratio)

    def set_busy(self, busy: bool):
        self.after(0, self._set_busy_impl, busy)

    def _set_busy_impl(self, busy: bool):
        self.is_running = busy
        state = "disabled" if busy else "normal"
        try:
            self.btn_run.configure(state=state, text="Generating..." if busy else "Generate")
            self.btn_save_top.configure(state=state)
            self.btn_load_top.configure(state=state)
        except AttributeError:
            pass  # UI being recreated
        self.configure(cursor="watch" if busy else "")
        if not busy:
            self.pbar.set(0.0)

    # ------------- Browse & Load -------------

    def browse_source(self):
        f = filedialog.askopenfilename(
            title="Pilih source Excel",
            filetypes=[("Excel files", "*.xlsx;*.xls;*.xlsm;*.xlsb")]
        )
        if f:
            self.var_source.set(f)
            self.log(f"Source: {f}")

    def browse_template(self):
        f = filedialog.askopenfilename(
            title="Pilih template Excel",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if f:
            self.var_template.set(f)
            self.log(f"Template: {f}")

    def browse_outdir(self):
        d = filedialog.askdirectory(title="Pilih output folder")
        if d:
            self.var_outdir.set(d)
            self.log(f"Output: {d}")

    def browse_soffice(self):
        f = filedialog.askopenfilename(
            title="Pilih soffice.exe (LibreOffice)",
            filetypes=[("Executable", "soffice.exe"), ("All files", "*.*")]
        )
        if f:
            self.var_lo_path.set(f)
            self.log(f"LibreOffice: {f}")

    def load_sheets(self):
        src = self.var_source.get().strip()
        if not src:
            messagebox.showwarning("Perhatian", "Pilih source Excel dulu.")
            return
        try:
            xls = pd.ExcelFile(src)
            sheets = xls.sheet_names
            self.cmb_sheet.configure(values=sheets)
            if sheets:
                self.var_sheet.set(sheets[0])
            self.log(f"Sheets loaded: {', '.join(sheets)}")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def load_headers(self):
        src = self.var_source.get().strip()
        sheet = self.var_sheet.get().strip()
        if not src or not sheet:
            messagebox.showwarning("Perhatian", "Pastikan source & sheet sudah dipilih.")
            return
        try:
            df = pd.read_excel(src, sheet_name=sheet, nrows=0)
            headers = list(df.columns.astype(str))
            index_vals = [str(i+1) for i in range(len(headers))]
            values = headers + index_vals
            self.cmb_key.configure(values=values)
            if headers:
                self.var_keycol.set(headers[0])
            self.log(f"Headers loaded: {headers}")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    # ------------- Run Generate (Threaded) -------------

    def on_run_clicked(self):
        if self.is_running:
            return
        try:
            source_path = Path(self.var_source.get().strip())
            template_path = Path(self.var_template.get().strip())
            out_dir = Path(self.var_outdir.get().strip())
            sheet_name = self.var_sheet.get().strip()
            key_raw = self.var_keycol.get().strip()
            header_rows = int(self.var_header_rows.get())
            pdf_engine = self.var_pdf_engine.get().strip().lower()

            if not source_path.exists():
                messagebox.showerror("Error", "Source Excel tidak ditemukan.")
                return
            if not template_path.exists():
                messagebox.showerror("Error", "Template Excel tidak ditemukan.")
                return
            if not out_dir:
                messagebox.showerror("Error", "Output folder belum dipilih.")
                return
            if not sheet_name:
                messagebox.showerror("Error", "Sheet belum dipilih.")
                return
            if not key_raw:
                messagebox.showerror("Error", "Key Column belum dipilih/diisi.")
                return

            try:
                key_col = int(key_raw)
            except ValueError:
                key_col = key_raw  # pakai header name

            # Validasi ReportLab bila dipilih
            if pdf_engine == "reportlab" and not REPORTLAB_AVAILABLE:
                messagebox.showwarning(
                    "ReportLab tidak tersedia",
                    "ReportLab belum terpasang di environment ini.\n"
                    "Jalankan: pip install reportlab\n"
                    "Atau pilih PDF Engine: 'libreoffice' atau 'none'."
                )
                return

            # Deteksi LibreOffice jika dipilih
            soffice_path = None
            if pdf_engine == "libreoffice":
                lo_explicit = self.var_lo_path.get().strip()
                soffice_path = find_soffice(lo_explicit)
                if not soffice_path:
                    ans = messagebox.askyesno(
                        "LibreOffice tidak ditemukan",
                        "Tidak menemukan 'soffice'. Mau pilih lokasi secara manual?"
                    )
                    if ans:
                        self.browse_soffice()
                        lo_explicit = self.var_lo_path.get().strip()
                        soffice_path = find_soffice(lo_explicit)
                if not soffice_path:
                    messagebox.showerror(
                        "Error",
                        "LibreOffice (soffice.exe) tidak ditemukan.\n"
                        "Isi path LibreOffice atau pilih PDF Engine: 'reportlab' / 'none'."
                    )
                    return

            self.set_busy(True)
            self.log("Mulai generate...")

            def worker():
                try:
                    split_excel_with_template(
                        source_path=source_path,
                        sheet_name=sheet_name,
                        key_col=key_col,
                        template_path=template_path,
                        out_dir=out_dir,
                        header_rows=header_rows,
                        pdf_engine=pdf_engine,
                        soffice_path=soffice_path,
                        prefix=self.var_prefix.get().strip(),
                        suffix=self.var_suffix.get().strip(),
                        status_cb=self.log,
                        progress_cb=self.set_progress
                    )
                    self.log("Selesai.")
                    self.after(0, lambda: messagebox.showinfo("Selesai", "Proses selesai."))
                except subprocess.CalledProcessError as e:
                    try:
                        out = e.stdout.decode("utf-8", errors="ignore")
                        err = e.stderr.decode("utf-8", errors="ignore")
                    except Exception:
                        out, err = str(e), ""
                    self.log(out); self.log(err)
                    self.after(0, lambda: messagebox.showerror("LibreOffice Error", "Gagal export PDF. Lihat log."))
                except Exception as e:
                    error_msg = str(e)
                    self.after(0, lambda: messagebox.showerror("Error", error_msg))
                finally:
                    self.set_busy(False)

            self.worker_thread = threading.Thread(target=worker, daemon=True)
            self.worker_thread.start()

        except Exception as e:
            messagebox.showerror("Error", str(e))
            self.set_busy(False)

    # ------------- INI Save / Load -------------

    def save_ini(self):
        f = filedialog.asksaveasfilename(
            defaultextension=".ini",
            filetypes=[("INI files", "*.ini")],
            title="Simpan konfigurasi"
        )
        if not f:
            return
        cfg = configparser.ConfigParser()
        cfg["template"] = {
            "template_path": self.var_template.get().strip(),
            "header_rows": str(self.var_header_rows.get()),
        }
        cfg["source"] = {
            "source_path": self.var_source.get().strip(),
            "sheet_name": self.var_sheet.get().strip(),
            "key_col": self.var_keycol.get().strip()
        }
        cfg["output"] = {
            "output_dir": self.var_outdir.get().strip(),
            "pdf_engine": self.var_pdf_engine.get().strip().lower(),
            "libreoffice_path": self.var_lo_path.get().strip(),
            "prefix": self.var_prefix.get().strip(),
            "suffix": self.var_suffix.get().strip()
        }
        with open(f, "w", encoding="utf-8") as fp:
            cfg.write(fp)
        self.log(f"Konfigurasi tersimpan: {f}")

    def load_ini(self):
        f = filedialog.askopenfilename(
            title="Muat konfigurasi",
            filetypes=[("INI files", "*.ini")]
        )
        if not f:
            return
        cfg = configparser.ConfigParser()
        cfg.read(f, encoding="utf-8")

        try:
            self.var_template.set(cfg.get("template", "template_path", fallback=""))
            self.var_header_rows.set(cfg.getint("template", "header_rows", fallback=5))

            self.var_source.set(cfg.get("source", "source_path", fallback=""))
            self.var_sheet.set(cfg.get("source", "sheet_name", fallback=""))
            self.var_keycol.set(cfg.get("source", "key_col", fallback=""))

            self.var_outdir.set(cfg.get("output", "output_dir", fallback=""))
            self.var_pdf_engine.set(cfg.get("output", "pdf_engine", fallback="reportlab").lower())
            self.var_lo_path.set(cfg.get("output", "libreoffice_path", fallback=""))
            self.var_prefix.set(cfg.get("output", "prefix", fallback=""))
            self.var_suffix.set(cfg.get("output", "suffix", fallback=""))

            self.var_loaded_ini.set(f"Loaded: {Path(f).name}")
            self.log(f"Konfigurasi dimuat: {f}")
        except Exception as e:
            messagebox.showerror("Error", f"Format .ini tidak valid: {e}")


if __name__ == "__main__":
    app = SplitApp()
    app.mainloop()
