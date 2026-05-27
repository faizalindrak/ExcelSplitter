# split_gui.py
# GUI untuk split Excel per nilai unik dengan template-based rendering.
# Build exe: pyinstaller split_gui.spec
# Dependencies: PySide6, PySide6-Fluent-Widgets, pandas, openpyxl

import os
import re
import shutil
import subprocess
import sys
import json
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from PySide6.QtCore import Qt, Signal, QThread, QSettings
from PySide6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout,
    QFileDialog, QGridLayout, QSizePolicy
)
from qfluentwidgets import (
    ScrollArea, SimpleCardWidget,
    LineEdit, ComboBox, PushButton, PrimaryPushButton,
    ProgressBar, SpinBox, TextEdit, InfoBar, InfoBarPosition,
    ToolButton, SubtitleLabel, BodyLabel, CaptionLabel,
    isDarkTheme, qconfig, setTheme, Theme, FluentIcon as FIF
)

from mail_merge import (
    all_jobs_valid,
    AttachmentSelection,
    EmailJob,
    EmailTemplate,
    SplitResult,
    build_email_jobs,
    load_recipient_rows,
    read_recipient_headers,
)

TEMPLATE_MODE_TEMPLATE_FILE = "template_file"
TEMPLATE_MODE_SOURCE_TEMPLATE = "source_template"
TEMPLATE_MODE_LABELS = {
    TEMPLATE_MODE_TEMPLATE_FILE: "Use Template File",
    TEMPLATE_MODE_SOURCE_TEMPLATE: "Use Source as Template",
}
TEMPLATE_MODE_BY_LABEL = {label: key for key, label in TEMPLATE_MODE_LABELS.items()}
OUTPUT_TYPE_EXCEL = "excel"
OUTPUT_TYPE_PDF = "pdf"
OUTPUT_TYPE_EXCEL_AND_PDF = "excel_and_pdf"
OUTPUT_TYPE_LABELS = {
    OUTPUT_TYPE_EXCEL: "Excel",
    OUTPUT_TYPE_PDF: "PDF",
    OUTPUT_TYPE_EXCEL_AND_PDF: "Excel + PDF",
}
OUTPUT_TYPE_BY_LABEL = {label: key for key, label in OUTPUT_TYPE_LABELS.items()}
PATH_FIELD_WIDTH = 460
SECONDARY_PATH_FIELD_WIDTH = 320
COMBO_FIELD_WIDTH = 190
SMALL_FIELD_WIDTH = 120
NAME_FIELD_WIDTH = 160
FIELD_CONTROL_HEIGHT = 40
DASHBOARD_LIGHT_PALETTE = {
    "root": "#f5f7fb",
    "card": "#ffffff",
    "border": "#d8dee8",
    "text": "#111827",
}
DASHBOARD_DARK_PALETTE = {
    "root": "#202020",
    "card": "#2b2b2b",
    "border": "#3f3f46",
    "text": "#f5f5f5",
}

# ==== (Opsional) xlwings untuk PDF via Excel COM ====
try:
    import xlwings as xw
    XLWINGS_AVAILABLE = True
except Exception:
    XLWINGS_AVAILABLE = False


# ----------------- Helpers -----------------

def safe_file_part(s: str) -> str:
    s = "" if s is None else str(s)
    return re.sub(r'[:\\/\?\*\[\]<>|"]', "_", s).strip() or "Key"

def join_output_name_parts(prefix: str, key_part: str, suffix: str) -> str:
    parts = []
    if prefix:
        parts.append(prefix)
    parts.append(key_part)
    if suffix:
        parts.append(suffix)
    return " ".join(parts)

def build_output_stem(prefix: str, key_value, suffix: str) -> str:
    return join_output_name_parts(prefix, safe_file_part(key_value), suffix)

def output_extension(output_file_type: str) -> str:
    if output_file_type == OUTPUT_TYPE_PDF:
        return ".pdf"
    if output_file_type == OUTPUT_TYPE_EXCEL_AND_PDF:
        return ".xlsx + PDF"
    return ".xlsx"

def output_requires_pdf(output_file_type: str) -> bool:
    return output_file_type in {OUTPUT_TYPE_PDF, OUTPUT_TYPE_EXCEL_AND_PDF}

def remove_intermediate_workbook_for_pdf(xlsx_path: Path, output_file_type: str):
    if output_file_type != OUTPUT_TYPE_PDF:
        return
    if xlsx_path.with_suffix(".pdf").exists() and xlsx_path.exists():
        xlsx_path.unlink()

def set_print_titles_and_area(ws, header_rows: int, last_col_idx: int, last_data_row: int):
    ws.print_title_rows = f"1:{header_rows}"
    last_col_letter = get_column_letter(last_col_idx if last_col_idx > 0 else 1)
    last_row = last_data_row if last_data_row >= (header_rows + 1) else (header_rows + 1)
    ws.print_area = f"A1:{last_col_letter}{last_row}"

def normalize_header(value) -> str:
    value = "" if value is None else str(value)
    return re.sub(r"[^a-z0-9]+", "", value.strip().lower())

def auto_map_columns(template_headers, source_headers):
    source_by_key = {}
    for source in source_headers:
        key = normalize_header(source)
        if key and key not in source_by_key:
            source_by_key[key] = source

    return {
        template: source_by_key.get(normalize_header(template))
        for template in template_headers
    }

def validate_column_mapping(template_headers, mapping):
    mapping = mapping or {}
    return [
        template
        for template in template_headers
        if not mapping.get(template)
    ]

def read_excel_headers(path: Path, sheet_name: str, header_rows: int) -> list[str]:
    df = pd.read_excel(path, sheet_name=sheet_name, header=header_rows - 1, nrows=0)
    return [str(col) for col in df.columns]

def read_template_headers(path: Path, header_rows: int) -> tuple[list[str], int]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        headers, col_idx, empty_streak = [], 1, 0
        first_col = None
        while col_idx <= 500 and empty_streak < 5:
            value = ws.cell(row=header_rows, column=col_idx).value
            if value is None or str(value).strip() == "":
                empty_streak += 1
            else:
                if first_col is None:
                    first_col = col_idx
                headers.append(str(value).strip())
                empty_streak = 0
            col_idx += 1
        return headers, first_col or 1
    finally:
        wb.close()

def detect_excel_header_row(path: Path, sheet_name: str | None = None, max_rows: int = 20) -> int:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name] if sheet_name else wb.active
        best_row, best_score = 1, -1
        scan_limit = min(max_rows, ws.max_row)
        rows = list(ws.iter_rows(min_row=1, max_row=scan_limit, values_only=True))
        for offset, row in enumerate(rows):
            row_number = offset + 1
            values = [str(value).strip() for value in row if value is not None and str(value).strip()]
            if not values:
                continue

            text_count = sum(1 for value in values if re.search(r"[A-Za-z]", value))
            numeric_count = len(values) - text_count
            unique_count = len({normalize_header(value) or value.lower() for value in values})
            next_values = []
            if offset + 1 < len(rows):
                next_values = [
                    str(value).strip()
                    for value in rows[offset + 1]
                    if value is not None and str(value).strip()
                ]

            score = (len(values) * 4) + (text_count * 2) + unique_count
            if next_values:
                score += min(len(next_values), len(values))
            if len(values) == 1:
                score -= 4
            if numeric_count > text_count:
                score -= 3

            if score > best_score:
                best_row, best_score = row_number, score

        return best_row
    finally:
        wb.close()

def find_soffice(explicit_path: str | None = None) -> str | None:
    """Cari soffice.exe dari explicit, env, PATH, atau lokasi umum."""
    def _normalize(p):
        if not p:
            return None
        p = str(Path(p))
        if p.lower().endswith("soffice.exe") and Path(p).exists():
            return p
        prog = Path(p) / "soffice.exe"
        if prog.exists():
            return str(prog)
        prog = Path(p) / "program" / "soffice.exe"
        if prog.exists():
            return str(prog)
        return None

    # explicit
    s = _normalize(explicit_path)
    if s: return s
    # env
    s = _normalize(os.environ.get("LIBREOFFICE_PATH"))
    if s: return s
    # PATH
    s = shutil.which("soffice")
    if s: return s
    # lokasi umum
    common = [
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
        r"C:\PortableApps\LibreOfficePortable\App\libreoffice\program\soffice.exe",
        r"D:\PortableApps\LibreOfficePortable\App\libreoffice\program\soffice.exe",
    ]
    for c in common:
        if Path(c).exists():
            return c
    return None

def export_pdf_via_lo(xlsx_path: Path, soffice_path: str | None = None):
    exe = soffice_path or "soffice"
    cmd = [exe, "--headless", "--convert-to", "pdf", "--outdir", str(xlsx_path.parent), str(xlsx_path)]
    subprocess.run(cmd, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)

def cleanup_excel_com():
    """Clean up Excel COM objects and release resources"""
    try:
        # Try to release any existing COM objects
        import gc
        import pythoncom

        # Force garbage collection
        gc.collect()

        # Uninitialize COM
        try:
            pythoncom.CoUninitialize()
        except:
            pass

        # Re-initialize COM for next use
        try:
            pythoncom.CoInitialize()
        except:
            pass

    except ImportError:
        # pythoncom not available, try basic cleanup
        import gc
        gc.collect()
    except:
        pass

def debug_excel_detection():
    """Debug function to check Excel detection methods"""
    results = []

    results.append(f"XLWINGS_AVAILABLE: {XLWINGS_AVAILABLE}")

    # Test Method 1: xlwings
    if XLWINGS_AVAILABLE:
        try:
            import xlwings as xw
            app = xw.App(visible=False, add_book=False)
            if app is not None:
                app.quit()
                results.append("Method 1 (xlwings): SUCCESS")
            else:
                results.append("Method 1 (xlwings): FAILED - app is None")
        except Exception as e:
            results.append(f"Method 1 (xlwings): FAILED - {str(e)}")
    else:
        results.append("Method 1 (xlwings): SKIPPED - xlwings not available")

    # Test Method 2: win32com.client.Dispatch
    try:
        import win32com.client
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.Quit()
        del excel
        results.append("Method 2 (win32com Dispatch): SUCCESS")
    except Exception as e:
        results.append(f"Method 2 (win32com Dispatch): FAILED - {str(e)}")

    # Test Method 3: win32com.client.gencache.EnsureDispatch
    try:
        import win32com.client
        excel = win32com.client.gencache.EnsureDispatch("Excel.Application")
        excel.Visible = False
        excel.Quit()
        del excel
        results.append("Method 3 (win32com EnsureDispatch): SUCCESS")
    except Exception as e:
        results.append(f"Method 3 (win32com EnsureDispatch): FAILED - {str(e)}")

    # Test Method 4: Registry check
    try:
        import winreg
        found_keys = []
        key_paths = [
            r"SOFTWARE\Microsoft\Office\ClickToRun\Configuration",
            r"SOFTWARE\Microsoft\Office\16.0\Excel\InstallRoot",
            r"SOFTWARE\Microsoft\Office\15.0\Excel\InstallRoot",
            r"SOFTWARE\Microsoft\Office\14.0\Excel\InstallRoot",
            r"SOFTWARE\WOW6432Node\Microsoft\Office\16.0\Excel\InstallRoot",
            r"SOFTWARE\WOW6432Node\Microsoft\Office\15.0\Excel\InstallRoot",
            r"SOFTWARE\WOW6432Node\Microsoft\Office\14.0\Excel\InstallRoot",
        ]

        for key_path in key_paths:
            try:
                key = winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE, key_path)
                winreg.CloseKey(key)
                found_keys.append(key_path)
            except:
                continue

        if found_keys:
            results.append(f"Method 4 (Registry): Found {len(found_keys)} keys: {found_keys}")
        else:
            results.append("Method 4 (Registry): No Excel keys found")

    except ImportError:
        results.append("Method 4 (Registry): SKIPPED - winreg not available")
    except Exception as e:
        results.append(f"Method 4 (Registry): FAILED - {str(e)}")

    return results

def check_excel_availability():
    """Check if Microsoft Excel is available for COM automation"""
    try:
        # Method 1: Try xlwings first (most reliable for xlwings usage)
        if XLWINGS_AVAILABLE:
            try:
                import xlwings as xw
                app = xw.App(visible=False, add_book=False)
                if app is not None:
                    app.quit()
                    return True
            except Exception:
                pass

        # Method 2: Try win32com.client
        try:
            import win32com.client
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.Quit()
            del excel
            return True
        except Exception:
            pass

        # Method 3: Try alternative COM approach
        try:
            import win32com.client
            excel = win32com.client.gencache.EnsureDispatch("Excel.Application")
            excel.Visible = False
            excel.Quit()
            del excel
            return True
        except Exception:
            pass

        # Method 4: Check registry for Excel installation
        try:
            import winreg
            key_paths = [
                r"SOFTWARE\Microsoft\Office\ClickToRun\Configuration",
                r"SOFTWARE\Microsoft\Office\16.0\Excel\InstallRoot",
                r"SOFTWARE\Microsoft\Office\15.0\Excel\InstallRoot",
                r"SOFTWARE\Microsoft\Office\14.0\Excel\InstallRoot",
                r"SOFTWARE\WOW6432Node\Microsoft\Office\16.0\Excel\InstallRoot",
                r"SOFTWARE\WOW6432Node\Microsoft\Office\15.0\Excel\InstallRoot",
                r"SOFTWARE\WOW6432Node\Microsoft\Office\14.0\Excel\InstallRoot",
            ]

            for key_path in key_paths:
                try:
                    key = winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE, key_path)
                    winreg.CloseKey(key)
                    # Found Excel in registry, but still need to test COM
                    try:
                        import win32com.client
                        excel = win32com.client.Dispatch("Excel.Application")
                        excel.Visible = False
                        excel.Quit()
                        return True
                    except:
                        pass
                except:
                    continue
        except ImportError:
            pass  # winreg not available
        except Exception:
            pass

        return False

    except Exception:
        return False
    finally:
        # Clean up after test
        try:
            cleanup_excel_com()
        except:
            pass

def export_pdf_via_xlwings(xlsx_path: Path):
    """Export Excel to PDF using xlwings (requires Excel installed on Windows)"""
    if not XLWINGS_AVAILABLE:
        raise RuntimeError("xlwings belum terpasang. Jalankan: pip install xlwings")

    # Clean up COM objects only (safe cleanup)
    cleanup_excel_com()

    # Check if Excel is available before proceeding
    if not check_excel_availability():
        raise RuntimeError("Microsoft Excel tidak terinstall atau tidak dapat diakses. "
                          "Gunakan PDF Engine 'libreoffice' atau 'none' sebagai alternatif.")

    pdf_path = xlsx_path.with_suffix(".pdf")

    # Use xlwings with invisible Excel application
    app = None
    wb = None
    try:
        # Create Excel application instance (invisible)
        app = xw.App(visible=False, add_book=False)

        # Verify app was created successfully
        if app is None:
            raise RuntimeError("Failed to create Excel application instance")

        # Open the workbook
        wb = app.books.open(str(xlsx_path))

        if wb is None:
            raise RuntimeError(f"Failed to open workbook: {xlsx_path}")

        # Get the active worksheet
        ws = wb.sheets.active

        if ws is None:
            raise RuntimeError("Failed to get active worksheet")

        # Export to PDF with Excel's native formatting
        # This preserves all Excel formatting, styles, colors, etc.
        ws.to_pdf(str(pdf_path))

        # Verify PDF was created
        if not pdf_path.exists():
            raise RuntimeError(f"PDF was not created: {pdf_path}")

    except Exception as e:
        # More detailed error reporting
        error_details = str(e)
        if "apps" in error_details.lower() or "nonetype" in error_details.lower():
            error_details = ("Excel COM interface tidak dapat diakses. "
                           "Pastikan Microsoft Excel terinstall dan tidak sedang digunakan aplikasi lain. "
                           "Alternatif: gunakan PDF Engine 'libreoffice' atau 'none'.")
        raise RuntimeError(f"Gagal export PDF via xlwings: {error_details}")
    finally:
        # Clean up: close workbook and quit Excel
        if wb is not None:
            try:
                wb.close()
            except:
                pass
        if app is not None:
            try:
                app.quit()
            except:
                pass

        # Safe COM cleanup only after each PDF export
        cleanup_excel_com()

        # Give a moment for cleanup to complete
        import time
        time.sleep(0.5)


# ----------------- Split Logic -----------------

def split_excel_with_template(
    source_path: Path, sheet_name: str, key_col, template_path: Path, out_dir: Path,
    header_rows: int, pdf_engine: str = "xlwings", soffice_path: str | None = None,
    prefix: str = "", suffix: str = "", status_cb=None, progress_cb=None,
    template_mode: str = TEMPLATE_MODE_TEMPLATE_FILE, column_mapping: dict | None = None,
    source_header_rows: int | None = None, template_header_rows: int | None = None,
    output_file_type: str | None = None
):
    if status_cb is None: status_cb = lambda msg: None
    if progress_cb is None: progress_cb = lambda t, c: None
    source_header_rows = source_header_rows or header_rows
    template_header_rows = template_header_rows or header_rows
    if output_file_type is None:
        output_file_type = (
            OUTPUT_TYPE_EXCEL
            if (pdf_engine or "none").lower() == "none"
            else OUTPUT_TYPE_PDF
        )
    if output_file_type not in {OUTPUT_TYPE_EXCEL, OUTPUT_TYPE_PDF, OUTPUT_TYPE_EXCEL_AND_PDF}:
        raise ValueError(f"Tipe output tidak didukung: {output_file_type}")
    effective_pdf_engine = "none" if not output_requires_pdf(output_file_type) else (pdf_engine or "xlwings")
    if output_requires_pdf(output_file_type) and effective_pdf_engine == "none":
        effective_pdf_engine = "xlwings"

    if template_mode not in {TEMPLATE_MODE_TEMPLATE_FILE, TEMPLATE_MODE_SOURCE_TEMPLATE}:
        raise ValueError(f"Template mode tidak didukung: {template_mode}")
    if not source_path.exists():
        raise FileNotFoundError(f"Sumber tidak ditemukan: {source_path}")
    if template_mode == TEMPLATE_MODE_TEMPLATE_FILE and not template_path.exists():
        raise FileNotFoundError(f"Template tidak ditemukan: {template_path}")
    out_dir.mkdir(parents=True, exist_ok=True)
    split_results: list[SplitResult] = []

    status_cb("Membaca sumber...")

    # Diagnostic logging
    try:
        file_size = source_path.stat().st_size if source_path.exists() else 0
        status_cb(f"Debug: File path: {source_path}")
        status_cb(f"Debug: File exists: {source_path.exists()}")
        status_cb(f"Debug: File size: {file_size:,} bytes ({file_size/1024/1024:.2f} MB)")
        status_cb(f"Debug: Sheet name: '{sheet_name}'")
        status_cb("Debug: Starting pd.read_excel...")

        # Try reading with timeout and error handling
        import time
        start_time = time.time()

        # First try to read just the header to test file accessibility
        status_cb("Debug: Testing file accessibility...")
        try:
            df_test = pd.read_excel(source_path, sheet_name=sheet_name, nrows=5, dtype=object)
            status_cb(f"Debug: Successfully read {len(df_test)} rows for testing")
        except Exception as test_e:
            status_cb(f"Debug: Test read failed: {str(test_e)}")
            raise test_e

        # Read with header at the correct row (source_header_rows is 1-indexed)
        df = pd.read_excel(source_path, sheet_name=sheet_name, header=source_header_rows - 1, dtype=object)

        elapsed = time.time() - start_time
        status_cb(f"Debug: Successfully read {len(df)} rows in {elapsed:.2f} seconds")

    except FileNotFoundError as e:
        status_cb(f"Debug: File not found: {e}")
        raise FileNotFoundError(f"File tidak ditemukan: {source_path}")
    except PermissionError as e:
        status_cb(f"Debug: Permission denied: {e}")
        raise PermissionError(f"Tidak ada akses ke file: {source_path}")
    except Exception as e:
        status_cb(f"Debug: Error reading Excel: {type(e).__name__}: {str(e)}")
        raise e

    # Tentukan kolom kunci
    if isinstance(key_col, int):
        if key_col < 1 or key_col > df.shape[1]:
            raise ValueError("Index kolom kunci di luar jangkauan DataFrame.")
        key_series = df.iloc[:, key_col - 1]
    else:
        if key_col not in df.columns:
            raise ValueError(f"Header kolom kunci '{key_col}' tidak ditemukan.")
        key_series = df[key_col]

    # Debug: Check for categorical data issues
    status_cb(f"Debug: Key column '{key_col}' data type: {key_series.dtype}")
    status_cb(f"Debug: Key column unique values: {len(key_series.unique())}")
    status_cb(f"Debug: Key column has null values: {key_series.isnull().sum()}")

    # Check for categorical columns in the entire DataFrame
    categorical_cols = []
    for col in df.columns:
        if df[col].dtype.name == 'category':
            categorical_cols.append(col)
            status_cb(f"Debug: Found categorical column: {col}")

    if categorical_cols:
        status_cb(f"Debug: Converting {len(categorical_cols)} categorical columns to string...")
        for col in categorical_cols:
            try:
                df[col] = df[col].astype(str)
                status_cb(f"Debug: Converted {col} to string")
            except Exception as conv_e:
                status_cb(f"Debug: Failed to convert {col}: {conv_e}")

    # Selaraskan urutan kolom ke header template.
    templ_cols = None
    templ_col_start = 1
    if template_mode == TEMPLATE_MODE_TEMPLATE_FILE:
        templ_cols, templ_col_start = read_template_headers(template_path, template_header_rows)
        if not templ_cols:
            raise ValueError("Header template tidak ditemukan untuk mapping kolom.")

        source_headers = [str(col) for col in df.columns]
        effective_mapping = column_mapping or auto_map_columns(templ_cols, source_headers)
        missing = validate_column_mapping(templ_cols, effective_mapping)
        if missing:
            raise ValueError("Mapping kolom template belum lengkap: " + ", ".join(missing))

        mapped_data = {}
        for template_col in templ_cols:
            source_col = effective_mapping[template_col]
            if source_col not in df.columns:
                raise ValueError(
                    f"Kolom sumber untuk template '{template_col}' tidak ditemukan: {source_col}"
                )
            mapped_data[template_col] = df[source_col]
        df = pd.DataFrame(mapped_data, index=df.index)

    # Debug: Check groupby operation
    status_cb("Debug: Starting groupby operation...")
    try:
        groups = df.groupby(key_series, dropna=False, sort=False)
        total, current = len(groups), 0
        status_cb(f"Debug: Groupby successful, found {total} groups")
        progress_cb(total, 0)
    except Exception as groupby_e:
        status_cb(f"Debug: Groupby error: {type(groupby_e).__name__}: {str(groupby_e)}")

        # Try multiple approaches to fix the issue
        if "categorical" in str(groupby_e).lower():
            status_cb("Debug: Attempting to fix categorical issue...")

            # Method 1: Try converting all categorical columns to object
            try:
                status_cb("Debug: Method 1 - Converting all categorical columns to object...")
                df_no_cat = df.copy()
                for col in df_no_cat.columns:
                    if df_no_cat[col].dtype.name == 'category':
                        df_no_cat[col] = df_no_cat[col].astype('object')
                groups = df_no_cat.groupby(key_series, dropna=False, sort=False)
                total, current = len(groups), 0
                status_cb(f"Debug: Method 1 successful, found {total} groups")
                progress_cb(total, 0)
                # Update df to use the fixed version
                df = df_no_cat
            except Exception as method1_e:
                status_cb(f"Debug: Method 1 failed: {method1_e}")

                # Method 2: Try using string conversion for groupby
                try:
                    status_cb("Debug: Method 2 - Using string keys for groupby...")
                    string_keys = key_series.astype(str)
                    groups = df.groupby(string_keys, dropna=False, sort=False)
                    total, current = len(groups), 0
                    status_cb(f"Debug: Method 2 successful, found {total} groups")
                    progress_cb(total, 0)
                except Exception as method2_e:
                    status_cb(f"Debug: Method 2 failed: {method2_e}")
                    raise groupby_e
        else:
            raise groupby_e

    for key_val, group in groups:
        current += 1
        status_cb(f"Proses [{current}/{total}] key={key_val}")
        progress_cb(total, current)

        if template_mode == TEMPLATE_MODE_SOURCE_TEMPLATE:
            wb = load_workbook(source_path)
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet sumber '{sheet_name}' tidak ditemukan.")

            for sheet in list(wb.worksheets):
                if sheet.title != sheet_name:
                    wb.remove(sheet)

            ws = wb[sheet_name]
            start_row = source_header_rows + 1
            keep_rows = {int(idx) + start_row for idx in group.index}
            for row_idx in range(ws.max_row, start_row - 1, -1):
                if row_idx not in keep_rows:
                    ws.delete_rows(row_idx)

            set_print_titles_and_area(ws, source_header_rows, ws.max_column, ws.max_row)

            out_name = build_output_stem(prefix, key_val, suffix)
            xlsx_out = out_dir / f"{out_name}.xlsx"
            wb.save(xlsx_out)

            eng = (effective_pdf_engine or "none").lower()
            if eng != "none":
                if eng == "libreoffice":
                    export_pdf_via_lo(xlsx_out, soffice_path=soffice_path)
                elif eng == "xlwings":
                    export_pdf_via_xlwings(xlsx_out)
                remove_intermediate_workbook_for_pdf(xlsx_out, output_file_type)
            pdf_out = xlsx_out.with_suffix(".pdf")
            split_results.append(
                SplitResult(
                    key=str(key_val),
                    excel_path=xlsx_out if xlsx_out.exists() else None,
                    pdf_path=pdf_out if pdf_out.exists() else None,
                    output_file_type=output_file_type,
                )
            )
            continue

        # 1) Tulis XLSX dari template
        wb = load_workbook(template_path)
        ws = wb.active
        start_row = template_header_rows + 1

        # Comprehensive cleanup to eliminate all Excel repair warnings
        try:
            # 1. Remove all named ranges (common cause of repair warnings)
            if hasattr(wb, 'defined_names'):
                wb.defined_names = {}

            # 2. Remove external links
            if hasattr(wb, '_external_links') and wb._external_links:
                wb._external_links.clear()

            # 3. Remove external link relationships
            if hasattr(wb, '_external_link_rels'):
                wb._external_link_rels.clear()

            # 4. Remove all drawings completely (they often cause repair issues)
            if hasattr(ws, '_drawings'):
                ws._drawings = []

            # 5. Clean up worksheet-level named ranges
            if hasattr(ws, 'defined_names'):
                ws.defined_names = []

            # 6. Remove external link parts from the package
            try:
                from openpyxl.packaging.relationship import Relationship
                if hasattr(wb, '_rels'):
                    # Remove external link relationships
                    rels_to_remove = []
                    for rel in wb._rels:
                        if ('externalLink' in str(rel.target) or
                            'externalLinks' in str(rel.target) or
                            'drawing' in str(rel.target).lower()):
                            rels_to_remove.append(rel)
                    for rel in rels_to_remove:
                        wb._rels.remove(rel)
            except:
                pass

            # 7. Clean up any hyperlinks that might reference external data
            for row in ws.iter_rows():
                for cell in row:
                    if cell.hyperlink:
                        cell.hyperlink = None

        except Exception as cleanup_e:
            # If cleanup fails, continue anyway
            pass

        values = group.fillna("").values.tolist()

        # Copy formatting from template row to data rows
        # Use the first data row from template as formatting template
        template_row = start_row  # This should be the first data row in template

        for r_off, row_vals in enumerate(values, start=0):
            row_idx = start_row + r_off

            for c_idx, v in enumerate(row_vals, start=templ_col_start):
                ws.cell(row=row_idx, column=c_idx, value=v)

            if r_off > 0:
                try:
                    for col_idx in range(templ_col_start, templ_col_start + len(row_vals)):
                        template_cell = ws.cell(row=template_row, column=col_idx)
                        current_cell = ws.cell(row=row_idx, column=col_idx)

                        # Copy all formatting properties
                        if template_cell.has_style:
                            current_cell.font = template_cell.font.copy()
                            current_cell.fill = template_cell.fill.copy()
                            current_cell.border = template_cell.border.copy()
                            current_cell.alignment = template_cell.alignment.copy()
                            current_cell.number_format = template_cell.number_format
                            current_cell.protection = template_cell.protection.copy()

                except Exception as format_e:
                    # If formatting copy fails, continue without it
                    pass

        last_data_row = start_row + len(values) - 1
        last_col = templ_col_start + group.shape[1] - 1
        set_print_titles_and_area(ws, template_header_rows, max(1, last_col), last_data_row)

        # Build filename with prefix and suffix
        out_name = build_output_stem(prefix, key_val, suffix)
        xlsx_out = out_dir / f"{out_name}.xlsx"
        wb.save(xlsx_out)

        # 2) PDF (opsional)
        eng = (effective_pdf_engine or "none").lower()
        if eng != "none":
            if eng == "libreoffice":
                export_pdf_via_lo(xlsx_out, soffice_path=soffice_path)
            elif eng == "xlwings":
                export_pdf_via_xlwings(xlsx_out)
            remove_intermediate_workbook_for_pdf(xlsx_out, output_file_type)
        pdf_out = xlsx_out.with_suffix(".pdf")
        split_results.append(
            SplitResult(
                key=str(key_val),
                excel_path=xlsx_out if xlsx_out.exists() else None,
                pdf_path=pdf_out if pdf_out.exists() else None,
                output_file_type=output_file_type,
            )
        )

    status_cb("Selesai.")
    progress_cb(total, total)

    # Final cleanup for Excel COM sessions
    if (effective_pdf_engine or "none").lower() == "xlwings":
        cleanup_excel_com()

    return split_results


# ----------------- GUI -----------------

class SplitWorker(QThread):
    status = Signal(str)
    progress = Signal(int, int)
    finished = Signal()
    error = Signal(str)

    def __init__(self, params):
        super().__init__()
        self.params = params
        self.results = []

    def run(self):
        try:
            self.results = split_excel_with_template(
                source_path=self.params['source_path'],
                sheet_name=self.params['sheet_name'],
                key_col=self.params['key_col'],
                template_path=self.params['template_path'],
                out_dir=self.params['out_dir'],
                header_rows=self.params['header_rows'],
                pdf_engine=self.params['pdf_engine'],
                soffice_path=self.params['soffice_path'],
                prefix=self.params['prefix'],
                suffix=self.params['suffix'],
                template_mode=self.params.get('template_mode', TEMPLATE_MODE_TEMPLATE_FILE),
                column_mapping=self.params.get('column_mapping'),
                source_header_rows=self.params.get('source_header_rows'),
                template_header_rows=self.params.get('template_header_rows'),
                output_file_type=self.params.get('output_file_type'),
                status_cb=self.status.emit,
                progress_cb=lambda t, c: self.progress.emit(t, c)
            )
            self.finished.emit()
        except Exception as e:
            self.error.emit(str(e))


class SplitApp(QWidget):
    def __init__(self, settings=None):
        super().__init__()
        self.setWindowTitle("Excel Splitter")
        self.setObjectName("appRoot")
        self.resize(1000, 750)
        setTheme(Theme.AUTO)
        self._apply_dashboard_styles()
        qconfig.themeChanged.connect(lambda *_: self._apply_dashboard_styles())

        self.is_running = False
        self.worker = None
        self.settings = settings or QSettings("Faizalindrak", "ExcelSplitter")
        self.saved_column_mapping = {}
        self._loading_settings = False
        self.source_headers = []
        self.template_headers = []
        self.template_col_start = 1
        self.mapping_combos = {}
        self.mapping_status_labels = {}
        self.field_action_buttons = []
        self.current_split_results = []
        self.current_mail_jobs = []
        self.current_mail_warnings = []
        self.current_preview_index = 0

        self._build_ui()
        self.load_settings()
        self._connect_settings_signals()

    def _apply_dashboard_styles(self):
        palette = DASHBOARD_DARK_PALETTE if isDarkTheme() else DASHBOARD_LIGHT_PALETTE
        self.setStyleSheet(f"""
            QWidget#appRoot {{
                background-color: {palette["root"]};
                color: {palette["text"]};
            }}
            QWidget#footerBar {{
                background-color: {palette["root"]};
                border-top: 1px solid {palette["border"]};
            }}
            QScrollArea#mainScrollArea {{
                background-color: {palette["root"]};
                border: none;
            }}
            QWidget#mainPanelHost {{
                background-color: {palette["root"]};
            }}
            SimpleCardWidget#dashboardPanel,
            SimpleCardWidget#workflowRail {{
                background-color: {palette["card"]};
                border: 1px solid {palette["border"]};
                border-radius: 8px;
            }}
            QLabel {{
                color: {palette["text"]};
            }}
        """)

    def _fixed_width(self, widget, width):
        widget.setMinimumWidth(width)
        widget.setMaximumWidth(width)
        if isinstance(widget, (LineEdit, ComboBox, SpinBox)):
            widget.setMinimumHeight(FIELD_CONTROL_HEIGHT)
            widget.setMaximumHeight(FIELD_CONTROL_HEIGHT)
        return widget

    def _field_action_button(self, button):
        button.setMinimumHeight(FIELD_CONTROL_HEIGHT)
        button.setMaximumHeight(FIELD_CONTROL_HEIGHT)
        self.field_action_buttons.append(button)
        return button

    def _panel(self, title, icon=None):
        card = SimpleCardWidget()
        card.setObjectName("dashboardPanel")
        card.setBorderRadius(8)
        layout = QVBoxLayout(card)
        layout.setContentsMargins(14, 12, 14, 14)
        layout.setSpacing(10)

        header = QHBoxLayout()
        if icon:
            icon_widget = ToolButton(icon)
            icon_widget.setFixedSize(20, 20)
            icon_widget.setEnabled(False)
            header.addWidget(icon_widget)
        header.addWidget(SubtitleLabel(title))
        header.addStretch()
        layout.addLayout(header)
        return card, layout

    def _labeled(self, label, widget):
        wrap = QWidget()
        layout = QVBoxLayout(wrap)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(4)
        layout.addWidget(CaptionLabel(label))
        layout.addWidget(widget)
        return wrap

    def _build_ui(self):
        root_layout = QVBoxLayout(self)
        root_layout.setContentsMargins(0, 0, 0, 0)
        root_layout.setSpacing(0)

        toolbar = QHBoxLayout()
        toolbar.setContentsMargins(18, 12, 18, 8)
        toolbar.setSpacing(10)
        toolbar.addWidget(SubtitleLabel("Excel Splitter"))
        self.lbl_workflow_status = CaptionLabel("Ready")
        toolbar.addWidget(self.lbl_workflow_status)
        toolbar.addStretch()
        self.btn_reset_settings = PushButton("Reset Settings")
        self.btn_reset_settings.clicked.connect(self.reset_settings)
        toolbar.addWidget(self.btn_reset_settings)
        root_layout.addLayout(toolbar)

        body = QHBoxLayout()
        body.setContentsMargins(16, 0, 16, 10)
        body.setSpacing(14)
        self.workflow_rail = self._build_workflow_rail()
        body.addWidget(self.workflow_rail)

        self.scroll_area = ScrollArea()
        self.scroll_area.setObjectName("mainScrollArea")
        self.scroll_area.setWidgetResizable(True)
        self.scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        scroll_widget = QWidget()
        scroll_widget.setObjectName("mainPanelHost")
        self.main_panel_layout = QVBoxLayout(scroll_widget)
        self.main_panel_layout.setContentsMargins(0, 0, 0, 0)
        self.main_panel_layout.setSpacing(10)
        self.scroll_area.setWidget(scroll_widget)
        body.addWidget(self.scroll_area, 1)
        root_layout.addLayout(body, 1)

        self._build_source_card()
        self._build_template_card()
        self._build_mapping_card()
        self._build_output_card()
        self._build_log_card()
        self._build_mail_merge_card()

        self.main_panel_layout.addStretch()

        self.footer_bar = QWidget()
        self.footer_bar.setObjectName("footerBar")
        footer = QHBoxLayout(self.footer_bar)
        footer.setContentsMargins(16, 10, 16, 14)
        footer.setSpacing(10)
        self._build_actions_card(footer)
        root_layout.addWidget(self.footer_bar)

    def _build_workflow_rail(self):
        rail = SimpleCardWidget()
        rail.setObjectName("workflowRail")
        rail.setBorderRadius(8)
        rail.setFixedWidth(148)
        layout = QVBoxLayout(rail)
        layout.setContentsMargins(12, 14, 12, 14)
        layout.setSpacing(10)

        layout.addWidget(CaptionLabel("Workflow"))
        self.workflow_steps = {}
        for name in ["Source", "Template", "Output", "Run"]:
            row = QWidget()
            row_layout = QHBoxLayout(row)
            row_layout.setContentsMargins(0, 0, 0, 0)
            row_layout.setSpacing(8)
            dot = CaptionLabel("○")
            label = BodyLabel(name)
            row_layout.addWidget(dot)
            row_layout.addWidget(label)
            row_layout.addStretch()
            layout.addWidget(row)
            self.workflow_steps[name] = dot

        layout.addStretch()
        return rail

    def _build_source_card(self):
        card, layout = self._panel("Source", FIF.DOCUMENT)

        row1 = QHBoxLayout()
        self.edit_source = self._fixed_width(LineEdit(), PATH_FIELD_WIDTH)
        self.edit_source.setPlaceholderText("Source Excel file")
        self.btn_browse_source = self._field_action_button(ToolButton(FIF.FOLDER))
        self.btn_browse_source.clicked.connect(self.browse_source)
        row1.addWidget(self._labeled("Workbook", self.edit_source))
        row1.addWidget(self.btn_browse_source, 0, Qt.AlignBottom)
        row1.addStretch()
        layout.addLayout(row1)

        grid = QGridLayout()
        grid.setHorizontalSpacing(10)
        grid.setVerticalSpacing(8)
        self.cmb_sheet = self._fixed_width(ComboBox(), COMBO_FIELD_WIDTH)
        self.cmb_sheet.setPlaceholderText("Sheet")
        self.btn_load_sheets = self._field_action_button(ToolButton(FIF.SYNC))
        self.btn_load_sheets.setToolTip("Load Sheets")
        self.btn_load_sheets.clicked.connect(self.load_sheets)
        self.cmb_key = self._fixed_width(ComboBox(), COMBO_FIELD_WIDTH)
        self.cmb_key.setPlaceholderText("Key Column")
        self.btn_load_headers = self._field_action_button(ToolButton(FIF.SYNC))
        self.btn_load_headers.setToolTip("Load Headers")
        self.btn_load_headers.clicked.connect(self.load_headers)
        self.spin_source_header_rows = self._fixed_width(SpinBox(), SMALL_FIELD_WIDTH)
        self.spin_source_header_rows.setRange(1, 100)
        self.spin_source_header_rows.setValue(5)
        self.spin_header_rows = self.spin_source_header_rows
        self.btn_detect_source_header = self._field_action_button(PushButton("Detect"))
        self.btn_detect_source_header.clicked.connect(self.detect_source_header)

        grid.addWidget(self._labeled("Sheet", self.cmb_sheet), 0, 0)
        grid.addWidget(self.btn_load_sheets, 0, 1, Qt.AlignBottom)
        grid.addWidget(self._labeled("Source Header Row", self.spin_source_header_rows), 0, 2)
        grid.addWidget(self.btn_detect_source_header, 0, 3, Qt.AlignBottom)
        grid.addWidget(self._labeled("Key Column", self.cmb_key), 0, 4)
        grid.addWidget(self.btn_load_headers, 0, 5, Qt.AlignBottom)
        grid.setColumnStretch(6, 1)
        layout.addLayout(grid)

        self.main_panel_layout.addWidget(card)

    def _build_template_card(self):
        card, layout = self._panel("Template", FIF.EDIT)

        mode_row = QHBoxLayout()
        self.cmb_template_mode = self._fixed_width(ComboBox(), 220)
        self.cmb_template_mode.addItems([
            TEMPLATE_MODE_LABELS[TEMPLATE_MODE_TEMPLATE_FILE],
            TEMPLATE_MODE_LABELS[TEMPLATE_MODE_SOURCE_TEMPLATE],
        ])
        self.cmb_template_mode.setCurrentIndex(0)
        self.cmb_template_mode.currentTextChanged.connect(self.on_template_mode_changed)
        mode_row.addWidget(self._labeled("Template Option", self.cmb_template_mode))
        mode_row.addStretch()
        layout.addLayout(mode_row)

        self.template_file_row_widget = QWidget()
        row1 = QHBoxLayout(self.template_file_row_widget)
        row1.setContentsMargins(0, 0, 0, 0)
        row1.setSpacing(8)
        self.edit_template = self._fixed_width(LineEdit(), PATH_FIELD_WIDTH)
        self.edit_template.setPlaceholderText("Template Excel file")
        self.btn_browse_template = self._field_action_button(ToolButton(FIF.FOLDER))
        self.btn_browse_template.clicked.connect(self.browse_template)
        self.template_header_row_widget = QWidget()
        header_row = QHBoxLayout(self.template_header_row_widget)
        header_row.setContentsMargins(0, 0, 0, 0)
        header_row.setSpacing(8)
        self.spin_template_header_rows = self._fixed_width(SpinBox(), SMALL_FIELD_WIDTH)
        self.spin_template_header_rows.setRange(1, 100)
        self.spin_template_header_rows.setValue(5)
        self.btn_detect_template_header = self._field_action_button(PushButton("Detect"))
        self.btn_detect_template_header.clicked.connect(self.detect_template_header)
        header_row.addWidget(self._labeled("Template Header Row", self.spin_template_header_rows))
        header_row.addWidget(self.btn_detect_template_header, 0, Qt.AlignBottom)
        row1.addWidget(self._labeled("Template Workbook", self.edit_template))
        row1.addWidget(self.btn_browse_template, 0, Qt.AlignBottom)
        row1.addWidget(self.template_header_row_widget)
        row1.addStretch()
        layout.addWidget(self.template_file_row_widget)

        self.main_panel_layout.addWidget(card)

    def _build_mapping_card(self):
        self.mapping_card, layout = self._panel("Column Mapping", FIF.EDIT)

        row = QHBoxLayout()
        self.btn_auto_map = PushButton("Auto Map")
        self.btn_auto_map.clicked.connect(lambda: self.refresh_template_mapping(auto=True))
        self.lbl_mapping_status = CaptionLabel("Map template columns to source columns.")
        row.addWidget(self.btn_auto_map)
        row.addWidget(self.lbl_mapping_status)
        row.addStretch()
        layout.addLayout(row)

        self.mapping_rows_widget = QWidget()
        self.mapping_rows_layout = QVBoxLayout(self.mapping_rows_widget)
        self.mapping_rows_layout.setContentsMargins(0, 0, 0, 0)
        self.mapping_rows_layout.setSpacing(8)
        layout.addWidget(self.mapping_rows_widget)

        self.main_panel_layout.addWidget(self.mapping_card)

    def _build_output_card(self):
        card, layout = self._panel("Output", FIF.FOLDER)

        row1 = QGridLayout()
        row1.setHorizontalSpacing(10)
        row1.setVerticalSpacing(8)
        self.edit_outdir = self._fixed_width(LineEdit(), PATH_FIELD_WIDTH)
        self.edit_outdir.setPlaceholderText("Output folder")
        self.btn_browse_outdir = self._field_action_button(ToolButton(FIF.FOLDER))
        self.btn_browse_outdir.clicked.connect(self.browse_outdir)
        self.cmb_output_type = self._fixed_width(ComboBox(), 140)
        self.cmb_output_type.addItems([
            OUTPUT_TYPE_LABELS[OUTPUT_TYPE_EXCEL],
            OUTPUT_TYPE_LABELS[OUTPUT_TYPE_PDF],
            OUTPUT_TYPE_LABELS[OUTPUT_TYPE_EXCEL_AND_PDF],
        ])
        self.cmb_output_type.setCurrentIndex(0)
        self.cmb_output_type.currentTextChanged.connect(self.on_output_type_changed)
        self.cmb_pdf_engine = self._fixed_width(ComboBox(), 160)
        self.cmb_pdf_engine.addItems(["xlwings", "libreoffice"])
        self.cmb_pdf_engine.setCurrentIndex(0)
        self.cmb_pdf_engine.currentTextChanged.connect(self.on_pdf_engine_changed)
        self.pdf_engine_field_widget = self._labeled("PDF Engine", self.cmb_pdf_engine)

        row1.addWidget(self._labeled("Folder", self.edit_outdir), 0, 0)
        row1.addWidget(self.btn_browse_outdir, 0, 1, Qt.AlignBottom)
        row1.addWidget(self._labeled("Output Type", self.cmb_output_type), 0, 2)
        row1.addWidget(self.pdf_engine_field_widget, 0, 3)
        row1.setColumnStretch(4, 1)
        layout.addLayout(row1)

        options = QGridLayout()
        options.setHorizontalSpacing(10)
        options.setVerticalSpacing(8)
        self.edit_prefix = self._fixed_width(LineEdit(), NAME_FIELD_WIDTH)
        self.edit_prefix.setPlaceholderText("Prefix")
        self.edit_suffix = self._fixed_width(LineEdit(), NAME_FIELD_WIDTH)
        self.edit_suffix.setPlaceholderText("Suffix")
        self.edit_lo_path = self._fixed_width(LineEdit(), SECONDARY_PATH_FIELD_WIDTH)
        self.edit_lo_path.setPlaceholderText("soffice.exe")
        self.btn_browse_soffice = self._field_action_button(ToolButton(FIF.FOLDER))
        self.btn_browse_soffice.clicked.connect(self.browse_soffice)
        self.lbl_filename_preview = CaptionLabel()

        options.addWidget(self._labeled("Prefix", self.edit_prefix), 0, 0)
        options.addWidget(self._labeled("Suffix", self.edit_suffix), 0, 1)
        options.addWidget(self._labeled("Preview", self.lbl_filename_preview), 0, 2)
        options.setColumnStretch(3, 1)
        layout.addLayout(options)

        self.lo_path_row_widget = QWidget()
        lo_row = QHBoxLayout(self.lo_path_row_widget)
        lo_row.setContentsMargins(0, 0, 0, 0)
        lo_row.setSpacing(8)
        lo_row.addWidget(self._labeled("LibreOffice", self.edit_lo_path))
        lo_row.addWidget(self.btn_browse_soffice, 0, Qt.AlignBottom)
        lo_row.addStretch()
        layout.addWidget(self.lo_path_row_widget)
        self.on_output_type_changed()

        self.main_panel_layout.addWidget(card)

    def _build_log_card(self):
        card, layout = self._panel("Log")
        self.txt_log = TextEdit()
        self.txt_log.setReadOnly(True)
        self.txt_log.setMinimumHeight(130)
        self.txt_log.setMaximumHeight(180)
        layout.addWidget(self.txt_log)
        self.main_panel_layout.addWidget(card)

    def _build_mail_merge_card(self):
        self.mail_merge_card, layout = self._panel("Mail Merge", FIF.MAIL)
        self.lbl_mail_merge_summary = CaptionLabel("No split results loaded.")
        layout.addWidget(self.lbl_mail_merge_summary)

        recipient_grid = QGridLayout()
        recipient_grid.setHorizontalSpacing(10)
        recipient_grid.setVerticalSpacing(8)
        self.edit_recipient_path = self._fixed_width(LineEdit(), PATH_FIELD_WIDTH)
        self.edit_recipient_path.setPlaceholderText("Recipient mapping Excel file")
        self.btn_browse_recipients = self._field_action_button(ToolButton(FIF.FOLDER))
        self.btn_browse_recipients.clicked.connect(self.browse_recipient_mapping)
        self.cmb_recipient_sheet = self._fixed_width(ComboBox(), COMBO_FIELD_WIDTH)
        self.cmb_recipient_sheet.setPlaceholderText("Sheet")
        self.btn_load_recipient_sheets = self._field_action_button(ToolButton(FIF.SYNC))
        self.btn_load_recipient_sheets.setToolTip("Load Recipient Sheets")
        self.btn_load_recipient_sheets.clicked.connect(self.load_recipient_sheets)
        self.spin_recipient_header_row = self._fixed_width(SpinBox(), SMALL_FIELD_WIDTH)
        self.spin_recipient_header_row.setRange(1, 100)
        self.spin_recipient_header_row.setValue(1)
        self.btn_load_recipient_headers = self._field_action_button(ToolButton(FIF.SYNC))
        self.btn_load_recipient_headers.setToolTip("Load Recipient Headers")
        self.btn_load_recipient_headers.clicked.connect(self.load_recipient_headers)
        self.cmb_recipient_key = self._fixed_width(ComboBox(), COMBO_FIELD_WIDTH)
        self.cmb_recipient_to = self._fixed_width(ComboBox(), COMBO_FIELD_WIDTH)
        self.cmb_recipient_cc = self._fixed_width(ComboBox(), COMBO_FIELD_WIDTH)
        self.cmb_recipient_bcc = self._fixed_width(ComboBox(), COMBO_FIELD_WIDTH)
        recipient_grid.addWidget(self._labeled("Recipient Workbook", self.edit_recipient_path), 0, 0)
        recipient_grid.addWidget(self.btn_browse_recipients, 0, 1, Qt.AlignBottom)
        recipient_grid.addWidget(self._labeled("Sheet", self.cmb_recipient_sheet), 0, 2)
        recipient_grid.addWidget(self.btn_load_recipient_sheets, 0, 3, Qt.AlignBottom)
        recipient_grid.addWidget(self._labeled("Header Row", self.spin_recipient_header_row), 0, 4)
        recipient_grid.addWidget(self.btn_load_recipient_headers, 0, 5, Qt.AlignBottom)
        recipient_grid.addWidget(self._labeled("Key", self.cmb_recipient_key), 1, 0)
        recipient_grid.addWidget(self._labeled("To", self.cmb_recipient_to), 1, 1)
        recipient_grid.addWidget(self._labeled("CC", self.cmb_recipient_cc), 1, 2)
        recipient_grid.addWidget(self._labeled("BCC", self.cmb_recipient_bcc), 1, 3)
        recipient_grid.setColumnStretch(6, 1)
        layout.addLayout(recipient_grid)

        content_grid = QGridLayout()
        content_grid.setHorizontalSpacing(10)
        content_grid.setVerticalSpacing(8)
        self.edit_mail_subject = self._fixed_width(LineEdit(), PATH_FIELD_WIDTH)
        self.edit_mail_subject.setPlaceholderText("Subject, e.g. Report {key}")
        self.edit_mail_body = TextEdit()
        self.edit_mail_body.setMinimumHeight(90)
        self.edit_mail_html_template = self._fixed_width(LineEdit(), PATH_FIELD_WIDTH)
        self.edit_mail_html_template.setPlaceholderText("Optional HTML template")
        self.btn_browse_mail_html = self._field_action_button(ToolButton(FIF.FOLDER))
        self.btn_browse_mail_html.clicked.connect(self.browse_mail_html_template)
        content_grid.addWidget(self._labeled("Subject", self.edit_mail_subject), 0, 0)
        content_grid.addWidget(self._labeled("Body", self.edit_mail_body), 1, 0)
        content_grid.addWidget(self._labeled("HTML Template", self.edit_mail_html_template), 2, 0)
        content_grid.addWidget(self.btn_browse_mail_html, 2, 1, Qt.AlignBottom)
        layout.addLayout(content_grid)

        self.mail_merge_card.setVisible(False)
        self.main_panel_layout.addWidget(self.mail_merge_card)

    def update_mail_merge_entry_state(self):
        has_results = bool(self.current_split_results)
        self.btn_mail_merge.setVisible(has_results)

    def show_mail_merge_panel(self):
        count = len(self.current_split_results)
        suffix = "file" if count == 1 else "files"
        self.lbl_mail_merge_summary.setText(f"{count} split {suffix} loaded for mail merge.")
        self.mail_merge_card.setVisible(True)
        self.update_mail_merge_entry_state()

    def _build_actions_card(self, layout):
        self.btn_generate = PrimaryPushButton(FIF.PLAY, "Generate")
        self.btn_generate.setFixedHeight(38)
        self.btn_generate.clicked.connect(self.on_run_clicked)
        self.progress_bar = ProgressBar()
        self.progress_bar.setFixedWidth(240)
        self.progress_bar.setValue(0)
        self.progress_bar.setVisible(False)
        self.btn_open_output = PushButton(FIF.FOLDER, "Open Output Folder")
        self.btn_open_output.setFixedHeight(36)
        self.btn_open_output.clicked.connect(self.open_output_folder)
        self.btn_open_output.setVisible(False)
        self.btn_mail_merge = PushButton(FIF.MAIL, "Mail Merge")
        self.btn_mail_merge.setFixedHeight(36)
        self.btn_mail_merge.clicked.connect(self.show_mail_merge_panel)
        self.btn_mail_merge.setVisible(False)
        self.btn_debug = PushButton("Debug Excel")
        self.btn_debug.setFixedHeight(36)
        self.btn_debug.clicked.connect(self.debug_excel)

        layout.addWidget(self.btn_generate)
        layout.addWidget(self.progress_bar)
        layout.addWidget(self.btn_open_output)
        layout.addWidget(self.btn_mail_merge)
        layout.addWidget(self.btn_debug)
        layout.addStretch()
    def log(self, msg):
        self.txt_log.append(msg)

    def set_progress(self, total, current):
        if self.is_running:
            self.progress_bar.setVisible(True)
        if total <= 0:
            self.progress_bar.setValue(0)
        else:
            self.progress_bar.setValue(int(100 * current / total))

    def set_busy(self, busy):
        self.is_running = busy
        self.btn_generate.setEnabled(not busy)
        self.btn_generate.setText("Generating..." if busy else "Generate")
        self.btn_reset_settings.setEnabled(not busy)
        self.progress_bar.setVisible(busy)
        if not busy:
            self.progress_bar.setValue(0)
        self.update_workflow_status()

    def _connect_settings_signals(self):
        for edit in [
            self.edit_source,
            self.edit_template,
            self.edit_outdir,
            self.edit_lo_path,
            self.edit_prefix,
            self.edit_suffix,
        ]:
            edit.editingFinished.connect(self.save_settings)
        self.edit_source.editingFinished.connect(lambda: self.refresh_source_options(silent=True))

        for edit in [self.edit_prefix, self.edit_suffix]:
            edit.textChanged.connect(self.update_filename_preview)

        for combo in [
            self.cmb_sheet,
            self.cmb_key,
            self.cmb_template_mode,
            self.cmb_output_type,
            self.cmb_pdf_engine,
        ]:
            combo.currentTextChanged.connect(self.save_settings)
        self.cmb_key.currentTextChanged.connect(self.update_filename_preview)
        self.cmb_sheet.currentTextChanged.connect(lambda *_: self.load_headers(silent=True))

        for edit in [self.edit_recipient_path, self.edit_mail_subject, self.edit_mail_html_template]:
            edit.editingFinished.connect(self.save_settings)
        self.edit_mail_body.textChanged.connect(self.save_settings)
        for combo in [
            self.cmb_recipient_sheet,
            self.cmb_recipient_key,
            self.cmb_recipient_to,
            self.cmb_recipient_cc,
            self.cmb_recipient_bcc,
        ]:
            combo.currentTextChanged.connect(self.save_settings)
        self.spin_recipient_header_row.valueChanged.connect(self.save_settings)

        self.spin_source_header_rows.valueChanged.connect(self.save_settings)
        self.spin_source_header_rows.valueChanged.connect(lambda *_: self.refresh_template_mapping(auto=True))
        self.spin_template_header_rows.valueChanged.connect(self.save_settings)
        self.spin_template_header_rows.valueChanged.connect(lambda *_: self.refresh_template_mapping(auto=True))

    def save_settings(self):
        if self._loading_settings:
            return

        mapping = self.collect_column_mapping() if self.mapping_combos else self.saved_column_mapping
        self.saved_column_mapping = mapping

        self.settings.setValue("source_path", self.edit_source.text().strip())
        self.settings.setValue("sheet_name", self.cmb_sheet.currentText().strip())
        self.settings.setValue("key_col", self.cmb_key.currentText().strip())
        self.settings.setValue("template_mode", self.current_template_mode())
        self.settings.setValue("template_path", self.edit_template.text().strip())
        self.settings.setValue("header_rows", self.spin_source_header_rows.value())
        self.settings.setValue("source_header_rows", self.spin_source_header_rows.value())
        self.settings.setValue("template_header_rows", self.spin_template_header_rows.value())
        self.settings.setValue("output_dir", self.edit_outdir.text().strip())
        self.settings.setValue("output_file_type", self.current_output_file_type())
        self.settings.setValue("pdf_engine", self.cmb_pdf_engine.currentText().strip().lower())
        self.settings.setValue("libreoffice_path", self.edit_lo_path.text().strip())
        self.settings.setValue("prefix", self.edit_prefix.text().strip())
        self.settings.setValue("suffix", self.edit_suffix.text().strip())
        self.settings.setValue("column_mapping", json.dumps(mapping))
        self.settings.setValue("mail_recipient_path", self.edit_recipient_path.text().strip())
        self.settings.setValue("mail_recipient_sheet", self.cmb_recipient_sheet.currentText().strip())
        self.settings.setValue("mail_recipient_header_row", self.spin_recipient_header_row.value())
        self.settings.setValue("mail_recipient_key_col", self.cmb_recipient_key.currentText().strip())
        self.settings.setValue("mail_recipient_to_col", self.cmb_recipient_to.currentText().strip())
        self.settings.setValue("mail_recipient_cc_col", self.cmb_recipient_cc.currentText().strip())
        self.settings.setValue("mail_recipient_bcc_col", self.cmb_recipient_bcc.currentText().strip())
        self.settings.setValue("mail_subject", self.edit_mail_subject.text().strip())
        self.settings.setValue("mail_body", self.edit_mail_body.toPlainText())
        self.settings.setValue("mail_html_template", self.edit_mail_html_template.text().strip())
        self.settings.sync()
        self.update_workflow_status()

    def load_settings(self):
        self._loading_settings = True
        try:
            self.edit_source.setText(self.settings.value("source_path", ""))
            self.edit_template.setText(self.settings.value("template_path", ""))
            legacy_header_rows = self.settings.value("header_rows", 5)
            self.spin_source_header_rows.setValue(
                int(self.settings.value("source_header_rows", legacy_header_rows))
            )
            self.spin_template_header_rows.setValue(
                int(self.settings.value("template_header_rows", legacy_header_rows))
            )
            self.edit_outdir.setText(self.settings.value("output_dir", ""))
            self.edit_lo_path.setText(self.settings.value("libreoffice_path", ""))
            self.edit_prefix.setText(self.settings.value("prefix", ""))
            self.edit_suffix.setText(self.settings.value("suffix", ""))
            self.edit_recipient_path.setText(self.settings.value("mail_recipient_path", ""))
            self.spin_recipient_header_row.setValue(int(self.settings.value("mail_recipient_header_row", 1)))
            self.edit_mail_subject.setText(self.settings.value("mail_subject", ""))
            self.edit_mail_body.setPlainText(self.settings.value("mail_body", ""))
            self.edit_mail_html_template.setText(self.settings.value("mail_html_template", ""))

            sheet = self.settings.value("sheet_name", "")
            if sheet:
                self.cmb_sheet.clear()
                self.cmb_sheet.addItem(sheet)
                self.cmb_sheet.setCurrentIndex(0)

            key = self.settings.value("key_col", "")
            if key:
                self.cmb_key.clear()
                self.cmb_key.addItem(key)
                self.cmb_key.setCurrentIndex(0)

            mail_sheet = self.settings.value("mail_recipient_sheet", "")
            if mail_sheet:
                self.cmb_recipient_sheet.clear()
                self.cmb_recipient_sheet.addItem(mail_sheet)
                self.cmb_recipient_sheet.setCurrentIndex(0)

            for setting_key, combo in [
                ("mail_recipient_key_col", self.cmb_recipient_key),
                ("mail_recipient_to_col", self.cmb_recipient_to),
                ("mail_recipient_cc_col", self.cmb_recipient_cc),
                ("mail_recipient_bcc_col", self.cmb_recipient_bcc),
            ]:
                value = self.settings.value(setting_key, "")
                if value:
                    combo.clear()
                    combo.addItem(value)
                    combo.setCurrentIndex(0)

            mode = self.settings.value("template_mode", TEMPLATE_MODE_TEMPLATE_FILE)
            mode_label = TEMPLATE_MODE_LABELS.get(mode, TEMPLATE_MODE_LABELS[TEMPLATE_MODE_TEMPLATE_FILE])
            mode_idx = self.cmb_template_mode.findText(mode_label)
            if mode_idx >= 0:
                self.cmb_template_mode.setCurrentIndex(mode_idx)

            pdf_engine = self.settings.value("pdf_engine", "xlwings")
            output_file_type = self.settings.value("output_file_type", "")
            if not output_file_type:
                has_legacy_pdf_engine = self.settings.contains("pdf_engine")
                output_file_type = (
                    OUTPUT_TYPE_PDF
                    if has_legacy_pdf_engine and str(pdf_engine).strip().lower() not in {"", "none"}
                    else OUTPUT_TYPE_EXCEL
                )
            output_label = OUTPUT_TYPE_LABELS.get(output_file_type, OUTPUT_TYPE_LABELS[OUTPUT_TYPE_EXCEL])
            output_idx = self.cmb_output_type.findText(output_label)
            if output_idx >= 0:
                self.cmb_output_type.setCurrentIndex(output_idx)

            if str(pdf_engine).strip().lower() == "none":
                pdf_engine = "xlwings"
            pdf_idx = self.cmb_pdf_engine.findText(pdf_engine)
            if pdf_idx >= 0:
                self.cmb_pdf_engine.setCurrentIndex(pdf_idx)

            mapping_raw = self.settings.value("column_mapping", "{}")
            try:
                mapping = json.loads(mapping_raw) if mapping_raw else {}
                self.saved_column_mapping = mapping if isinstance(mapping, dict) else {}
            except Exception:
                self.saved_column_mapping = {}
        finally:
            self._loading_settings = False

        self.on_template_mode_changed()
        self.on_output_type_changed()
        self.update_workflow_status()
        self.update_filename_preview()

    def reset_settings(self):
        self.settings.clear()
        self.settings.sync()
        self._loading_settings = True
        try:
            self.edit_source.clear()
            self.edit_template.clear()
            self.edit_outdir.clear()
            self.edit_lo_path.clear()
            self.edit_prefix.clear()
            self.edit_suffix.clear()
            self.edit_recipient_path.clear()
            self.edit_mail_subject.clear()
            self.edit_mail_body.clear()
            self.edit_mail_html_template.clear()
            self.cmb_sheet.clear()
            self.cmb_key.clear()
            self.cmb_recipient_sheet.clear()
            self.cmb_recipient_key.clear()
            self.cmb_recipient_to.clear()
            self.cmb_recipient_cc.clear()
            self.cmb_recipient_bcc.clear()
            self.cmb_template_mode.setCurrentIndex(0)
            self.cmb_output_type.setCurrentIndex(0)
            self.cmb_pdf_engine.setCurrentIndex(0)
            self.spin_source_header_rows.setValue(5)
            self.spin_template_header_rows.setValue(5)
            self.spin_recipient_header_row.setValue(1)
            self.source_headers = []
            self.template_headers = []
            self.saved_column_mapping = {}
            self.render_mapping_rows({})
        finally:
            self._loading_settings = False
        self.on_template_mode_changed()
        self.on_output_type_changed()
        self.update_workflow_status()
        self.update_filename_preview()

    def current_template_mode(self):
        return TEMPLATE_MODE_BY_LABEL.get(
            self.cmb_template_mode.currentText(),
            TEMPLATE_MODE_TEMPLATE_FILE,
        )

    def current_output_file_type(self):
        return OUTPUT_TYPE_BY_LABEL.get(
            self.cmb_output_type.currentText(),
            OUTPUT_TYPE_EXCEL,
        )

    def on_template_mode_changed(self):
        if not hasattr(self, "mapping_card"):
            return
        use_template_file = self.current_template_mode() == TEMPLATE_MODE_TEMPLATE_FILE
        self.template_header_row_widget.setVisible(use_template_file)
        self.template_file_row_widget.setVisible(use_template_file)
        self.edit_template.setVisible(use_template_file)
        self.btn_browse_template.setVisible(use_template_file)
        self.spin_template_header_rows.setVisible(use_template_file)
        self.btn_detect_template_header.setVisible(use_template_file)
        self.mapping_card.setVisible(use_template_file)
        if use_template_file:
            self.refresh_template_mapping(auto=True)
        self.update_workflow_status()

    def on_output_type_changed(self, *_):
        if not hasattr(self, "pdf_engine_field_widget"):
            return
        use_pdf = output_requires_pdf(self.current_output_file_type())
        self.pdf_engine_field_widget.setVisible(use_pdf)
        self.cmb_pdf_engine.setVisible(use_pdf)
        self.on_pdf_engine_changed()
        self.update_filename_preview()
        self.update_workflow_status()

    def on_pdf_engine_changed(self, *_):
        if not hasattr(self, "lo_path_row_widget"):
            return
        use_libreoffice = (
            output_requires_pdf(self.current_output_file_type())
            and self.cmb_pdf_engine.currentText().strip().lower() == "libreoffice"
        )
        self.lo_path_row_widget.setVisible(use_libreoffice)
        self.edit_lo_path.setVisible(use_libreoffice)
        self.btn_browse_soffice.setVisible(use_libreoffice)
        self.update_filename_preview()

    def update_filename_preview(self, *_):
        if not hasattr(self, "lbl_filename_preview"):
            return
        key_label = self.cmb_key.currentText().strip() or "key"
        key_preview = f"<{key_label} value>"
        filename = join_output_name_parts(
            self.edit_prefix.text().strip(),
            key_preview,
            self.edit_suffix.text().strip(),
        ) + output_extension(self.current_output_file_type())
        self.lbl_filename_preview.setText(filename)

    def _clear_mapping_rows(self):
        while self.mapping_rows_layout.count():
            item = self.mapping_rows_layout.takeAt(0)
            widget = item.widget()
            if widget is not None:
                widget.deleteLater()
        self.mapping_combos = {}
        self.mapping_status_labels = {}

    def render_mapping_rows(self, mapping=None):
        self._clear_mapping_rows()
        mapping = mapping or {}
        if not self.template_headers:
            self.mapping_rows_layout.addWidget(CaptionLabel("No template headers loaded."))
            return

        source_choices = [""] + self.source_headers
        for template_header in self.template_headers:
            row_widget = QWidget()
            row_widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
            row = QHBoxLayout(row_widget)
            row.setContentsMargins(0, 0, 0, 0)
            row.addWidget(self._fixed_width(BodyLabel(template_header), 220))
            row.addWidget(BodyLabel("->"))
            combo = self._fixed_width(ComboBox(), 240)
            combo.addItems(source_choices)
            selected = mapping.get(template_header)
            if selected:
                idx = combo.findText(selected)
                if idx >= 0:
                    combo.setCurrentIndex(idx)
            status = CaptionLabel("Mapped" if combo.currentText().strip() else "Missing")

            def on_mapping_changed(value, label=status):
                label.setText("Mapped" if value.strip() else "Missing")
                self.save_settings()

            combo.currentTextChanged.connect(on_mapping_changed)
            row.addWidget(combo)
            row.addWidget(status)
            row.addStretch()
            self.mapping_rows_layout.addWidget(row_widget)
            self.mapping_combos[template_header] = combo
            self.mapping_status_labels[template_header] = status

    def collect_column_mapping(self):
        return {
            header: combo.currentText().strip()
            for header, combo in self.mapping_combos.items()
            if combo.currentText().strip()
        }

    def refresh_template_mapping(self, auto=True):
        if self.current_template_mode() != TEMPLATE_MODE_TEMPLATE_FILE:
            self.mapping_card.setVisible(False)
            return

        self.mapping_card.setVisible(True)
        src = self.edit_source.text().strip()
        sheet = self.cmb_sheet.currentText().strip()
        template = self.edit_template.text().strip()

        try:
            if src and sheet:
                self.source_headers = read_excel_headers(
                    Path(src),
                    sheet,
                    self.spin_source_header_rows.value(),
                )
            if template and Path(template).exists():
                self.template_headers, self.template_col_start = read_template_headers(
                    Path(template),
                    self.spin_template_header_rows.value(),
                )
            else:
                self.template_headers = []

            if auto:
                mapping = auto_map_columns(self.template_headers, self.source_headers)
                for key, value in self.saved_column_mapping.items():
                    if key in mapping and value in self.source_headers:
                        mapping[key] = value
            else:
                mapping = self.collect_column_mapping()
            self.saved_column_mapping = {k: v for k, v in mapping.items() if v}
            self.render_mapping_rows(mapping)
        except Exception as e:
            self.log(f"Mapping error: {e}")
            self.template_headers = []
            self.render_mapping_rows({})

    def update_workflow_status(self):
        if not hasattr(self, "workflow_steps"):
            return

        states = {
            "Source": bool(
                self.edit_source.text().strip()
                and self.cmb_sheet.currentText().strip()
                and self.cmb_key.currentText().strip()
            ),
            "Template": (
                self.current_template_mode() == TEMPLATE_MODE_SOURCE_TEMPLATE
                or bool(self.edit_template.text().strip())
            ),
            "Output": bool(self.edit_outdir.text().strip()),
            "Run": not self.is_running,
        }
        for name, ready in states.items():
            self.workflow_steps[name].setText("●" if ready else "○")
        missing = [name for name, ready in states.items() if not ready and name != "Run"]
        self.lbl_workflow_status.setText("Ready" if not missing else "Missing: " + ", ".join(missing))

    def browse_source(self):
        f, _ = QFileDialog.getOpenFileName(
            self, "Pilih source Excel",
            "", "Excel files (*.xlsx *.xls *.xlsm *.xlsb)"
        )
        if f:
            self.edit_source.setText(f)
            self.log(f"Source: {f}")
            self.refresh_source_options()

    def browse_template(self):
        f, _ = QFileDialog.getOpenFileName(
            self, "Pilih template Excel",
            "", "Excel files (*.xlsx)"
        )
        if f:
            self.edit_template.setText(f)
            self.log(f"Template: {f}")
            self.detect_template_header(silent=True)
            self.refresh_template_mapping(auto=True)
            self.update_workflow_status()

    def browse_outdir(self):
        d = QFileDialog.getExistingDirectory(self, "Pilih output folder")
        if d:
            self.edit_outdir.setText(d)
            self.log(f"Output: {d}")
            self.update_workflow_status()

    def browse_soffice(self):
        f, _ = QFileDialog.getOpenFileName(
            self, "Pilih soffice.exe (LibreOffice)",
            "", "Executable (soffice.exe);;All files (*.*)"
        )
        if f:
            self.edit_lo_path.setText(f)
            self.log(f"LibreOffice: {f}")

    def browse_recipient_mapping(self):
        f, _ = QFileDialog.getOpenFileName(
            self, "Pilih recipient mapping Excel",
            "", "Excel files (*.xlsx *.xls *.xlsm)"
        )
        if f:
            self.edit_recipient_path.setText(f)
            self.load_recipient_sheets()

    def load_recipient_sheets(self):
        path = self.edit_recipient_path.text().strip()
        if not path:
            InfoBar.warning("Perhatian", "Pilih recipient mapping Excel dulu.", parent=self, duration=3000, position=InfoBarPosition.TOP)
            return
        try:
            with pd.ExcelFile(path) as xls:
                sheets = list(xls.sheet_names)
            self.cmb_recipient_sheet.clear()
            self.cmb_recipient_sheet.addItems(sheets)
            if sheets:
                self.cmb_recipient_sheet.setCurrentIndex(0)
                self.load_recipient_headers()
            self.save_settings()
        except Exception as e:
            InfoBar.error("Error", str(e), parent=self, duration=5000, position=InfoBarPosition.TOP)

    def load_recipient_headers(self):
        path = self.edit_recipient_path.text().strip()
        sheet = self.cmb_recipient_sheet.currentText().strip()
        if not path or not sheet:
            return
        try:
            headers = read_recipient_headers(Path(path), sheet, self.spin_recipient_header_row.value())
            choices = [""] + headers
            for combo in [self.cmb_recipient_key, self.cmb_recipient_to, self.cmb_recipient_cc, self.cmb_recipient_bcc]:
                combo.clear()
                combo.addItems(choices)
            self._select_combo_text(self.cmb_recipient_key, "Key")
            self._select_combo_text(self.cmb_recipient_to, "To")
            self._select_combo_text(self.cmb_recipient_cc, "CC")
            self._select_combo_text(self.cmb_recipient_bcc, "BCC")
            self.save_settings()
        except Exception as e:
            InfoBar.error("Error", str(e), parent=self, duration=5000, position=InfoBarPosition.TOP)

    def _select_combo_text(self, combo, text):
        idx = combo.findText(text)
        if idx >= 0:
            combo.setCurrentIndex(idx)

    def browse_mail_html_template(self):
        f, _ = QFileDialog.getOpenFileName(self, "Pilih HTML template", "", "HTML files (*.html *.htm)")
        if f:
            self.edit_mail_html_template.setText(f)
            self.save_settings()

    def refresh_source_options(self, silent=False):
        self.load_sheets(load_headers=True, silent=silent)

    def load_sheets(self, *_, load_headers=True, silent=False):
        src = self.edit_source.text().strip()
        if not src:
            if not silent:
                InfoBar.warning("Perhatian", "Pilih source Excel dulu.", parent=self, duration=3000, position=InfoBarPosition.TOP)
            return
        try:
            with pd.ExcelFile(src) as xls:
                sheets = list(xls.sheet_names)
            was_blocked = self.cmb_sheet.blockSignals(True)
            self.cmb_sheet.clear()
            self.cmb_sheet.addItems(sheets)
            if sheets:
                self.cmb_sheet.setCurrentIndex(0)
            self.cmb_sheet.blockSignals(was_blocked)
            if sheets:
                self.detect_source_header(silent=True)
                if load_headers:
                    self.load_headers(silent=True)
            self.log(f"Sheets loaded: {', '.join(sheets)}")
            self.save_settings()
            self.update_workflow_status()
        except Exception as e:
            if not silent:
                InfoBar.error("Error", str(e), parent=self, duration=5000, position=InfoBarPosition.TOP)

    def detect_source_header(self, silent=False):
        src = self.edit_source.text().strip()
        sheet = self.cmb_sheet.currentText().strip()
        if not src or not sheet:
            if not silent:
                InfoBar.warning("Perhatian", "Pastikan source & sheet sudah dipilih.", parent=self, duration=3000, position=InfoBarPosition.TOP)
            return None
        try:
            row = detect_excel_header_row(Path(src), sheet)
            self.spin_source_header_rows.setValue(row)
            self.log(f"Source header row detected: {row}")
            return row
        except Exception as e:
            if not silent:
                InfoBar.error("Error", f"Gagal detect source header: {e}", parent=self, duration=5000, position=InfoBarPosition.TOP)
            return None

    def detect_template_header(self, silent=False):
        template = self.edit_template.text().strip()
        if not template:
            if not silent:
                InfoBar.warning("Perhatian", "Pilih template Excel dulu.", parent=self, duration=3000, position=InfoBarPosition.TOP)
            return None
        try:
            row = detect_excel_header_row(Path(template))
            self.spin_template_header_rows.setValue(row)
            self.log(f"Template header row detected: {row}")
            self.refresh_template_mapping(auto=True)
            return row
        except Exception as e:
            if not silent:
                InfoBar.error("Error", f"Gagal detect template header: {e}", parent=self, duration=5000, position=InfoBarPosition.TOP)
            return None

    def load_headers(self, *_, silent=False):
        src = self.edit_source.text().strip()
        sheet = self.cmb_sheet.currentText().strip()
        if not src or not sheet:
            if not silent:
                InfoBar.warning("Perhatian", "Pastikan source & sheet sudah dipilih.", parent=self, duration=3000, position=InfoBarPosition.TOP)
            return
        try:
            header_row_idx = self.spin_source_header_rows.value() - 1
            df = pd.read_excel(src, sheet_name=sheet, header=header_row_idx, nrows=0)
            headers = list(df.columns.astype(str))
            self.source_headers = headers
            index_vals = [str(i+1) for i in range(len(headers))]
            values = headers + index_vals
            self.cmb_key.clear()
            self.cmb_key.addItems(values)
            if headers:
                self.cmb_key.setCurrentIndex(0)
            self.log(f"Headers loaded: {headers}")
            self.refresh_template_mapping(auto=True)
            self.update_workflow_status()
            self.update_filename_preview()
        except Exception as e:
            if not silent:
                InfoBar.error("Error", str(e), parent=self, duration=5000, position=InfoBarPosition.TOP)

    def debug_excel(self):
        try:
            self.log("=== Excel Detection Debug ===")
            results = debug_excel_detection()
            for result in results:
                self.log(result)
            excel_available = check_excel_availability()
            self.log(f"Final check_excel_availability(): {excel_available}")
            self.log("=== Debug selesai ===")
            if excel_available:
                InfoBar.success("Debug Excel", "Excel terdeteksi dan dapat diakses!", parent=self, duration=3000, position=InfoBarPosition.TOP)
            else:
                InfoBar.warning("Debug Excel", "Excel tidak dapat diakses. Lihat log.", parent=self, duration=5000, position=InfoBarPosition.TOP)
        except Exception as e:
            self.log(f"Error saat debug: {str(e)}")
            InfoBar.error("Debug Error", f"Gagal debug: {str(e)}", parent=self, duration=5000, position=InfoBarPosition.TOP)

    def open_output_folder(self):
        out_dir = self.edit_outdir.text().strip()
        if out_dir and Path(out_dir).exists():
            try:
                win_path = str(Path(out_dir).resolve())
                subprocess.run(f'explorer "{win_path}"', shell=True)
            except Exception as e:
                InfoBar.error("Error", f"Failed to open folder: {str(e)}", parent=self, duration=5000, position=InfoBarPosition.TOP)
        else:
            InfoBar.warning("Warning", "Output folder not set or doesn't exist", parent=self, duration=3000, position=InfoBarPosition.TOP)
    def on_run_clicked(self):
        if self.is_running:
            return
        try:
            source_path = Path(self.edit_source.text().strip())
            template_path = Path(self.edit_template.text().strip())
            out_dir = Path(self.edit_outdir.text().strip())
            sheet_name = self.cmb_sheet.currentText().strip()
            key_raw = self.cmb_key.currentText().strip()
            source_header_rows = self.spin_source_header_rows.value()
            template_header_rows = self.spin_template_header_rows.value()
            output_file_type = self.current_output_file_type()
            pdf_engine = (
                self.cmb_pdf_engine.currentText().strip().lower()
                if output_requires_pdf(output_file_type)
                else "none"
            )
            template_mode = self.current_template_mode()

            if not source_path.exists():
                InfoBar.error("Error", "Source Excel tidak ditemukan.", parent=self, duration=5000, position=InfoBarPosition.TOP)
                return
            if template_mode == TEMPLATE_MODE_TEMPLATE_FILE and not template_path.exists():
                InfoBar.error("Error", "Template Excel tidak ditemukan.", parent=self, duration=5000, position=InfoBarPosition.TOP)
                return
            if not self.edit_outdir.text().strip():
                InfoBar.error("Error", "Output folder belum dipilih.", parent=self, duration=5000, position=InfoBarPosition.TOP)
                return
            if not sheet_name:
                InfoBar.error("Error", "Sheet belum dipilih.", parent=self, duration=5000, position=InfoBarPosition.TOP)
                return
            if not key_raw:
                InfoBar.error("Error", "Key Column belum dipilih/diisi.", parent=self, duration=5000, position=InfoBarPosition.TOP)
                return

            try:
                key_col = int(key_raw)
            except ValueError:
                key_col = key_raw

            column_mapping = None
            if template_mode == TEMPLATE_MODE_TEMPLATE_FILE:
                self.refresh_template_mapping(auto=False)
                if not self.template_headers:
                    InfoBar.error(
                        "Mapping",
                        "Header template tidak ditemukan. Periksa Header Rows atau file template.",
                        parent=self,
                        duration=8000,
                        position=InfoBarPosition.TOP,
                    )
                    return
                column_mapping = self.collect_column_mapping()
                missing = validate_column_mapping(self.template_headers, column_mapping)
                if missing:
                    InfoBar.error(
                        "Mapping",
                        "Lengkapi mapping kolom: " + ", ".join(missing),
                        parent=self,
                        duration=8000,
                        position=InfoBarPosition.TOP,
                    )
                    return

            if output_requires_pdf(output_file_type) and pdf_engine == "xlwings":
                if not XLWINGS_AVAILABLE:
                    InfoBar.warning("xlwings", "xlwings belum terpasang. Gunakan LibreOffice atau output Excel.", parent=self, duration=5000, position=InfoBarPosition.TOP)
                    return
                elif not check_excel_availability():
                    InfoBar.warning("Excel", "Microsoft Excel tidak dapat diakses via COM.", parent=self, duration=5000, position=InfoBarPosition.TOP)
                    return

            soffice_path = None
            if output_requires_pdf(output_file_type) and pdf_engine == "libreoffice":
                lo_explicit = self.edit_lo_path.text().strip()
                soffice_path = find_soffice(lo_explicit)
                if not soffice_path:
                    InfoBar.error("Error", "LibreOffice (soffice.exe) tidak ditemukan.", parent=self, duration=5000, position=InfoBarPosition.TOP)
                    return

            self.set_busy(True)
            self.save_settings()
            self.log("Mulai generate...")

            if pdf_engine == "xlwings":
                self.log("Membersihkan Excel COM sessions...")
                cleanup_excel_com()

            params = {
                'source_path': source_path,
                'sheet_name': sheet_name,
                'key_col': key_col,
                'template_path': template_path,
                'out_dir': out_dir,
                'header_rows': source_header_rows,
                'source_header_rows': source_header_rows,
                'template_header_rows': template_header_rows,
                'output_file_type': output_file_type,
                'pdf_engine': pdf_engine,
                'soffice_path': soffice_path,
                'prefix': self.edit_prefix.text().strip(),
                'suffix': self.edit_suffix.text().strip(),
                'template_mode': template_mode,
                'column_mapping': column_mapping,
            }

            self.worker = SplitWorker(params)
            self.worker.status.connect(self.log)
            self.worker.progress.connect(self.set_progress)
            self.worker.finished.connect(self._on_worker_finished)
            self.worker.error.connect(self._on_worker_error)
            self.worker.start()

        except Exception as e:
            InfoBar.error("Error", str(e), parent=self, duration=5000, position=InfoBarPosition.TOP)
            self.set_busy(False)

    def _on_worker_finished(self):
        self.set_busy(False)
        self.log("Selesai.")
        self.current_split_results = list(getattr(self.worker, "results", []))
        self.update_mail_merge_entry_state()
        self.btn_open_output.setVisible(True)
        if (
            output_requires_pdf(self.current_output_file_type())
            and self.cmb_pdf_engine.currentText().strip().lower() == "xlwings"
        ):
            cleanup_excel_com()
        InfoBar.success("Selesai", "Proses selesai.", parent=self, duration=5000, position=InfoBarPosition.TOP)

    def _on_worker_error(self, error_msg):
        self.set_busy(False)
        self.log(f"Error: {error_msg}")
        InfoBar.error("Error", error_msg, parent=self, duration=8000, position=InfoBarPosition.TOP)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = SplitApp()
    window.show()
    sys.exit(app.exec())
