from contextlib import redirect_stdout
from datetime import datetime
from io import StringIO
from pathlib import Path
import tempfile
import unittest

from openpyxl import Workbook, load_workbook


with redirect_stdout(StringIO()):
    import main


class HeaderMappingTests(unittest.TestCase):
    def test_normalize_header_ignores_case_spaces_and_punctuation(self):
        self.assertEqual(main.normalize_header(" Employee  Name!"), "employeename")

    def test_auto_map_columns_matches_normalized_headers(self):
        mapping = main.auto_map_columns(
            ["Employee Name", "Dept"],
            ["employee_name", "Dept", "Amount"],
        )

        self.assertEqual(mapping, {"Employee Name": "employee_name", "Dept": "Dept"})

    def test_validate_column_mapping_reports_unmapped_template_columns(self):
        missing = main.validate_column_mapping(
            ["Employee Name", "Dept"],
            {"Employee Name": "Name", "Dept": None},
        )

        self.assertEqual(missing, ["Dept"])

    def test_detect_excel_header_row_prefers_row_with_header_like_cells(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "source.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Data"
            ws.append(["Monthly report", None, None])
            ws.append([None, None, None])
            ws.append(["Name", "Dept", "Amount"])
            ws.append(["Alice", "A", 10])
            wb.save(path)

            self.assertEqual(main.detect_excel_header_row(path, "Data"), 3)


class TemplateFileSplitTests(unittest.TestCase):
    def make_source_workbook(self, path: Path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(["Name", "Dept", "Amount"])
        ws.append(["Alice", "A", 10])
        ws.append(["Bob", "B", 20])
        ws.append(["Ana", "A", 30])
        wb.save(path)

    def make_template_workbook(self, path: Path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Template"
        ws.append(["Worker", "Team"])
        wb.save(path)

    def test_template_file_mode_uses_manual_mapping_for_different_headers(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            source = tmp_path / "source.xlsx"
            template = tmp_path / "template.xlsx"
            out_dir = tmp_path / "out"
            self.make_source_workbook(source)
            self.make_template_workbook(template)

            main.split_excel_with_template(
                source,
                "Data",
                "Dept",
                template,
                out_dir,
                1,
                pdf_engine="none",
                template_mode="template_file",
                column_mapping={"Worker": "Name", "Team": "Dept"},
            )

            wb = load_workbook(out_dir / "A.xlsx", data_only=True)
            ws = wb.active
            self.assertEqual(ws["A1"].value, "Worker")
            self.assertEqual(ws["B1"].value, "Team")
            self.assertEqual(ws["A2"].value, "Alice")
            self.assertEqual(ws["B2"].value, "A")
            self.assertEqual(ws["A3"].value, "Ana")
            self.assertEqual(ws["B3"].value, "A")

    def test_template_file_mode_supports_separate_source_and_template_header_rows(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            source = tmp_path / "source.xlsx"
            template = tmp_path / "template.xlsx"
            out_dir = tmp_path / "out"

            wb = Workbook()
            ws = wb.active
            ws.title = "Data"
            ws.append(["Report export", None, None])
            ws.append(["Name", "Dept", "Amount"])
            ws.append(["Alice", "A", 10])
            ws.append(["Bob", "B", 20])
            ws.append(["Ana", "A", 30])
            wb.save(source)

            wb = Workbook()
            ws = wb.active
            ws.title = "Template"
            ws.append(["Company report", None])
            ws.append([None, None])
            ws.append(["Worker", "Team"])
            wb.save(template)

            main.split_excel_with_template(
                source,
                "Data",
                "Dept",
                template,
                out_dir,
                1,
                pdf_engine="none",
                template_mode="template_file",
                column_mapping={"Worker": "Name", "Team": "Dept"},
                source_header_rows=2,
                template_header_rows=3,
            )

            wb = load_workbook(out_dir / "A.xlsx", data_only=True)
            ws = wb.active
            self.assertEqual(ws["A1"].value, "Company report")
            self.assertEqual(ws["A3"].value, "Worker")
            self.assertEqual(ws["B3"].value, "Team")
            self.assertEqual(ws["A4"].value, "Alice")
            self.assertEqual(ws["B4"].value, "A")
            self.assertEqual(ws["A5"].value, "Ana")
            self.assertEqual(ws["B5"].value, "A")

    def test_template_file_mode_maps_date_headers_read_as_datetime_objects(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            source = tmp_path / "source.xlsx"
            template = tmp_path / "template.xlsx"
            out_dir = tmp_path / "out"
            date_header = datetime(2026, 6, 1)

            wb = Workbook()
            ws = wb.active
            ws.title = "Data"
            ws.append(["Dept", date_header])
            ws.append(["A", 10])
            ws.append(["B", 20])
            ws.append(["A", 30])
            wb.save(source)

            wb = Workbook()
            ws = wb.active
            ws.title = "Template"
            ws.append(["Dept", date_header])
            wb.save(template)

            main.split_excel_with_template(
                source,
                "Data",
                "Dept",
                template,
                out_dir,
                1,
                pdf_engine="none",
                template_mode="template_file",
            )

            wb = load_workbook(out_dir / "A.xlsx", data_only=True)
            ws = wb.active
            self.assertEqual(ws["B1"].value, date_header)
            self.assertEqual(ws["B2"].value, 10)
            self.assertEqual(ws["B3"].value, 30)

    def test_template_file_mode_accepts_date_key_column_selected_by_display_text(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            source = tmp_path / "source.xlsx"
            template = tmp_path / "template.xlsx"
            out_dir = tmp_path / "out"
            date_header = datetime(2026, 6, 1)

            wb = Workbook()
            ws = wb.active
            ws.title = "Data"
            ws.append([date_header, "Name"])
            ws.append(["A", "Alice"])
            ws.append(["B", "Bob"])
            ws.append(["A", "Ana"])
            wb.save(source)

            wb = Workbook()
            ws = wb.active
            ws.title = "Template"
            ws.append(["Name"])
            wb.save(template)

            main.split_excel_with_template(
                source,
                "Data",
                str(date_header),
                template,
                out_dir,
                1,
                pdf_engine="none",
                template_mode="template_file",
                column_mapping={"Name": "Name"},
            )

            wb = load_workbook(out_dir / "A.xlsx", data_only=True)
            ws = wb.active
            self.assertEqual(ws["A2"].value, "Alice")
            self.assertEqual(ws["A3"].value, "Ana")

    def test_template_file_mode_rejects_duplicate_template_headers(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            source = tmp_path / "source.xlsx"
            template = tmp_path / "template.xlsx"
            out_dir = tmp_path / "out"
            self.make_source_workbook(source)

            wb = Workbook()
            ws = wb.active
            ws.title = "Template"
            ws.append(["Name", "Name"])
            wb.save(template)

            with self.assertRaisesRegex(ValueError, "Header template duplikat"):
                main.split_excel_with_template(
                    source,
                    "Data",
                    "Dept",
                    template,
                    out_dir,
                    1,
                    pdf_engine="none",
                    template_mode="template_file",
                    column_mapping={"Name": "Name"},
                )

    def test_template_file_mode_preserves_blank_hidden_template_columns(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            source = tmp_path / "source.xlsx"
            template = tmp_path / "template.xlsx"
            out_dir = tmp_path / "out"
            self.make_source_workbook(source)

            wb = Workbook()
            ws = wb.active
            ws.title = "Template"
            ws["A1"] = "Worker"
            ws["C1"] = "Team"
            ws.column_dimensions["B"].hidden = True
            wb.save(template)

            main.split_excel_with_template(
                source,
                "Data",
                "Dept",
                template,
                out_dir,
                1,
                pdf_engine="none",
                template_mode="template_file",
                column_mapping={"Worker": "Name", "Team": "Dept"},
            )

            wb = load_workbook(out_dir / "A.xlsx", data_only=True)
            ws = wb.active
            self.assertEqual(ws["A2"].value, "Alice")
            self.assertIsNone(ws["B2"].value)
            self.assertTrue(ws.column_dimensions["B"].hidden)
            self.assertEqual(ws["C2"].value, "A")

    def test_template_file_mode_requires_complete_manual_mapping(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            source = tmp_path / "source.xlsx"
            template = tmp_path / "template.xlsx"
            out_dir = tmp_path / "out"
            self.make_source_workbook(source)
            self.make_template_workbook(template)

            with self.assertRaisesRegex(ValueError, "Team"):
                main.split_excel_with_template(
                    source,
                    "Data",
                    "Dept",
                    template,
                    out_dir,
                    1,
                    pdf_engine="none",
                    template_mode="template_file",
                    column_mapping={"Worker": "Name"},
                )

    def test_template_file_mode_requires_template_headers_for_mapping(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            source = tmp_path / "source.xlsx"
            template = tmp_path / "template.xlsx"
            out_dir = tmp_path / "out"
            self.make_source_workbook(source)

            wb = Workbook()
            ws = wb.active
            ws.title = "Template"
            ws["A2"] = "data row style only"
            wb.save(template)

            with self.assertRaisesRegex(ValueError, "Header template"):
                main.split_excel_with_template(
                    source,
                    "Data",
                    "Dept",
                    template,
                    out_dir,
                    1,
                    pdf_engine="none",
                    template_mode="template_file",
                    column_mapping={"Worker": "Name"},
                )


class SourceTemplateSplitTests(unittest.TestCase):
    def test_source_template_mode_keeps_selected_sheet_and_matching_key_rows(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            source = tmp_path / "source.xlsx"
            out_dir = tmp_path / "out"

            wb = Workbook()
            ws = wb.active
            ws.title = "Data"
            ws.column_dimensions["A"].width = 22
            ws.append(["Dept", "Name", "Amount"])
            ws.append(["A", "Alice", 10])
            ws.append(["B", "Bob", 20])
            ws.append(["A", "Ana", 30])
            other = wb.create_sheet("Other")
            other["A1"] = "Should not be copied"
            wb.save(source)

            main.split_excel_with_template(
                source,
                "Data",
                "Dept",
                source,
                out_dir,
                1,
                pdf_engine="none",
                template_mode="source_template",
            )

            output = load_workbook(out_dir / "A.xlsx", data_only=True)
            self.assertEqual(output.sheetnames, ["Data"])
            ws_out = output["Data"]
            self.assertEqual(ws_out.column_dimensions["A"].width, 22)
            self.assertEqual([cell.value for cell in ws_out[1]], ["Dept", "Name", "Amount"])
            self.assertEqual(ws_out.max_row, 3)
            self.assertEqual([ws_out["A2"].value, ws_out["B2"].value], ["A", "Alice"])
            self.assertEqual([ws_out["A3"].value, ws_out["B3"].value], ["A", "Ana"])

    def test_pdf_output_type_removes_intermediate_workbook_after_export(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            source = tmp_path / "source.xlsx"
            out_dir = tmp_path / "out"

            wb = Workbook()
            ws = wb.active
            ws.title = "Data"
            ws.append(["Dept", "Name"])
            ws.append(["A", "Alice"])
            wb.save(source)

            original_export = main.export_pdf_via_lo

            def fake_export(xlsx_path, soffice_path=None):
                xlsx_path.with_suffix(".pdf").write_bytes(b"%PDF-1.4\n")

            try:
                main.export_pdf_via_lo = fake_export
                main.split_excel_with_template(
                    source,
                    "Data",
                    "Dept",
                    source,
                    out_dir,
                    1,
                    pdf_engine="libreoffice",
                    template_mode="source_template",
                    output_file_type=main.OUTPUT_TYPE_PDF,
                )
            finally:
                main.export_pdf_via_lo = original_export

            self.assertTrue((out_dir / "A.pdf").exists())
            self.assertFalse((out_dir / "A.xlsx").exists())

    def test_excel_and_pdf_output_type_keeps_workbook_and_pdf(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            source = tmp_path / "source.xlsx"
            out_dir = tmp_path / "out"

            wb = Workbook()
            ws = wb.active
            ws.title = "Data"
            ws.append(["Dept", "Name"])
            ws.append(["A", "Alice"])
            wb.save(source)

            original_export = main.export_pdf_via_lo

            def fake_export(xlsx_path, soffice_path=None):
                xlsx_path.with_suffix(".pdf").write_bytes(b"%PDF-1.4\n")

            try:
                main.export_pdf_via_lo = fake_export
                main.split_excel_with_template(
                    source,
                    "Data",
                    "Dept",
                    source,
                    out_dir,
                    1,
                    pdf_engine="libreoffice",
                    template_mode="source_template",
                    output_file_type=main.OUTPUT_TYPE_EXCEL_AND_PDF,
                )
            finally:
                main.export_pdf_via_lo = original_export

            self.assertTrue((out_dir / "A.xlsx").exists())
            self.assertTrue((out_dir / "A.pdf").exists())

    def test_split_returns_manifest_for_excel_and_pdf_outputs(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            source = tmp_path / "source.xlsx"
            out_dir = tmp_path / "out"
            wb = Workbook()
            ws = wb.active
            ws.title = "Data"
            ws.append(["Dept", "Name"])
            ws.append(["A", "Alice"])
            wb.save(source)

            original_export = main.export_pdf_via_lo

            def fake_export(xlsx_path, soffice_path=None):
                xlsx_path.with_suffix(".pdf").write_bytes(b"%PDF-1.4\n")

            try:
                main.export_pdf_via_lo = fake_export
                results = main.split_excel_with_template(
                    source,
                    "Data",
                    "Dept",
                    source,
                    out_dir,
                    1,
                    pdf_engine="libreoffice",
                    template_mode="source_template",
                    output_file_type=main.OUTPUT_TYPE_EXCEL_AND_PDF,
                )
            finally:
                main.export_pdf_via_lo = original_export

            self.assertEqual(len(results), 1)
            self.assertEqual(results[0].key, "A")
            self.assertEqual(results[0].excel_path, out_dir / "A.xlsx")
            self.assertEqual(results[0].pdf_path, out_dir / "A.pdf")
            self.assertEqual(results[0].output_file_type, main.OUTPUT_TYPE_EXCEL_AND_PDF)


class SplitControlTests(unittest.TestCase):
    def make_source_workbook(self, path: Path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(["Dept", "Name"])
        ws.append(["A", "Alice"])
        ws.append(["B", "Bob"])
        ws.append(["C", "Cara"])
        ws.append(["A", "Ana"])
        wb.save(path)

    def test_selected_keys_limits_generated_files(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            source = tmp_path / "source.xlsx"
            out_dir = tmp_path / "out"
            self.make_source_workbook(source)

            results = main.split_excel_with_template(
                source, "Data", "Dept", source, out_dir, 1,
                pdf_engine="none", template_mode="source_template",
                output_file_type=main.OUTPUT_TYPE_EXCEL,
                selected_keys={"A", "C"},
            )

            self.assertTrue((out_dir / "A.xlsx").exists())
            self.assertTrue((out_dir / "C.xlsx").exists())
            self.assertFalse((out_dir / "B.xlsx").exists())
            self.assertEqual({result.key for result in results}, {"A", "C"})

    def test_selected_keys_none_generates_all(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            source = tmp_path / "source.xlsx"
            out_dir = tmp_path / "out"
            self.make_source_workbook(source)

            results = main.split_excel_with_template(
                source, "Data", "Dept", source, out_dir, 1,
                pdf_engine="none", template_mode="source_template",
                output_file_type=main.OUTPUT_TYPE_EXCEL,
                selected_keys=None,
            )

            self.assertEqual({result.key for result in results}, {"A", "B", "C"})
            self.assertTrue((out_dir / "B.xlsx").exists())

    def test_stop_requested_halts_after_first_key(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            source = tmp_path / "source.xlsx"
            out_dir = tmp_path / "out"
            self.make_source_workbook(source)

            calls = {"count": 0}

            def stop():
                # Allow the first iteration, then request stop.
                should_stop = calls["count"] >= 1
                calls["count"] += 1
                return should_stop

            results = main.split_excel_with_template(
                source, "Data", "Dept", source, out_dir, 1,
                pdf_engine="none", template_mode="source_template",
                output_file_type=main.OUTPUT_TYPE_EXCEL,
                stop_requested=stop,
            )

            self.assertEqual(len(results), 1)
            self.assertEqual(results[0].key, "A")

    def test_read_key_values_returns_ordered_unique_strings(self):
        with tempfile.TemporaryDirectory() as tmp:
            source = Path(tmp) / "source.xlsx"
            self.make_source_workbook(source)

            values = main.read_key_values(source, "Data", "Dept", 1)

            self.assertEqual(values, ["A", "B", "C"])

    def test_debug_messages_suppressed_unless_verbose(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            source = tmp_path / "source.xlsx"
            out_dir = tmp_path / "out"
            self.make_source_workbook(source)

            quiet_messages = []
            main.split_excel_with_template(
                source, "Data", "Dept", source, out_dir, 1,
                pdf_engine="none", template_mode="source_template",
                output_file_type=main.OUTPUT_TYPE_EXCEL,
                status_cb=quiet_messages.append, verbose=False,
            )
            self.assertFalse(any(msg.startswith("Debug:") for msg in quiet_messages))

            verbose_messages = []
            main.split_excel_with_template(
                source, "Data", "Dept", source, out_dir, 1,
                pdf_engine="none", template_mode="source_template",
                output_file_type=main.OUTPUT_TYPE_EXCEL,
                status_cb=verbose_messages.append, verbose=True,
            )
            self.assertTrue(any(msg.startswith("Debug:") for msg in verbose_messages))

    def test_source_template_preserves_only_matching_rows_after_optimization(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            source = tmp_path / "source.xlsx"
            out_dir = tmp_path / "out"
            self.make_source_workbook(source)

            main.split_excel_with_template(
                source, "Data", "Dept", source, out_dir, 1,
                pdf_engine="none", template_mode="source_template",
                output_file_type=main.OUTPUT_TYPE_EXCEL,
            )

            wb = load_workbook(out_dir / "A.xlsx")
            ws = wb.active
            rows = [
                [cell.value for cell in row]
                for row in ws.iter_rows(values_only=False)
            ]
            self.assertEqual(rows[0], [cell for cell in ["Dept", "Name"]])
            data_rows = [[r[0], r[1]] for r in rows[1:]]
            self.assertEqual(data_rows, [["A", "Alice"], ["A", "Ana"]])

    def test_template_file_applies_template_styles_to_data_rows(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            source = tmp_path / "source.xlsx"
            template = tmp_path / "template.xlsx"
            out_dir = tmp_path / "out"

            wb = Workbook()
            ws = wb.active
            ws.title = "Data"
            ws.append(["Dept", "Amount"])
            ws.append(["A", 10])
            ws.append(["A", 20])
            wb.save(source)

            tpl = Workbook()
            tws = tpl.active
            tws.title = "Template"
            tws.append(["Dept", "Amount"])
            tws.cell(row=2, column=2).number_format = "#,##0.00"
            tpl.save(template)

            main.split_excel_with_template(
                source, "Data", "Dept", template, out_dir, 1,
                pdf_engine="none", template_mode="template_file",
                output_file_type=main.OUTPUT_TYPE_EXCEL,
                column_mapping={"Dept": "Dept", "Amount": "Amount"},
            )

            out_wb = load_workbook(out_dir / "A.xlsx")
            out_ws = out_wb.active
            # Second data row should inherit the template data-row number format.
            self.assertEqual(out_ws.cell(row=3, column=2).number_format, "#,##0.00")


if __name__ == "__main__":
    unittest.main()
