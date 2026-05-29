import os
from contextlib import redirect_stdout
from io import StringIO
from pathlib import Path
import tempfile
import unittest

from openpyxl import Workbook


os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
if os.name == "nt" and os.path.isdir(r"C:\Windows\Fonts"):
    os.environ.setdefault("QT_QPA_FONTDIR", r"C:\Windows\Fonts")

try:
    from PySide6.QtCore import QSettings
    from PySide6.QtWidgets import QApplication
    from qfluentwidgets import Theme, isDarkTheme, qconfig
    with redirect_stdout(StringIO()):
        import main
except ModuleNotFoundError as exc:
    raise unittest.SkipTest(f"GUI dependencies are not installed: {exc.name}")


class UISmokeTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.app = QApplication.instance() or QApplication([])

    def make_settings(self, path: Path):
        settings = QSettings(str(path), QSettings.IniFormat)
        settings.clear()
        return settings

    def test_split_app_constructs_with_installed_widgets(self):
        with tempfile.TemporaryDirectory() as tmp:
            window = main.SplitApp(settings=self.make_settings(Path(tmp) / "settings.ini"))
            self.addCleanup(window.deleteLater)

            self.assertEqual(window.windowTitle(), "Excel Splitter")
            self.assertTrue(hasattr(window, "cmb_template_mode"))
            self.assertTrue(hasattr(window, "mapping_card"))
            self.assertTrue(hasattr(window, "btn_auto_map"))

            modes = [
                window.cmb_template_mode.itemText(index)
                for index in range(window.cmb_template_mode.count())
            ]
            self.assertEqual(modes, ["Use Template File", "Use Source as Template"])

    def test_split_app_uses_compact_dashboard_layout(self):
        with tempfile.TemporaryDirectory() as tmp:
            window = main.SplitApp(settings=self.make_settings(Path(tmp) / "settings.ini"))
            self.addCleanup(window.deleteLater)

            self.assertTrue(hasattr(window, "workflow_rail"))
            self.assertTrue(hasattr(window, "main_panel_layout"))
            self.assertTrue(hasattr(window, "footer_bar"))

            self.assertLessEqual(window.edit_source.maximumWidth(), 520)
            self.assertLessEqual(window.edit_template.maximumWidth(), 520)
            self.assertLessEqual(window.edit_outdir.maximumWidth(), 520)
            self.assertLessEqual(window.edit_lo_path.maximumWidth(), 360)
            self.assertLessEqual(window.edit_prefix.maximumWidth(), 180)
            self.assertLessEqual(window.edit_suffix.maximumWidth(), 180)

    def test_template_mode_hides_template_file_controls_for_source_template(self):
        with tempfile.TemporaryDirectory() as tmp:
            window = main.SplitApp(settings=self.make_settings(Path(tmp) / "settings.ini"))
            self.addCleanup(window.deleteLater)

            self.assertFalse(window.edit_template.isHidden())
            self.assertFalse(window.btn_browse_template.isHidden())

            window.cmb_template_mode.setCurrentIndex(
                window.cmb_template_mode.findText("Use Source as Template")
            )
            window.on_template_mode_changed()

            self.assertTrue(window.edit_template.isHidden())
            self.assertTrue(window.btn_browse_template.isHidden())
            self.assertTrue(window.mapping_card.isHidden())

    def test_mapping_status_updates_when_user_selects_source_column(self):
        with tempfile.TemporaryDirectory() as tmp:
            window = main.SplitApp(settings=self.make_settings(Path(tmp) / "settings.ini"))
            self.addCleanup(window.deleteLater)

            window.source_headers = ["Name"]
            window.template_headers = ["Worker"]
            window.render_mapping_rows({})

            self.assertTrue(hasattr(window, "mapping_status_labels"))
            self.assertEqual(window.mapping_status_labels["Worker"].text(), "Missing")

            window.mapping_combos["Worker"].setCurrentIndex(
                window.mapping_combos["Worker"].findText("Name")
            )

            self.assertEqual(window.mapping_status_labels["Worker"].text(), "Mapped")

    def test_dashboard_uses_native_theme_mode_with_matching_surfaces(self):
        with tempfile.TemporaryDirectory() as tmp:
            window = main.SplitApp(settings=self.make_settings(Path(tmp) / "settings.ini"))
            self.addCleanup(window.deleteLater)

            self.assertEqual(window.objectName(), "appRoot")
            self.assertEqual(window.workflow_rail.objectName(), "workflowRail")
            self.assertEqual(qconfig.themeMode.value, Theme.AUTO)
            if isDarkTheme():
                self.assertIn("#202020", window.styleSheet())
                self.assertNotIn("#f5f7fb", window.styleSheet())
            else:
                self.assertIn("#f5f7fb", window.styleSheet())
                self.assertNotIn("#202020", window.styleSheet())

    def test_field_action_buttons_share_input_height_contract(self):
        with tempfile.TemporaryDirectory() as tmp:
            window = main.SplitApp(settings=self.make_settings(Path(tmp) / "settings.ini"))
            self.addCleanup(window.deleteLater)

            self.assertTrue(hasattr(window, "field_action_buttons"))
            self.assertGreaterEqual(len(window.field_action_buttons), 6)
            for button in window.field_action_buttons:
                self.assertEqual(button.minimumHeight(), main.FIELD_CONTROL_HEIGHT)
                self.assertEqual(button.maximumHeight(), main.FIELD_CONTROL_HEIGHT)

    def test_libreoffice_path_row_only_shows_for_libreoffice_pdf_engine(self):
        with tempfile.TemporaryDirectory() as tmp:
            window = main.SplitApp(settings=self.make_settings(Path(tmp) / "settings.ini"))
            self.addCleanup(window.deleteLater)

            window.cmb_output_type.setCurrentIndex(window.cmb_output_type.findText("PDF"))
            window.on_output_type_changed()

            self.assertTrue(hasattr(window, "lo_path_row_widget"))
            self.assertTrue(window.lo_path_row_widget.isHidden())

            window.cmb_pdf_engine.setCurrentIndex(window.cmb_pdf_engine.findText("libreoffice"))
            window.on_pdf_engine_changed()

            self.assertFalse(window.lo_path_row_widget.isHidden())

    def test_header_rows_are_independent_and_output_type_controls_preview(self):
        with tempfile.TemporaryDirectory() as tmp:
            window = main.SplitApp(settings=self.make_settings(Path(tmp) / "settings.ini"))
            self.addCleanup(window.deleteLater)

            self.assertTrue(hasattr(window, "spin_source_header_rows"))
            self.assertTrue(hasattr(window, "spin_template_header_rows"))
            self.assertTrue(hasattr(window, "btn_detect_source_header"))
            self.assertTrue(hasattr(window, "btn_detect_template_header"))
            self.assertTrue(hasattr(window, "cmb_output_type"))
            self.assertTrue(hasattr(window, "lbl_filename_preview"))

            output_types = [
                window.cmb_output_type.itemText(index)
                for index in range(window.cmb_output_type.count())
            ]
            self.assertEqual(output_types, ["Excel", "PDF", "Excel + PDF"])
            self.assertEqual(window.cmb_output_type.currentText(), "Excel")
            self.assertTrue(window.cmb_pdf_engine.isHidden())

            window.cmb_key.clear()
            window.cmb_key.addItem("Dept")
            window.edit_prefix.setText("PRE")
            window.edit_suffix.setText("SUF")
            window.update_filename_preview()
            self.assertIn("PRE <Dept value> SUF.xlsx", window.lbl_filename_preview.text())

            window.cmb_output_type.setCurrentIndex(window.cmb_output_type.findText("PDF"))
            window.on_output_type_changed()
            self.assertFalse(window.cmb_pdf_engine.isHidden())
            self.assertIn("PRE <Dept value> SUF.pdf", window.lbl_filename_preview.text())

            window.cmb_output_type.setCurrentIndex(window.cmb_output_type.findText("Excel + PDF"))
            window.on_output_type_changed()
            self.assertFalse(window.cmb_pdf_engine.isHidden())
            self.assertIn("PRE <Dept value> SUF.xlsx + PDF", window.lbl_filename_preview.text())

    def test_source_refresh_controls_are_icon_buttons(self):
        with tempfile.TemporaryDirectory() as tmp:
            window = main.SplitApp(settings=self.make_settings(Path(tmp) / "settings.ini"))
            self.addCleanup(window.deleteLater)

            self.assertIsInstance(window.btn_load_sheets, main.ToolButton)
            self.assertIsInstance(window.btn_load_headers, main.ToolButton)
            self.assertEqual(window.btn_load_sheets.toolTip(), "Load Sheets")
            self.assertEqual(window.btn_load_headers.toolTip(), "Load Headers")

    def test_refresh_source_options_populates_sheets_and_key_headers(self):
        with tempfile.TemporaryDirectory() as tmp:
            source = Path(tmp) / "source.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Data"
            ws.append(["Dept", "Name"])
            ws.append(["A", "Alice"])
            wb.save(source)

            window = main.SplitApp(settings=self.make_settings(Path(tmp) / "settings.ini"))
            self.addCleanup(window.deleteLater)
            window.edit_source.setText(str(source))

            window.refresh_source_options()

            self.assertEqual(window.cmb_sheet.currentText(), "Data")
            self.assertEqual(window.spin_source_header_rows.value(), 1)
            key_values = [
                window.cmb_key.itemText(index)
                for index in range(window.cmb_key.count())
            ]
            self.assertEqual(key_values, ["Dept", "Name", "1", "2"])

    def test_template_header_row_sits_in_template_workbook_row(self):
        with tempfile.TemporaryDirectory() as tmp:
            window = main.SplitApp(settings=self.make_settings(Path(tmp) / "settings.ini"))
            self.addCleanup(window.deleteLater)

            row_layout = window.template_file_row_widget.layout()
            self.assertGreaterEqual(row_layout.indexOf(window.template_header_row_widget), 0)

    def test_footer_progress_is_hidden_until_generation_starts(self):
        with tempfile.TemporaryDirectory() as tmp:
            window = main.SplitApp(settings=self.make_settings(Path(tmp) / "settings.ini"))
            self.addCleanup(window.deleteLater)

            self.assertTrue(window.progress_bar.isHidden())
            window.set_busy(True)
            self.assertFalse(window.progress_bar.isHidden())

    def test_mail_merge_button_is_available_without_split_results(self):
        with tempfile.TemporaryDirectory() as tmp:
            window = main.SplitApp(settings=self.make_settings(Path(tmp) / "settings.ini"))
            self.addCleanup(window.deleteLater)

            self.assertTrue(hasattr(window, "btn_mail_merge"))
            self.assertFalse(window.btn_mail_merge.isHidden())

            window.show_mail_merge_panel()

            self.assertFalse(window.mail_merge_card.isHidden())
            self.assertIn("No split files", window.lbl_mail_merge_summary.text())

            window.current_split_results = [
                main.SplitResult(key="A", excel_path=Path(tmp) / "A.xlsx", output_file_type=main.OUTPUT_TYPE_EXCEL)
            ]
            window.update_mail_merge_entry_state()

            self.assertFalse(window.btn_mail_merge.isHidden())

    def test_show_mail_merge_panel_reveals_panel_with_loaded_count(self):
        with tempfile.TemporaryDirectory() as tmp:
            window = main.SplitApp(settings=self.make_settings(Path(tmp) / "settings.ini"))
            self.addCleanup(window.deleteLater)
            window.current_split_results = [
                main.SplitResult(key="A", excel_path=Path(tmp) / "A.xlsx", output_file_type=main.OUTPUT_TYPE_EXCEL)
            ]

            window.show_mail_merge_panel()

            self.assertFalse(window.mail_merge_card.isHidden())
            self.assertIn("1 split file", window.lbl_mail_merge_summary.text())

    def test_mail_merge_recipient_controls_exist(self):
        with tempfile.TemporaryDirectory() as tmp:
            window = main.SplitApp(settings=self.make_settings(Path(tmp) / "settings.ini"))
            self.addCleanup(window.deleteLater)

            self.assertTrue(hasattr(window, "edit_recipient_path"))
            self.assertTrue(hasattr(window, "cmb_recipient_sheet"))
            self.assertTrue(hasattr(window, "spin_recipient_header_row"))
            self.assertTrue(hasattr(window, "cmb_recipient_key"))
            self.assertTrue(hasattr(window, "cmb_recipient_to"))
            self.assertTrue(hasattr(window, "cmb_recipient_cc"))
            self.assertTrue(hasattr(window, "cmb_recipient_bcc"))

    def test_mail_merge_layout_fits_default_window_width(self):
        with tempfile.TemporaryDirectory() as tmp:
            window = main.SplitApp(settings=self.make_settings(Path(tmp) / "settings.ini"))
            self.addCleanup(window.deleteLater)

            window.show_mail_merge_panel()

            self.assertLessEqual(window.mail_merge_card.sizeHint().width(), 760)

    def test_mail_merge_preview_carousel_moves_between_jobs(self):
        with tempfile.TemporaryDirectory() as tmp:
            window = main.SplitApp(settings=self.make_settings(Path(tmp) / "settings.ini"))
            self.addCleanup(window.deleteLater)
            window.current_mail_jobs = [
                main.EmailJob("A", ["a@example.com"], [], [], "Subject A", "Body A", False, []),
                main.EmailJob("B", ["b@example.com"], [], [], "Subject B", "Body B", False, []),
            ]
            window.current_preview_index = 0

            window.render_mail_preview()
            self.assertIn("1 / 2", window.lbl_mail_preview_count.text())
            self.assertIn("Subject A", window.lbl_mail_preview_subject.text())

            window.next_mail_preview()

            self.assertIn("2 / 2", window.lbl_mail_preview_count.text())
            self.assertIn("Subject B", window.lbl_mail_preview_subject.text())

    def test_mail_merge_send_disabled_when_validation_fails(self):
        with tempfile.TemporaryDirectory() as tmp:
            window = main.SplitApp(settings=self.make_settings(Path(tmp) / "settings.ini"))
            self.addCleanup(window.deleteLater)
            window.current_mail_jobs = [
                main.EmailJob("A", [], [], [], "", "", False, [], validation_errors=["Required To is empty for key A"])
            ]

            window.render_mail_preview()

            self.assertFalse(window.btn_send_mail_merge.isEnabled())
            self.assertIn("1 issue", window.lbl_mail_validation_summary.text())

    def test_current_send_timing_reads_delay_and_throttle_controls(self):
        with tempfile.TemporaryDirectory() as tmp:
            window = main.SplitApp(settings=self.make_settings(Path(tmp) / "settings.ini"))
            self.addCleanup(window.deleteLater)
            window.chk_delay_delivery.setChecked(True)
            window.spin_delay_minutes.setValue(7)
            window.chk_throttle.setChecked(True)
            window.spin_throttle_seconds.setValue(3)

            timing = window.current_send_timing()

            self.assertTrue(timing.delay_delivery_enabled)
            self.assertEqual(timing.delay_delivery_minutes, 7)
            self.assertTrue(timing.throttle_enabled)
            self.assertEqual(timing.throttle_seconds, 3)

    def test_mail_merge_attachment_options_follow_split_outputs(self):
        with tempfile.TemporaryDirectory() as tmp:
            window = main.SplitApp(settings=self.make_settings(Path(tmp) / "settings.ini"))
            self.addCleanup(window.deleteLater)

            window.current_split_results = [
                main.SplitResult(key="A", excel_path=None, pdf_path=Path(tmp) / "A.pdf", output_file_type=main.OUTPUT_TYPE_PDF)
            ]
            window.show_mail_merge_panel()

            self.assertFalse(window.chk_attach_excel.isEnabled())
            self.assertFalse(window.chk_attach_excel.isChecked())
            self.assertTrue(window.chk_attach_pdf.isEnabled())
            self.assertTrue(window.chk_attach_pdf.isChecked())

    def test_mail_merge_split_folder_controls_exist(self):
        with tempfile.TemporaryDirectory() as tmp:
            window = main.SplitApp(settings=self.make_settings(Path(tmp) / "settings.ini"))
            self.addCleanup(window.deleteLater)

            self.assertTrue(hasattr(window, "edit_split_folder"))
            self.assertTrue(hasattr(window, "btn_browse_split_folder"))
            self.assertTrue(hasattr(window, "edit_detect_prefix"))
            self.assertTrue(hasattr(window, "edit_detect_suffix"))
            self.assertTrue(hasattr(window, "btn_scan_split_folder"))

    def test_scan_split_folder_loads_paired_split_results(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            excel_a = tmp_path / "Report A Final.xlsx"
            pdf_a = tmp_path / "Report A Final.pdf"
            excel_b = tmp_path / "Report B Final.xlsx"
            
            wb = Workbook()
            wb.save(excel_a)
            pdf_a.write_bytes(b"%PDF-1.4\n")
            wb2 = Workbook()
            wb2.save(excel_b)

            window = main.SplitApp(settings=self.make_settings(tmp_path / "settings.ini"))
            self.addCleanup(window.deleteLater)
            
            window.edit_split_folder.setText(str(tmp_path))
            window.edit_detect_prefix.setText("Report")
            window.edit_detect_suffix.setText("Final")
            window.scan_split_folder()

            self.assertEqual(len(window.current_split_results), 2)
            keys = sorted([r.key for r in window.current_split_results])
            self.assertEqual(keys, ["A", "B"])
            
            result_a = next(r for r in window.current_split_results if r.key == "A")
            self.assertIsNotNone(result_a.excel_path)
            self.assertIsNotNone(result_a.pdf_path)
            self.assertEqual(result_a.output_file_type, main.OUTPUT_TYPE_EXCEL_AND_PDF)
            
            self.assertIn("2 split files", window.lbl_mail_merge_summary.text())


if __name__ == "__main__":
    unittest.main()
