from pathlib import Path
import tempfile
import unittest
from datetime import datetime

from openpyxl import Workbook

import mail_merge


class MailMergeCoreTests(unittest.TestCase):
    def test_parse_email_list_splits_semicolon_values(self):
        self.assertEqual(
            mail_merge.parse_email_list("a@example.com; b@example.com ;"),
            ["a@example.com", "b@example.com"],
        )

    def test_parse_email_list_handles_empty_values(self):
        self.assertEqual(mail_merge.parse_email_list(" ; "), [])
        self.assertEqual(mail_merge.parse_email_list(None), [])

    def test_is_valid_email_rejects_obvious_invalid_values(self):
        self.assertTrue(mail_merge.is_valid_email("person@example.com"))
        self.assertFalse(mail_merge.is_valid_email("missing-at.example.com"))
        self.assertFalse(mail_merge.is_valid_email("person@"))
        self.assertFalse(mail_merge.is_valid_email("@example.com"))

    def test_render_placeholders_uses_case_insensitive_keys_and_leaves_unknown(self):
        context = {
            "key": "A",
            "To": "alice@example.com",
            "Dept": "Finance",
        }

        rendered = mail_merge.render_placeholders(
            "Hello {to}, key={KEY}, dept={dept}, keep={missing}",
            context,
        )

        self.assertEqual(
            rendered,
            "Hello alice@example.com, key=A, dept=Finance, keep={missing}",
        )

    def test_read_recipient_headers_from_excel_sheet(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "recipients.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Recipients"
            ws.append(["Report", None, None, None])
            ws.append(["Key", "To", "CC", "Department"])
            wb.save(path)

            headers = mail_merge.read_recipient_headers(path, "Recipients", 2)

            self.assertEqual(headers, ["Key", "To", "CC", "Department"])

    def test_load_recipient_rows_uses_manual_column_mapping(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "recipients.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Recipients"
            ws.append(["Customer", "Email", "Copy", "Team"])
            ws.append(["A", "a@example.com; aa@example.com", "lead@example.com", "Finance"])
            ws.append(["B", "b@example.com", None, "Ops"])
            wb.save(path)

            rows = mail_merge.load_recipient_rows(
                path,
                "Recipients",
                1,
                {
                    "key": "Customer",
                    "to": "Email",
                    "cc": "Copy",
                    "bcc": "",
                },
            )

            self.assertEqual(len(rows), 2)
            self.assertEqual(rows[0].key, "A")
            self.assertEqual(rows[0].to, ["a@example.com", "aa@example.com"])
            self.assertEqual(rows[0].cc, ["lead@example.com"])
            self.assertEqual(rows[0].raw["Team"], "Finance")
            self.assertEqual(rows[1].bcc, [])

    def test_build_email_jobs_renders_content_and_selected_attachments(self):
        with tempfile.TemporaryDirectory() as tmp:
            excel_path = Path(tmp) / "A.xlsx"
            pdf_path = Path(tmp) / "A.pdf"
            excel_path.write_text("excel", encoding="utf-8")
            pdf_path.write_text("pdf", encoding="utf-8")

            jobs, warnings = mail_merge.build_email_jobs(
                split_results=[
                    mail_merge.SplitResult("A", excel_path=excel_path, pdf_path=pdf_path, output_file_type="excel_and_pdf")
                ],
                recipients=[
                    mail_merge.RecipientRow(
                        key="A",
                        to=["a@example.com"],
                        cc=[],
                        bcc=[],
                        raw={"Department": "Finance"},
                    )
                ],
                template=mail_merge.EmailTemplate(
                    subject="Report {key} {Department}",
                    body="Hello {to}, file {excel_file}",
                    is_html=False,
                ),
                attachments=mail_merge.AttachmentSelection(attach_excel=True, attach_pdf=True),
            )

            self.assertEqual(warnings, [])
            self.assertEqual(len(jobs), 1)
            self.assertEqual(jobs[0].subject, "Report A Finance")
            self.assertEqual(jobs[0].body, f"Hello a@example.com, file {excel_path}")
            self.assertEqual(jobs[0].attachments, [excel_path, pdf_path])
            self.assertTrue(jobs[0].is_valid)

    def test_build_email_jobs_reports_missing_mapping_and_ignores_extra_mapping_rows(self):
        split_results = [mail_merge.SplitResult("A")]
        recipients = [
            mail_merge.RecipientRow(key="B", to=["b@example.com"], raw={}),
        ]

        jobs, warnings = mail_merge.build_email_jobs(
            split_results=split_results,
            recipients=recipients,
            template=mail_merge.EmailTemplate(subject="Subject", body="Body"),
            attachments=mail_merge.AttachmentSelection(attach_excel=False, attach_pdf=False),
        )

        self.assertEqual(len(jobs), 1)
        self.assertIn("No recipient mapping for key A", jobs[0].validation_errors)
        self.assertEqual(warnings, ["Recipient mapping key B does not match a generated split file"])

    def test_build_email_jobs_without_split_results_uses_recipient_rows(self):
        jobs, warnings = mail_merge.build_email_jobs(
            split_results=[],
            recipients=[
                mail_merge.RecipientRow(
                    key="A",
                    to=["a@example.com"],
                    raw={"Department": "Finance"},
                ),
            ],
            template=mail_merge.EmailTemplate(
                subject="Report {key} {Department}",
                body="Hello {to}",
            ),
            attachments=mail_merge.AttachmentSelection(attach_excel=False, attach_pdf=False),
        )

        self.assertEqual(warnings, [])
        self.assertEqual(len(jobs), 1)
        self.assertEqual(jobs[0].key, "A")
        self.assertEqual(jobs[0].subject, "Report A Finance")
        self.assertEqual(jobs[0].body, "Hello a@example.com")
        self.assertEqual(jobs[0].attachments, [])
        self.assertTrue(jobs[0].is_valid)

    def test_build_email_jobs_validates_email_subject_body_and_attachments(self):
        with tempfile.TemporaryDirectory() as tmp:
            missing_pdf = Path(tmp) / "A.pdf"
            jobs, warnings = mail_merge.build_email_jobs(
                split_results=[mail_merge.SplitResult("A", pdf_path=missing_pdf, output_file_type="pdf")],
                recipients=[mail_merge.RecipientRow(key="A", to=["bad-email"], raw={})],
                template=mail_merge.EmailTemplate(subject=" ", body=" "),
                attachments=mail_merge.AttachmentSelection(attach_excel=False, attach_pdf=True),
            )

            self.assertEqual(warnings, [])
            self.assertIn("Invalid To email: bad-email", jobs[0].validation_errors)
            self.assertIn("Selected PDF attachment is missing for key A", jobs[0].validation_errors)
            self.assertIn("Rendered subject is empty for key A", jobs[0].validation_errors)
            self.assertIn("Rendered body is empty for key A", jobs[0].validation_errors)

    def test_fake_provider_records_sent_jobs(self):
        provider = mail_merge.FakeMailProvider()
        job = mail_merge.EmailJob(
            key="A",
            to=["a@example.com"],
            cc=[],
            bcc=[],
            subject="Subject",
            body="Body",
            is_html=False,
            attachments=[],
        )

        result = provider.send(job, mail_merge.SendTimingOptions())

        self.assertEqual(result.status, "sent")
        self.assertEqual(provider.sent_jobs, [job])

    def test_outlook_provider_sets_recipients_attachments_and_deferred_delivery(self):
        calls = []

        class FakeAttachments:
            def Add(self, path):
                calls.append(("attach", path))

        class FakeRecipients:
            def Add(self, value):
                calls.append(("recipient", value))
                recipient = type("Recipient", (), {})()
                recipient.Type = None
                return recipient

        class FakeMessage:
            def __init__(self):
                self.To = ""
                self.CC = ""
                self.BCC = ""
                self.Subject = ""
                self.Body = ""
                self.HTMLBody = ""
                self.DeferredDeliveryTime = None
                self.Attachments = FakeAttachments()
                self.Recipients = FakeRecipients()

            def Send(self):
                calls.append(("send", self.Subject, self.DeferredDeliveryTime))

        class FakeOutlook:
            def CreateItem(self, item_type):
                calls.append(("create", item_type))
                return FakeMessage()

        now = datetime(2026, 5, 27, 15, 0, 0)
        provider = mail_merge.OutlookMailProvider(
            dispatcher=lambda name: FakeOutlook(),
            now_fn=lambda: now,
        )
        attachment = Path("C:/tmp/A.pdf")
        job = mail_merge.EmailJob(
            key="A",
            to=["a@example.com"],
            cc=["c@example.com"],
            bcc=["b@example.com"],
            subject="Subject A",
            body="<p>Body</p>",
            is_html=True,
            attachments=[attachment],
        )

        result = provider.send(
            job,
            mail_merge.SendTimingOptions(
                delay_delivery_enabled=True,
                delay_delivery_minutes=5,
                throttle_enabled=False,
                throttle_seconds=0,
            ),
        )

        self.assertEqual(result.status, "sent")
        self.assertIn(("create", 0), calls)
        self.assertIn(("attach", str(attachment)), calls)
        self.assertIn(("send", "Subject A", now + mail_merge.timedelta(minutes=5)), calls)

    def test_send_jobs_throttles_between_jobs_and_can_cancel(self):
        provider = mail_merge.FakeMailProvider()
        sleeps = []
        statuses = []
        jobs = [
            mail_merge.EmailJob("A", ["a@example.com"], [], [], "A", "Body", False, []),
            mail_merge.EmailJob("B", ["b@example.com"], [], [], "B", "Body", False, []),
        ]

        results = mail_merge.send_jobs(
            jobs,
            provider,
            mail_merge.SendTimingOptions(
                delay_delivery_enabled=False,
                delay_delivery_minutes=0,
                throttle_enabled=True,
                throttle_seconds=5,
            ),
            status_cb=statuses.append,
            sleep_fn=sleeps.append,
        )

        self.assertEqual([result.key for result in results], ["A", "B"])
        self.assertEqual(sleeps, [5])
        self.assertEqual(statuses, ["Sending 1/2: A", "Sending 2/2: B"])

    def test_detect_key_from_filename_with_prefix_and_suffix_both_present(self):
        result = mail_merge.detect_key_from_filename("Report 12345 Final", prefix="Report", suffix="Final")
        self.assertEqual(result, "12345")

    def test_detect_key_from_filename_with_no_prefix_suffix(self):
        result = mail_merge.detect_key_from_filename("12345", prefix="", suffix="")
        self.assertEqual(result, "12345")

    def test_detect_key_from_filename_with_prefix_only(self):
        result = mail_merge.detect_key_from_filename("Report 12345", prefix="Report", suffix="")
        self.assertEqual(result, "12345")

    def test_detect_key_from_filename_with_suffix_only(self):
        result = mail_merge.detect_key_from_filename("12345 Final", prefix="", suffix="Final")
        self.assertEqual(result, "12345")

    def test_detect_key_from_filename_prefix_not_matching_but_suffix_matches(self):
        result = mail_merge.detect_key_from_filename("12345 Final", prefix="Report", suffix="Final")
        self.assertEqual(result, "12345")

    def test_detect_key_from_filename_suffix_not_matching_but_prefix_matches(self):
        result = mail_merge.detect_key_from_filename("Report 12345", prefix="Report", suffix="Final")
        self.assertEqual(result, "12345")

    def test_detect_key_from_filename_key_containing_spaces(self):
        result = mail_merge.detect_key_from_filename("Report North East Final", prefix="Report", suffix="Final")
        self.assertEqual(result, "North East")

    def test_detect_key_from_filename_returns_original_if_empty_after_strip(self):
        result = mail_merge.detect_key_from_filename("Report Final", prefix="Report", suffix="Final")
        self.assertEqual(result, "Report Final")

    def test_detect_key_from_filename_case_sensitive_matching(self):
        result = mail_merge.detect_key_from_filename("report 12345 final", prefix="Report", suffix="Final")
        self.assertEqual(result, "report 12345 final")

    def test_discover_split_results_pairs_xlsx_and_pdf_same_key(self):
        with tempfile.TemporaryDirectory() as tmp:
            folder = Path(tmp)
            xlsx_path = folder / "12345.xlsx"
            pdf_path = folder / "12345.pdf"
            wb = Workbook()
            wb.save(xlsx_path)
            pdf_path.write_bytes(b"%PDF-1.4\n")

            results = mail_merge.discover_split_results_from_folder(folder, prefix="", suffix="")

            self.assertEqual(len(results), 1)
            self.assertEqual(results[0].key, "12345")
            self.assertEqual(results[0].excel_path, xlsx_path)
            self.assertEqual(results[0].pdf_path, pdf_path)
            self.assertEqual(results[0].output_file_type, "excel_and_pdf")

    def test_discover_split_results_excel_only(self):
        with tempfile.TemporaryDirectory() as tmp:
            folder = Path(tmp)
            xlsx_path = folder / "12345.xlsx"
            wb = Workbook()
            wb.save(xlsx_path)

            results = mail_merge.discover_split_results_from_folder(folder, prefix="", suffix="")

            self.assertEqual(len(results), 1)
            self.assertEqual(results[0].key, "12345")
            self.assertEqual(results[0].excel_path, xlsx_path)
            self.assertIsNone(results[0].pdf_path)
            self.assertEqual(results[0].output_file_type, "excel")

    def test_discover_split_results_pdf_only(self):
        with tempfile.TemporaryDirectory() as tmp:
            folder = Path(tmp)
            pdf_path = folder / "12345.pdf"
            pdf_path.write_bytes(b"%PDF-1.4\n")

            results = mail_merge.discover_split_results_from_folder(folder, prefix="", suffix="")

            self.assertEqual(len(results), 1)
            self.assertEqual(results[0].key, "12345")
            self.assertIsNone(results[0].excel_path)
            self.assertEqual(results[0].pdf_path, pdf_path)
            self.assertEqual(results[0].output_file_type, "pdf")

    def test_discover_split_results_ignores_temp_files(self):
        with tempfile.TemporaryDirectory() as tmp:
            folder = Path(tmp)
            xlsx_path = folder / "12345.xlsx"
            temp_path = folder / "~$12345.xlsx"
            wb = Workbook()
            wb.save(xlsx_path)
            wb.save(temp_path)

            results = mail_merge.discover_split_results_from_folder(folder, prefix="", suffix="")

            self.assertEqual(len(results), 1)
            self.assertEqual(results[0].key, "12345")

    def test_discover_split_results_returns_empty_for_nonexistent_folder(self):
        results = mail_merge.discover_split_results_from_folder(Path("/nonexistent/folder"), prefix="", suffix="")
        self.assertEqual(results, [])

    def test_discover_split_results_with_prefix_suffix_and_sorts_by_key(self):
        with tempfile.TemporaryDirectory() as tmp:
            folder = Path(tmp)
            (folder / "Report 12345 Final.xlsx").write_bytes(b"")
            wb1 = Workbook()
            wb1.save(folder / "Report 12345 Final.xlsx")
            (folder / "Report 12345 Final.pdf").write_bytes(b"%PDF-1.4\n")
            wb2 = Workbook()
            wb2.save(folder / "Report 67890 Final.xlsx")
            (folder / "Report 67890 Final.pdf").write_bytes(b"%PDF-1.4\n")
            wb3 = Workbook()
            wb3.save(folder / "Report AAAAA Final.xlsx")

            results = mail_merge.discover_split_results_from_folder(folder, prefix="Report", suffix="Final")

            self.assertEqual(len(results), 3)
            self.assertEqual(results[0].key, "12345")
            self.assertEqual(results[0].output_file_type, "excel_and_pdf")
            self.assertEqual(results[1].key, "67890")
            self.assertEqual(results[1].output_file_type, "excel_and_pdf")
            self.assertEqual(results[2].key, "AAAAA")
            self.assertEqual(results[2].output_file_type, "excel")

    def test_discover_split_results_handles_duplicate_keys_deterministically(self):
        with tempfile.TemporaryDirectory() as tmp:
            folder = Path(tmp)
            wb1 = Workbook()
            wb1.save(folder / "12345.xlsx")
            wb2 = Workbook()
            wb2.save(folder / "12345_copy.xlsx")

            results = mail_merge.discover_split_results_from_folder(folder, prefix="", suffix="")

            self.assertEqual(len(results), 2)
            keys = [r.key for r in results]
            self.assertIn("12345", keys)
            self.assertIn("12345_copy", keys)

    def test_discover_split_results_recurse_into_subfolders(self):
        with tempfile.TemporaryDirectory() as tmp:
            folder = Path(tmp)
            subfolder = folder / "sub"
            subfolder.mkdir()
            wb1 = Workbook()
            wb1.save(folder / "12345.xlsx")
            wb2 = Workbook()
            wb2.save(subfolder / "67890.xlsx")

            results = mail_merge.discover_split_results_from_folder(folder, prefix="", suffix="", recurse=True)

            self.assertEqual(len(results), 2)
            keys = sorted([r.key for r in results])
            self.assertEqual(keys, ["12345", "67890"])

    def test_discover_split_results_no_recurse_ignores_subfolders(self):
        with tempfile.TemporaryDirectory() as tmp:
            folder = Path(tmp)
            subfolder = folder / "sub"
            subfolder.mkdir()
            wb1 = Workbook()
            wb1.save(folder / "12345.xlsx")
            wb2 = Workbook()
            wb2.save(subfolder / "67890.xlsx")

            results = mail_merge.discover_split_results_from_folder(folder, prefix="", suffix="", recurse=False)

            self.assertEqual(len(results), 1)
            self.assertEqual(results[0].key, "12345")


if __name__ == "__main__":
    unittest.main()
