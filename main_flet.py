# main_flet.py
# Modern Flet-based GUI for Excel splitter with Fluent Design styling
# This is the modernized version of the CustomTkinter application

import os
import re
import shutil
import subprocess
import threading
import asyncio
from pathlib import Path
import configparser
from typing import Optional, List

import flet as ft
from flet import Colors, Icons
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ==== (Opsional) ReportLab untuk PDF pure-Python ====
try:
    from reportlab.lib.pagesizes import A4, LETTER, landscape, portrait
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
    REPORTLAB_AVAILABLE = True
except Exception:
    REPORTLAB_AVAILABLE = False


# ----------------- Helpers (unchanged from original) -----------------
def safe_file_part(s: str) -> str:
    s = "" if s is None else str(s)
    return re.sub(r'[:\\/\?\*\[\]<>|"]', "_", s).strip() or "Key"

def set_print_titles_and_area(ws, header_rows: int, last_col_idx: int, last_data_row: int):
    ws.print_title_rows = f"1:{header_rows}"
    last_col_letter = get_column_letter(last_col_idx if last_col_idx > 0 else 1)
    last_row = last_data_row if last_data_row >= (header_rows + 1) else (header_rows + 1)
    ws.print_area = f"A1:{last_col_letter}{last_row}"

def find_soffice(explicit_path: str | None = None) -> str | None:
    """Cari soffice.exe dari explicit, env, PATH, atau lokasi umum."""
    def _normalize(p):
        if not p:
            return None
        p = str(Path(p))
        if p.lower().endswith("soffice.exe") and Path(p).exists():
            return p
        prog = Path(p) / "soffice.exe"
        if prog.exists():
            return str(prog)
        prog = Path(p) / "program" / "soffice.exe"
        if prog.exists():
            return str(prog)
        return None

    # explicit
    s = _normalize(explicit_path)
    if s: return s
    # env
    s = _normalize(os.environ.get("LIBREOFFICE_PATH"))
    if s: return s
    # PATH
    s = shutil.which("soffice")
    if s: return s
    # lokasi umum
    common = [
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
        r"C:\PortableApps\LibreOfficePortable\App\libreoffice\program\soffice.exe",
        r"D:\PortableApps\LibreOfficePortable\App\libreoffice\program\soffice.exe",
    ]
    for c in common:
        if Path(c).exists():
            return c
    return None

def export_pdf_via_lo(xlsx_path: Path, soffice_path: str | None = None):
    exe = soffice_path or "soffice"
    cmd = [exe, "--headless", "--convert-to", "pdf", "--outdir", str(xlsx_path.parent), str(xlsx_path)]
    subprocess.run(cmd, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)

# ---------- ReportLab (pure-Python) PDF ----------
PAPER_MAP = {9: A4, 1: LETTER}

def _excel_col_width_to_points(w):
    return float(w) * 7.0 * 0.75 if w else 50.0

def _page_setup_from_template(ws):
    ps = ws.page_setup
    paper = PAPER_MAP.get(ps.paperSize, A4)
    orientation = str(getattr(ps, "orientation", "portrait")).lower()
    paper = landscape(paper) if orientation == "landscape" else portrait(paper)
    pm = ws.page_margins
    return paper, {
        "left": (pm.left or 0.7) * inch,
        "right": (pm.right or 0.7) * inch,
        "top": (pm.top or 0.75) * inch,
        "bottom": (pm.bottom or 0.75) * inch,
    }

def export_pdf_pure(group_df: pd.DataFrame, template_path: Path, header_rows: int, pdf_out: Path):
    if not REPORTLAB_AVAILABLE:
        raise RuntimeError("ReportLab belum terpasang. Jalankan: pip install reportlab")

    wb = load_workbook(template_path, read_only=True, data_only=True)
    ws = wb.active

    templ_headers, hidden_cols_idx, col_width_pts = [], set(), []
    empty_streak, c = 0, 1
    while c <= 500 and empty_streak < 5:
        val = ws.cell(row=header_rows, column=c).value
        col_dim = ws.column_dimensions.get(ws.cell(row=1, column=c).column_letter)
        hidden = bool(getattr(col_dim, "hidden", False))
        width = getattr(col_dim, "width", None)
        if val is None and width is None:
            empty_streak += 1
        else:
            empty_streak = 0
        templ_headers.append("" if val is None else str(val))
        if hidden:
            hidden_cols_idx.add(c - 1)
        col_width_pts.append(_excel_col_width_to_points(width))
        c += 1
    wb.close()

    # Map urutan kolom mengikuti header template bila cocok
    if templ_headers and any(h.strip() for h in templ_headers):
        ordered = [h for h in templ_headers if h and h in group_df.columns]
        if ordered:
            df = group_df[ordered].copy()
            keep_idx = [templ_headers.index(h) for h in ordered]
            col_width_pts = [col_width_pts[idx] for idx in keep_idx]
            hidden_cols_idx = {i for i in range(len(keep_idx)) if keep_idx[i] in hidden_cols_idx}
        else:
            df = group_df.copy()
    else:
        df = group_df.copy()

    # Drop kolom hidden
    keep_cols = [i for i in range(len(df.columns)) if i not in hidden_cols_idx] or list(range(len(df.columns)))
    df = df.iloc[:, keep_cols]
    col_width_pts = [col_width_pts[i] for i in keep_cols]

    # Susun data tabel (header + isi)
    table_data = [list(df.columns.astype(str))] + df.fillna("").astype(str).values.tolist()

    # Page setup
    pagesize, margins = _page_setup_from_template(ws)
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
    from reportlab.lib import colors
    doc = SimpleDocTemplate(
        str(pdf_out),
        pagesize=pagesize,
        leftMargin=margins["left"],
        rightMargin=margins["right"],
        topMargin=margins["top"],
        bottomMargin=margins["bottom"],
        title=pdf_out.stem,
    )

    tbl = Table(table_data, colWidths=col_width_pts or None, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTSIZE", (0, 1), (-1, -1), 9),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.25, Colors.black),
        ("LINEABOVE", (0, 0), (-1, 0), 0.75, Colors.black),
        ("BACKGROUND", (0, 0), (-1, 0), Colors.HexColor("#F0F0F0")),
    ]))
    doc.build([tbl])


# ----------------- Split Logic (unchanged) -----------------
def split_excel_with_template(
    source_path: Path, sheet_name: str, key_col, template_path: Path, out_dir: Path,
    header_rows: int, pdf_engine: str = "reportlab", soffice_path: str | None = None,
    prefix: str = "", suffix: str = "", status_cb=None, progress_cb=None
):
    if status_cb is None: status_cb = lambda msg: None
    if progress_cb is None: progress_cb = lambda t, c: None

    if not source_path.exists():
        raise FileNotFoundError(f"Sumber tidak ditemukan: {source_path}")
    if not template_path.exists():
        raise FileNotFoundError(f"Template tidak ditemukan: {template_path}")
    out_dir.mkdir(parents=True, exist_ok=True)

    status_cb("Membaca sumber...")

    # Diagnostic logging
    try:
        file_size = source_path.stat().st_size if source_path.exists() else 0
        status_cb(f"Debug: File path: {source_path}")
        status_cb(f"Debug: File exists: {source_path.exists()}")
        status_cb(f"Debug: File size: {file_size:,} bytes ({file_size/1024/1024:.2f} MB)")
        status_cb(f"Debug: Sheet name: '{sheet_name}'")
        status_cb("Debug: Starting pd.read_excel...")

        # Try reading with timeout and error handling
        import time
        start_time = time.time()

        # First try to read just the header to test file accessibility
        status_cb("Debug: Testing file accessibility...")
        try:
            # Try to read just first few rows to test
            df_test = pd.read_excel(source_path, sheet_name=sheet_name, nrows=5, dtype=object)
            status_cb(f"Debug: Successfully read {len(df_test)} rows for testing")
        except Exception as test_e:
            status_cb(f"Debug: Test read failed: {str(test_e)}")
            raise test_e

        # Now read the full file
        df = pd.read_excel(source_path, sheet_name=sheet_name, dtype=object)

        elapsed = time.time() - start_time
        status_cb(f"Debug: Successfully read {len(df)} rows in {elapsed:.2f} seconds")

    except FileNotFoundError as e:
        status_cb(f"Debug: File not found: {e}")
        raise FileNotFoundError(f"File tidak ditemukan: {source_path}")
    except PermissionError as e:
        status_cb(f"Debug: Permission denied: {e}")
        raise PermissionError(f"Tidak ada akses ke file: {source_path}")
    except Exception as e:
        status_cb(f"Debug: Error reading Excel: {type(e).__name__}: {str(e)}")
        raise e

    # Tentukan kolom kunci
    if isinstance(key_col, int):
        if key_col < 1 or key_col > df.shape[1]:
            raise ValueError("Index kolom kunci di luar jangkauan DataFrame.")
        key_series = df.iloc[:, key_col - 1]
    else:
        if key_col not in df.columns:
            raise ValueError(f"Header kolom kunci '{key_col}' tidak ditemukan.")
        key_series = df[key_col]

    # Debug: Check for categorical data issues
    status_cb(f"Debug: Key column '{key_col}' data type: {key_series.dtype}")
    status_cb(f"Debug: Key column unique values: {len(key_series.unique())}")
    status_cb(f"Debug: Key column has null values: {key_series.isnull().sum()}")

    # Check for categorical columns in the entire DataFrame
    categorical_cols = []
    for col in df.columns:
        if df[col].dtype.name == 'category':
            categorical_cols.append(col)
            status_cb(f"Debug: Found categorical column: {col}")

    if categorical_cols:
        status_cb(f"Debug: Converting {len(categorical_cols)} categorical columns to string...")
        for col in categorical_cols:
            try:
                df[col] = df[col].astype(str)
                status_cb(f"Debug: Converted {col} to string")
            except Exception as conv_e:
                status_cb(f"Debug: Failed to convert {col}: {conv_e}")

    # Selaraskan urutan kolom ke header template jika cocok
    templ_cols = None
    try:
        wb_probe = load_workbook(template_path, read_only=True, data_only=True)
        ws_probe = wb_probe.active
        tmp, c, empty_streak = [], 1, 0
        while c <= 500 and empty_streak < 5:
            val = ws_probe.cell(row=header_rows, column=c).value
            if val is None or str(val).strip() == "":
                empty_streak += 1
            else:
                tmp.append(str(val).strip()); empty_streak = 0
            c += 1
        wb_probe.close()
        if tmp:
            templ_cols = tmp
    except Exception:
        templ_cols = None

    if templ_cols:
        exist = [col for col in templ_cols if col in df.columns]
        if exist:
            df = df[exist]

    # Debug: Check groupby operation
    status_cb("Debug: Starting groupby operation...")
    try:
        groups = df.groupby(key_series, dropna=False, sort=False)
        total, current = len(groups), 0
        status_cb(f"Debug: Groupby successful, found {total} groups")
        progress_cb(total, 0)
    except Exception as groupby_e:
        status_cb(f"Debug: Groupby error: {type(groupby_e).__name__}: {str(groupby_e)}")

        # Try multiple approaches to fix the issue
        if "categorical" in str(groupby_e).lower():
            status_cb("Debug: Attempting to fix categorical issue...")

            # Method 1: Try converting all categorical columns to object
            try:
                status_cb("Debug: Method 1 - Converting all categorical columns to object...")
                df_no_cat = df.copy()
                for col in df_no_cat.columns:
                    if df_no_cat[col].dtype.name == 'category':
                        df_no_cat[col] = df_no_cat[col].astype('object')
                groups = df_no_cat.groupby(key_series, dropna=False, sort=False)
                total, current = len(groups), 0
                status_cb(f"Debug: Method 1 successful, found {total} groups")
                progress_cb(total, 0)
                # Update df to use the fixed version
                df = df_no_cat
            except Exception as method1_e:
                status_cb(f"Debug: Method 1 failed: {method1_e}")

                # Method 2: Try using string conversion for groupby
                try:
                    status_cb("Debug: Method 2 - Using string keys for groupby...")
                    string_keys = key_series.astype(str)
                    groups = df.groupby(string_keys, dropna=False, sort=False)
                    total, current = len(groups), 0
                    status_cb(f"Debug: Method 2 successful, found {total} groups")
                    progress_cb(total, 0)
                except Exception as method2_e:
                    status_cb(f"Debug: Method 2 failed: {method2_e}")
                    raise groupby_e
        else:
            raise groupby_e

    for key_val, group in groups:
        current += 1
        status_cb(f"Proses [{current}/{total}] key={key_val}")
        progress_cb(total, current)

        # 1) Tulis XLSX dari template
        wb = load_workbook(template_path)
        ws = wb.active
        start_row = header_rows + 1

        # Comprehensive cleanup to eliminate all Excel repair warnings
        try:
            # 1. Remove all named ranges (common cause of repair warnings)
            if hasattr(wb, 'defined_names'):
                wb.defined_names = {}

            # 2. Remove external links
            if hasattr(wb, '_external_links') and wb._external_links:
                wb._external_links.clear()

            # 3. Remove external link relationships
            if hasattr(wb, '_external_link_rels'):
                wb._external_link_rels.clear()

            # 4. Remove all drawings completely (they often cause repair issues)
            if hasattr(ws, '_drawings'):
                ws._drawings = []

            # 5. Clean up worksheet-level named ranges
            if hasattr(ws, 'defined_names'):
                ws.defined_names = []

            # 6. Remove external link parts from the package
            try:
                from openpyxl.packaging.relationship import Relationship
                if hasattr(wb, '_rels'):
                    # Remove external link relationships
                    rels_to_remove = []
                    for rel in wb._rels:
                        if ('externalLink' in str(rel.target) or
                            'externalLinks' in str(rel.target) or
                            'drawing' in str(rel.target).lower()):
                            rels_to_remove.append(rel)
                    for rel in rels_to_remove:
                        wb._rels.remove(rel)
            except:
                pass

            # 7. Clean up any hyperlinks that might reference external data
            for row in ws.iter_rows():
                for cell in row:
                    if cell.hyperlink:
                        cell.hyperlink = None

        except Exception as cleanup_e:
            # If cleanup fails, continue anyway
            pass

        values = group.fillna("").values.tolist()

        # Copy formatting from template row to data rows
        # Use the first data row from template as formatting template
        template_row = start_row  # This should be the first data row in template

        for r_off, row_vals in enumerate(values, start=0):
            row_idx = start_row + r_off

            # Copy values
            for c_idx, v in enumerate(row_vals, start=1):
                ws.cell(row=row_idx, column=c_idx, value=v)

            # Copy formatting from template row to current row
            if r_off > 0:  # Don't copy formatting to the template row itself
                try:
                    # Copy formatting from template row to current row
                    for col_idx in range(1, len(row_vals) + 1):
                        template_cell = ws.cell(row=template_row, column=col_idx)
                        current_cell = ws.cell(row=row_idx, column=col_idx)

                        # Copy all formatting properties
                        if template_cell.has_style:
                            current_cell.font = template_cell.font.copy()
                            current_cell.fill = template_cell.fill.copy()
                            current_cell.border = template_cell.border.copy()
                            current_cell.alignment = template_cell.alignment.copy()
                            current_cell.number_format = template_cell.number_format
                            current_cell.protection = template_cell.protection.copy()

                except Exception as format_e:
                    # If formatting copy fails, continue without it
                    pass

        last_data_row = start_row + len(values) - 1
        set_print_titles_and_area(ws, header_rows, max(1, group.shape[1]), last_data_row)

        # Build filename with prefix and suffix
        key_part = safe_file_part(key_val)
        parts = []
        if prefix:
            parts.append(prefix)
        parts.append(key_part)
        if suffix:
            parts.append(suffix)

        out_name = " ".join(parts)
        xlsx_out = out_dir / f"{out_name}.xlsx"
        wb.save(xlsx_out)

        # 2) PDF (opsional)
        eng = (pdf_engine or "none").lower()
        if eng != "none":
            if eng == "reportlab":
                export_pdf_pure(group, template_path, header_rows, xlsx_out.with_suffix(".pdf"))
            elif eng == "libreoffice":
                export_pdf_via_lo(xlsx_out, soffice_path=soffice_path)

    status_cb("Selesai.")
    progress_cb(total, total)


# ----------------- Modern Flet GUI -----------------

class ExcelSplitterApp:
    def __init__(self, page: ft.Page):
        self.page = page
        self.setup_page()
        
        # Application state
        self.source_path = ""
        self.sheet_name = ""
        self.key_column = ""
        self.template_path = ""
        self.output_dir = ""
        self.header_rows = 5
        self.pdf_engine = "reportlab"
        self.libreoffice_path = ""
        self.prefix = ""
        self.suffix = ""
        self.loaded_ini = ""
        
        # UI controls
        self.source_field = None
        self.sheet_dropdown = None
        self.key_dropdown = None
        self.template_field = None
        self.output_field = None
        self.header_rows_field = None
        self.pdf_engine_dropdown = None
        self.libreoffice_field = None
        self.prefix_field = None
        self.suffix_field = None
        self.progress_ring = None
        self.status_log = None
        self.generate_button = None
        
        self.is_running = False
        
        self.build_ui()

    def setup_page(self):
        """Configure the main page with modern Fluent Design settings"""
        self.page.title = "Excel Splitter - Modern UI"
        self.page.theme_mode = ft.ThemeMode.SYSTEM
        self.page.window.width = 1000
        self.page.window.height = 700
        self.page.window.min_width = 800
        self.page.window.min_height = 600
        
        # Enable adaptive design for Fluent Design styling
        self.page.adaptive = True
        
        # Set modern color scheme
        self.page.theme = ft.Theme(
            color_scheme_seed=Colors.BLUE,
            use_material3=True
        )

    def build_ui(self):
        """Build the modern UI using Flet components"""
        
        # App Bar with modern styling - Fixed contrast with proper Material Design color
        app_bar = ft.AppBar(
            leading=ft.Icon(Icons.SPLITSCREEN),
            title=ft.Text("Excel Splitter", size=20, weight=ft.FontWeight.W_500),
            center_title=False,
            bgcolor=Colors.SURFACE_CONTAINER_HIGHEST,
            actions=[
                ft.PopupMenuButton(
                    icon=Icons.MORE_VERT,
                    items=[
                        ft.PopupMenuItem(text="Save Configuration", on_click=self.save_config),
                        ft.PopupMenuItem(text="Load Configuration", on_click=self.load_config),
                        ft.PopupMenuItem(),  # Divider
                        ft.PopupMenuItem(text="About", on_click=self.show_about),
                    ]
                )
            ]
        )

        # Debug: Log header styling for contrast analysis
        self.log_status("DEBUG: AppBar created with bgcolor=SURFACE_CONTAINER_HIGHEST (fixed contrast)")
        self.log_status(f"DEBUG: Current theme mode: {self.page.theme_mode}")
        self.log_status(f"DEBUG: Color scheme seed: {Colors.BLUE}")
        
        self.page.appbar = app_bar
        
        # Main content with tabs
        main_tabs = ft.Tabs(
            selected_index=0,
            animation_duration=300,
            tabs=[
                ft.Tab(
                    text="Input",
                    icon=Icons.INPUT,
                    content=self.build_input_tab()
                ),
                ft.Tab(
                    text="Template", 
                    icon=Icons.DESIGN_SERVICES,
                    content=self.build_template_tab()
                ),
                ft.Tab(
                    text="Output",
                    icon=Icons.OUTPUT, 
                    content=self.build_output_tab()
                ),
                ft.Tab(
                    text="Generate",
                    icon=Icons.PLAY_ARROW,
                    content=self.build_generate_tab()
                ),
            ],
        )
        
        # Add to page
        self.page.add(
            ft.Container(
                content=main_tabs,
                padding=20,
                expand=True
            )
        )
        
        self.page.update()

    def build_input_tab(self):
        """Build the input configuration tab"""
        
        # Source file selection
        self.source_field = ft.TextField(
            label="Source Excel File",
            hint_text="Select your Excel file to split",
            read_only=True,
            prefix_icon=Icons.DESCRIPTION,
            expand=True,
            on_change=self.on_form_field_change
        )
        
        source_browse_btn = ft.FilledTonalButton(
            text="Browse",
            icon=Icons.FOLDER_OPEN,
            on_click=self.browse_source_file
        )
        
        # Sheet selection
        self.sheet_dropdown = ft.Dropdown(
            label="Sheet Name",
            hint_text="Select sheet from Excel file",
            prefix_icon=Icons.TABLE_CHART,
            expand=True,
            on_change=self.on_form_field_change
        )
        
        load_sheets_btn = ft.FilledTonalButton(
            text="Load Sheets",
            icon=Icons.REFRESH,
            on_click=self.load_sheets_with_loading
        )
        
        # Key column selection
        self.key_dropdown = ft.Dropdown(
            label="Key Column",
            hint_text="Column to split data by",
            prefix_icon=Icons.KEY,
            expand=True,
            on_change=self.on_key_column_change
        )

        # Debug: Log key dropdown creation
        self.log_status("DEBUG: Key column dropdown created")
        
        load_headers_btn = ft.FilledTonalButton(
            text="Load Headers",
            icon=Icons.VIEW_COLUMN,
            on_click=self.load_headers_with_loading
        )
        
        return ft.Container(
            content=ft.Column([
                ft.Text("Source Configuration", size=18, weight=ft.FontWeight.W_500),
                ft.Divider(height=20),
                
                # Source file row
                ft.Row([
                    self.source_field,
                    source_browse_btn
                ], spacing=10),
                
                # Sheet selection row
                ft.Row([
                    self.sheet_dropdown,
                    load_sheets_btn
                ], spacing=10),
                
                # Key column row
                ft.Row([
                    self.key_dropdown,
                    load_headers_btn
                ], spacing=10),
                
            ], spacing=20),
            padding=20
        )

    def build_template_tab(self):
        """Build the template configuration tab"""
        
        self.template_field = ft.TextField(
            label="Template Excel File",
            hint_text="Template file for formatting output",
            read_only=True,
            prefix_icon=Icons.DESIGN_SERVICES,
            expand=True,
            on_change=self.on_form_field_change
        )
        
        template_browse_btn = ft.FilledTonalButton(
            text="Browse",
            icon=Icons.FOLDER_OPEN,
            on_click=self.browse_template_file
        )
        
        self.header_rows_field = ft.TextField(
            label="Header Rows",
            hint_text="Number of header rows",
            value="5",
            prefix_icon=Icons.FORMAT_LIST_NUMBERED,
            width=200,
            input_filter=ft.NumbersOnlyInputFilter(),
            on_change=self.on_form_field_change
        )
        
        return ft.Container(
            content=ft.Column([
                ft.Text("Template Configuration", size=18, weight=ft.FontWeight.W_500),
                ft.Divider(height=20),
                
                # Template file row
                ft.Row([
                    self.template_field,
                    template_browse_btn
                ], spacing=10),
                
                # Header rows
                self.header_rows_field,
                
                # Info card
                ft.Card(
                    content=ft.Container(
                        content=ft.Column([
                            ft.ListTile(
                                leading=ft.Icon(Icons.INFO_OUTLINE, color=Colors.BLUE),
                                title=ft.Text("Template Usage", weight=ft.FontWeight.W_500),
                                subtitle=ft.Text("The template file provides formatting, styling, and column layout for the output files.")
                            )
                        ]),
                        padding=10
                    )
                )
                
            ], spacing=20),
            padding=20
        )

    def build_output_tab(self):
        """Build the output configuration tab"""
        
        self.output_field = ft.TextField(
            label="Output Folder",
            hint_text="Where to save split files",
            read_only=True,
            prefix_icon=Icons.FOLDER,
            expand=True,
            on_change=self.on_form_field_change
        )
        
        output_browse_btn = ft.FilledTonalButton(
            text="Browse",
            icon=Icons.FOLDER_OPEN,
            on_click=self.browse_output_folder
        )
        
        # PDF Engine selection
        self.pdf_engine_dropdown = ft.Dropdown(
            label="PDF Engine",
            hint_text="Choose PDF generation method",
            prefix_icon=Icons.PICTURE_AS_PDF,
            value="reportlab",
            options=[
                ft.dropdown.Option("none", "None - Excel only"),
                ft.dropdown.Option("reportlab", "ReportLab (Fast)"),
                ft.dropdown.Option("libreoffice", "LibreOffice (Better formatting)")
            ],
            width=300,
            on_change=self.on_form_field_change
        )
        
        # LibreOffice path
        self.libreoffice_field = ft.TextField(
            label="LibreOffice Path (Optional)",
            hint_text="Path to soffice.exe",
            prefix_icon=Icons.INTEGRATION_INSTRUCTIONS,
            expand=True,
            on_change=self.on_form_field_change
        )
        
        libreoffice_browse_btn = ft.FilledTonalButton(
            text="Browse",
            icon=Icons.FOLDER_OPEN,
            on_click=self.browse_libreoffice
        )
        
        # File naming
        self.prefix_field = ft.TextField(
            label="File Prefix",
            hint_text="Text to add before filename",
            prefix_icon=Icons.LABEL,
            width=200,
            on_change=self.on_form_field_change
        )

        self.suffix_field = ft.TextField(
            label="File Suffix",
            hint_text="Text to add after filename",
            prefix_icon=Icons.LABEL_OUTLINE,
            width=200,
            on_change=self.on_form_field_change
        )
        
        return ft.Container(
            content=ft.Column([
                ft.Text("Output Configuration", size=18, weight=ft.FontWeight.W_500),
                ft.Divider(height=20),
                
                # Output folder row
                ft.Row([
                    self.output_field,
                    output_browse_btn
                ], spacing=10),
                
                # PDF engine
                self.pdf_engine_dropdown,
                
                # LibreOffice path row
                ft.Row([
                    self.libreoffice_field,
                    libreoffice_browse_btn
                ], spacing=10),
                
                # File naming row
                ft.Row([
                    self.prefix_field,
                    self.suffix_field
                ], spacing=10),
                
            ], spacing=20),
            padding=20
        )

    def build_generate_tab(self):
        """Build the generate/action tab"""
        
        # Configuration summary card
        source_ref = ft.Ref[ft.Text]()
        sheet_key_ref = ft.Ref[ft.Text]()
        template_ref = ft.Ref[ft.Text]()
        output_ref = ft.Ref[ft.Text]()

        # Store references for later updates
        self.config_summary_refs = {
            'source': source_ref,
            'sheet_key': sheet_key_ref,
            'template': template_ref,
            'output': output_ref
        }

        config_summary = ft.Card(
            content=ft.Container(
                content=ft.Column([
                    ft.Text("Configuration Summary", size=16, weight=ft.FontWeight.W_500),
                    ft.Divider(height=10),

                    ft.ListTile(
                        leading=ft.Icon(Icons.DESCRIPTION),
                        title=ft.Text("Source File"),
                        subtitle=ft.Text("Not selected", ref=source_ref)
                    ),
                    ft.ListTile(
                        leading=ft.Icon(Icons.TABLE_CHART),
                        title=ft.Text("Sheet & Key Column"),
                        subtitle=ft.Text("Not configured", ref=sheet_key_ref)
                    ),
                    ft.ListTile(
                        leading=ft.Icon(Icons.DESIGN_SERVICES),
                        title=ft.Text("Template File"),
                        subtitle=ft.Text("Not selected", ref=template_ref)
                    ),
                    ft.ListTile(
                        leading=ft.Icon(Icons.FOLDER),
                        title=ft.Text("Output Folder"),
                        subtitle=ft.Text("Not selected", ref=output_ref)
                    ),
                ]),
                padding=20
            )
        )
        
        # Generate button
        self.generate_button = ft.FilledButton(
            text="Generate Split Files",
            icon=Icons.PLAY_ARROW,
            style=ft.ButtonStyle(
                padding=ft.padding.all(15),
                text_style=ft.TextStyle(size=16, weight=ft.FontWeight.W_500)
            ),
            on_click=self.start_generation,
            width=200
        )
        
        # Progress indicator
        self.progress_ring = ft.ProgressRing(
            visible=False,
            stroke_width=4
        )
        
        # Status log
        self.status_log = ft.ListView(
            height=300,
            spacing=5,
            padding=10
        )
        
        status_card = ft.Card(
            content=ft.Container(
                content=ft.Column([
                    ft.Text("Processing Status", size=16, weight=ft.FontWeight.W_500),
                    ft.Divider(height=5),
                    self.status_log
                ]),
                padding=10
            )
        )
        
        return ft.Container(
            content=ft.Column([
                config_summary,
                
                ft.Row([
                    self.generate_button,
                    self.progress_ring
                ], alignment=ft.MainAxisAlignment.CENTER, spacing=20),
                
                status_card
                
            ], spacing=20),
            padding=20
        )

    # File picker methods
    async def browse_source_file(self, e):
        """Browse for source Excel file"""
        file_picker = ft.FilePicker(on_result=self.on_source_file_selected)
        self.page.overlay.append(file_picker)
        self.page.update()

        # Ensure configuration summary is initialized
        self.page.after_next_render = self.initialize_config_summary
        
        await file_picker.pick_files(
            allow_multiple=False,
            file_type=ft.FilePickerFileType.CUSTOM,
            allowed_extensions=["xlsx", "xls", "xlsm", "xlsb"]
        )

    def on_source_file_selected(self, e: ft.FilePickerResultEvent):
        """Handle source file selection"""
        if e.files:
            self.source_path = e.files[0].path
            self.source_field.value = self.source_path
            self.log_status(f"Source selected: {Path(self.source_path).name}")
            self.source_field.update()

            # Update configuration summary when file is selected
            self.update_config_summary()

    async def browse_template_file(self, e):
        """Browse for template Excel file"""
        file_picker = ft.FilePicker(on_result=self.on_template_file_selected)
        self.page.overlay.append(file_picker)
        self.page.update()
        
        await file_picker.pick_files(
            allow_multiple=False,
            file_type=ft.FilePickerFileType.CUSTOM,
            allowed_extensions=["xlsx"]
        )

    def on_template_file_selected(self, e: ft.FilePickerResultEvent):
        """Handle template file selection"""
        if e.files:
            self.template_path = e.files[0].path
            self.template_field.value = self.template_path
            self.log_status(f"Template selected: {Path(self.template_path).name}")
            self.template_field.update()

            # Update configuration summary when file is selected
            self.update_config_summary()

    async def browse_output_folder(self, e):
        """Browse for output folder"""
        folder_picker = ft.FilePicker(on_result=self.on_output_folder_selected)
        self.page.overlay.append(folder_picker)
        self.page.update()
        
        await folder_picker.get_directory_path()

    def on_output_folder_selected(self, e: ft.FilePickerResultEvent):
        """Handle output folder selection"""
        if e.path:
            self.output_dir = e.path
            self.output_field.value = self.output_dir
            self.log_status(f"Output folder: {Path(self.output_dir).name}")
            self.output_field.update()

            # Update configuration summary when folder is selected
            self.update_config_summary()

    async def browse_libreoffice(self, e):
        """Browse for LibreOffice executable"""
        file_picker = ft.FilePicker(on_result=self.on_libreoffice_selected)
        self.page.overlay.append(file_picker)
        self.page.update()
        
        await file_picker.pick_files(
            allow_multiple=False,
            file_type=ft.FilePickerFileType.CUSTOM,
            allowed_extensions=["exe"]
        )

    def on_libreoffice_selected(self, e: ft.FilePickerResultEvent):
        """Handle LibreOffice path selection"""
        if e.files:
            self.libreoffice_path = e.files[0].path
            self.libreoffice_field.value = self.libreoffice_path
            self.log_status(f"LibreOffice path set")
            self.libreoffice_field.update()

            # Update configuration summary when LibreOffice path is selected
            self.update_config_summary()

    # Data loading methods
    def load_sheets(self, e):
        """Load sheet names from selected Excel file"""
        if not self.source_path:
            self.show_error("Please select a source Excel file first")
            return

        try:
            xls = pd.ExcelFile(self.source_path)
            sheets = xls.sheet_names

            self.sheet_dropdown.options = [
                ft.dropdown.Option(sheet, sheet) for sheet in sheets
            ]

            if sheets:
                self.sheet_dropdown.value = sheets[0]
                self.sheet_name = sheets[0]

            self.log_status(f"Loaded {len(sheets)} sheets: {', '.join(sheets)}")
            self.sheet_dropdown.update()

            # Update configuration summary
            self.update_config_summary()

        except Exception as ex:
            self.show_error(f"Error loading sheets: {str(ex)}")

    def load_headers(self, e):
        """Load column headers from selected sheet"""
        if not self.source_path or not self.sheet_name:
            self.show_error("Please select source file and sheet first")
            return

        try:
            # Get current sheet name
            self.sheet_name = self.sheet_dropdown.value
            self.log_status(f"DEBUG: Loading headers from sheet '{self.sheet_name}'")

            df = pd.read_excel(self.source_path, sheet_name=self.sheet_name, nrows=0)
            headers = list(df.columns.astype(str))
            index_vals = [f"{i+1}" for i in range(len(headers))]

            # Debug: Log header information
            self.log_status(f"DEBUG: Found {len(headers)} headers: {headers[:5]}{'...' if len(headers) > 5 else ''}")
            self.log_status(f"DEBUG: Index values: {index_vals}")

            # Combine headers and indices
            all_options = []
            for i, header in enumerate(headers):
                all_options.append(ft.dropdown.Option(header, f"{header} (Column {i+1})"))

            for i, idx in enumerate(index_vals):
                all_options.append(ft.dropdown.Option(idx, f"Column {idx}"))

            self.key_dropdown.options = all_options

            if headers:
                self.key_dropdown.value = headers[0]
                self.key_column = headers[0]
                self.log_status(f"DEBUG: Set default key column to: {headers[0]}")

            self.log_status(f"Loaded {len(headers)} headers")
            self.log_status(f"DEBUG: Key dropdown options count: {len(all_options)}")
            self.key_dropdown.update()

            # Update configuration summary
            self.update_config_summary()

        except Exception as ex:
            self.log_status(f"DEBUG: Error loading headers: {type(ex).__name__}: {str(ex)}")
            self.show_error(f"Error loading headers: {str(ex)}")

    def on_key_column_change(self, e):
        """Handle key column dropdown value changes"""
        if e.control.value:
            self.key_column = e.control.value
            self.log_status(f"DEBUG: Key column manually set to: '{self.key_column}'")
        else:
            self.log_status("DEBUG: Key column cleared")

        # Update configuration summary
        self.update_config_summary()

    def on_form_field_change(self, e):
        """Handle form field value changes"""
        # Debug: Log which field triggered the change
        self.log_status(f"DEBUG: Form field changed: {e.control.label if hasattr(e.control, 'label') else e.control.hint_text}")

        # Update configuration summary when any form field changes
        self.update_config_summary()

    async def load_sheets_with_loading(self, e):
        """Load sheets with loading indicator"""
        if not self.source_path:
            self.show_error("Please select a source Excel file first")
            return

        # Show loading state
        original_text = e.control.text
        e.control.text = "Loading..."
        e.control.disabled = True
        e.control.update()

        try:
            # Run the actual loading in a separate task
            await self.load_sheets_async(e)
        finally:
            # Restore button state
            e.control.text = original_text
            e.control.disabled = False
            e.control.update()

    async def load_sheets_async(self, e):
        """Async version of load_sheets"""
        try:
            xls = pd.ExcelFile(self.source_path)
            sheets = xls.sheet_names

            self.sheet_dropdown.options = [
                ft.dropdown.Option(sheet, sheet) for sheet in sheets
            ]

            if sheets:
                self.sheet_dropdown.value = sheets[0]
                self.sheet_name = sheets[0]

            self.log_status(f"Loaded {len(sheets)} sheets: {', '.join(sheets)}")
            self.sheet_dropdown.update()

            # Update configuration summary
            self.update_config_summary()

        except Exception as ex:
            self.show_error(f"Error loading sheets: {str(ex)}")

    async def load_headers_with_loading(self, e):
        """Load headers with loading indicator"""
        if not self.source_path or not self.sheet_name:
            self.show_error("Please select source file and sheet first")
            return

        # Show loading state
        original_text = e.control.text
        e.control.text = "Loading..."
        e.control.disabled = True
        e.control.update()

        try:
            # Run the actual loading in a separate task
            await self.load_headers_async(e)
        finally:
            # Restore button state
            e.control.text = original_text
            e.control.disabled = False
            e.control.update()

    async def load_headers_async(self, e):
        """Async version of load_headers"""
        try:
            # Get current sheet name
            self.sheet_name = self.sheet_dropdown.value
            self.log_status(f"DEBUG: Loading headers from sheet '{self.sheet_name}'")

            df = pd.read_excel(self.source_path, sheet_name=self.sheet_name, nrows=0)
            headers = list(df.columns.astype(str))
            index_vals = [f"{i+1}" for i in range(len(headers))]

            # Debug: Log header information
            self.log_status(f"DEBUG: Found {len(headers)} headers: {headers[:5]}{'...' if len(headers) > 5 else ''}")
            self.log_status(f"DEBUG: Index values: {index_vals}")

            # Combine headers and indices
            all_options = []
            for i, header in enumerate(headers):
                all_options.append(ft.dropdown.Option(header, f"{header} (Column {i+1})"))

            for i, idx in enumerate(index_vals):
                all_options.append(ft.dropdown.Option(idx, f"Column {idx}"))

            self.key_dropdown.options = all_options

            if headers:
                self.key_dropdown.value = headers[0]
                self.key_column = headers[0]
                self.log_status(f"DEBUG: Set default key column to: {headers[0]}")

            self.log_status(f"Loaded {len(headers)} headers")
            self.log_status(f"DEBUG: Key dropdown options count: {len(all_options)}")
            self.key_dropdown.update()

            # Update configuration summary
            self.update_config_summary()

        except Exception as ex:
            self.log_status(f"DEBUG: Error loading headers: {type(ex).__name__}: {str(ex)}")
            self.show_error(f"Error loading headers: {str(ex)}")

    # Generation methods
    def start_generation(self, e):
        """Start the Excel splitting process"""
        if self.is_running:
            return
            
        # Validate inputs
        validation_error = self.validate_inputs()
        if validation_error:
            self.show_error(validation_error)
            return
        
        # Collect current values
        self.collect_form_values()
        
        self.is_running = True
        self.generate_button.disabled = True
        self.progress_ring.visible = True
        self.generate_button.update()
        self.progress_ring.update()
        
        # Clear previous logs
        self.status_log.controls.clear()
        self.status_log.update()
        
        # Start processing in thread
        threading.Thread(target=self.run_generation, daemon=True).start()

    def collect_form_values(self):
        """Collect all form values"""
        self.source_path = self.source_field.value or ""
        self.sheet_name = self.sheet_dropdown.value or ""
        self.key_column = self.key_dropdown.value or ""
        self.template_path = self.template_field.value or ""
        self.output_dir = self.output_field.value or ""
        self.header_rows = int(self.header_rows_field.value or "5")
        self.pdf_engine = self.pdf_engine_dropdown.value or "reportlab"
        self.libreoffice_path = self.libreoffice_field.value or ""
        self.prefix = self.prefix_field.value or ""
        self.suffix = self.suffix_field.value or ""

        # Debug: Log collected form values
        self.log_status(f"DEBUG: Collected form values - source: {self.source_path}")
        self.log_status(f"DEBUG: Collected form values - sheet: {self.sheet_name}")
        self.log_status(f"DEBUG: Collected form values - key_column: '{self.key_column}'")
        self.log_status(f"DEBUG: Collected form values - template: {self.template_path}")
        self.log_status(f"DEBUG: Collected form values - output: {self.output_dir}")
        self.log_status(f"DEBUG: Collected form values - header_rows: {self.header_rows}")

        # Update configuration summary
        self.update_config_summary()

    def initialize_config_summary(self):
        """Initialize configuration summary after UI is rendered"""
        self.log_status("DEBUG: Initializing configuration summary")
        self.update_config_summary()

    def update_config_summary(self):
        """Update the configuration summary with current values"""
        if not hasattr(self, 'config_summary_refs') or not self.config_summary_refs:
            self.log_status("DEBUG: Config summary refs not initialized yet")
            return

        # Get summary text references
        source_ref = self.config_summary_refs['source']
        sheet_key_ref = self.config_summary_refs['sheet_key']
        template_ref = self.config_summary_refs['template']
        output_ref = self.config_summary_refs['output']

        # Update source file summary
        if self.source_path:
            source_ref.value = Path(self.source_path).name
        else:
            source_ref.value = "Not selected"

        # Update sheet & key column summary
        if self.sheet_name and self.key_column:
            sheet_key_ref.value = f"Sheet: {self.sheet_name}, Key: {self.key_column}"
        else:
            sheet_key_ref.value = "Not configured"

        # Update template file summary
        if self.template_path:
            template_ref.value = Path(self.template_path).name
        else:
            template_ref.value = "Not selected"

        # Update output folder summary
        if self.output_dir:
            output_ref.value = Path(self.output_dir).name
        else:
            output_ref.value = "Not selected"

        # Debug: Log current values
        self.log_status(f"DEBUG: Summary update - source: {self.source_path}, sheet: {self.sheet_name}, key: {self.key_column}")
        self.log_status(f"DEBUG: Summary update - template: {self.template_path}, output: {self.output_dir}")

        # Update the page to reflect changes
        try:
            self.page.update()
            self.log_status("DEBUG: Configuration summary updated successfully")
        except Exception as e:
            self.log_status(f"DEBUG: Error updating configuration summary: {str(e)}")

    def validate_inputs(self):
        """Validate all required inputs"""
        if not self.source_field.value:
            return "Source Excel file is required"
        if not self.sheet_dropdown.value:
            return "Sheet name is required"  
        if not self.key_dropdown.value:
            return "Key column is required"
        if not self.template_field.value:
            return "Template file is required"
        if not self.output_field.value:
            return "Output folder is required"
        
        # Check if files exist
        if not Path(self.source_field.value).exists():
            return "Source Excel file does not exist"
        if not Path(self.template_field.value).exists():
            return "Template file does not exist"
        if not Path(self.output_field.value).exists():
            return "Output folder does not exist"
        
        return None

    def run_generation(self):
        """Run the actual Excel splitting process"""
        try:
            # Convert key column to appropriate type
            try:
                key_col = int(self.key_column)
                self.log_status(f"DEBUG: Converting key column '{self.key_column}' to int: {key_col}")
            except ValueError:
                key_col = self.key_column
                self.log_status(f"DEBUG: Using key column as string: '{key_col}'")

            # Validate ReportLab if needed
            if self.pdf_engine == "reportlab" and not REPORTLAB_AVAILABLE:
                self.page.add(ft.SnackBar(
                    content=ft.Text("ReportLab not available. Install with: pip install reportlab"),
                    bgcolor=Colors.RED
                ))
                return

            # Find LibreOffice if needed
            soffice_path = None
            if self.pdf_engine == "libreoffice":
                soffice_path = find_soffice(self.libreoffice_path)
                if not soffice_path:
                    self.page.add(ft.SnackBar(
                        content=ft.Text("LibreOffice (soffice.exe) not found"),
                        bgcolor=Colors.RED
                    ))
                    return

            # Run the splitting process
            split_excel_with_template(
                source_path=Path(self.source_path),
                sheet_name=self.sheet_name,
                key_col=key_col,
                template_path=Path(self.template_path),
                out_dir=Path(self.output_dir),
                header_rows=self.header_rows,
                pdf_engine=self.pdf_engine,
                soffice_path=soffice_path,
                prefix=self.prefix,
                suffix=self.suffix,
                status_cb=self.log_status,
                progress_cb=self.update_progress
            )
            
            # Show success message
            self.page.add(ft.SnackBar(
                content=ft.Text("Excel splitting completed successfully!"),
                bgcolor=Colors.GREEN_600
            ))
            
        except Exception as ex:
            self.log_status(f"Error: {str(ex)}")
            self.page.add(ft.SnackBar(
                content=ft.Text(f"Error: {str(ex)}"),
                bgcolor=Colors.RED
            ))
        
        finally:
            # Re-enable UI
            self.is_running = False
            self.generate_button.disabled = False
            self.progress_ring.visible = False
            self.generate_button.update()
            self.progress_ring.update()

    def log_status(self, message: str):
        """Add status message to the log"""
        def add_log():
            log_item = ft.Container(
                content=ft.Text(
                    message,
                    size=12,
                    color=Colors.ON_SURFACE_VARIANT
                ),
                padding=ft.padding.symmetric(horizontal=10, vertical=5),
                border_radius=5,
                bgcolor=Colors.GREY_400
            )
            
            self.status_log.controls.append(log_item)
            
            # Keep only last 50 messages
            if len(self.status_log.controls) > 50:
                self.status_log.controls.pop(0)
            
            self.status_log.update()
            
            # Auto-scroll to bottom
            if hasattr(self.status_log, 'scroll_to'):
                self.status_log.scroll_to(offset=-1)
        
        # Use page update instead of add_async
        try:
            add_log()
        except:
            # Fallback if update fails
            pass

    def update_progress(self, total: int, current: int):
        """Update progress indicator"""
        if total > 0:
            percentage = (current / total) * 100
            self.log_status(f"Progress: {current}/{total} ({percentage:.1f}%)")

            # Update progress ring if it exists
            if hasattr(self, 'progress_ring') and self.progress_ring:
                self.progress_ring.value = percentage / 100.0
                self.progress_ring.update()

    # Utility methods
    def show_error(self, message: str):
        """Show error message"""
        self.page.add(ft.SnackBar(
            content=ft.Text(message),
            bgcolor=Colors.RED
        ))

    def show_about(self, e):
        """Show about dialog"""
        about_dialog = ft.AlertDialog(
            modal=True,
            title=ft.Text("About Excel Splitter"),
            content=ft.Column([
                ft.Text("Modern Excel Splitter with Flet UI"),
                ft.Text("Version 2.0 - Fluent Design"),
                ft.Divider(height=10),
                ft.Text("Features:"),
                ft.Text("• Template-based splitting"),
                ft.Text("• PDF export support"),
                ft.Text("• Modern Fluent Design UI"),
                ft.Text("• Configuration save/load"),
            ], tight=True),
            actions=[
                ft.TextButton("Close", on_click=lambda _: self.close_dialog(about_dialog))
            ]
        )
        
        self.page.add(about_dialog)
        about_dialog.open = True
        self.page.update()

    def close_dialog(self, dialog):
        """Close a dialog"""
        dialog.open = False
        self.page.update()

    # Configuration methods
    async def save_config(self, e):
        """Save current configuration to INI file"""
        try:
            # Collect current form values
            self.collect_form_values()

            file_picker = ft.FilePicker(on_result=self.on_config_save_result)
            self.page.overlay.append(file_picker)
            self.page.update()

            await file_picker.save_file(
                file_type=ft.FilePickerFileType.CUSTOM,
                allowed_extensions=["ini"],
                dialog_title="Save Configuration As"
            )
        except Exception as ex:
            self.show_error(f"Error saving configuration: {str(ex)}")

    def on_config_save_result(self, e: ft.FilePickerResultEvent):
        """Handle configuration save result"""
        if e.path:
            try:
                config = configparser.ConfigParser()
                config.add_section('ExcelSplitter')

                # Save all current configuration values
                config.set('ExcelSplitter', 'source_path', self.source_path or '')
                config.set('ExcelSplitter', 'sheet_name', self.sheet_name or '')
                config.set('ExcelSplitter', 'key_column', self.key_column or '')
                config.set('ExcelSplitter', 'template_path', self.template_path or '')
                config.set('ExcelSplitter', 'output_dir', self.output_dir or '')
                config.set('ExcelSplitter', 'header_rows', str(self.header_rows))
                config.set('ExcelSplitter', 'pdf_engine', self.pdf_engine or 'reportlab')
                config.set('ExcelSplitter', 'libreoffice_path', self.libreoffice_path or '')
                config.set('ExcelSplitter', 'prefix', self.prefix or '')
                config.set('ExcelSplitter', 'suffix', self.suffix or '')

                # Write to file
                with open(e.path, 'w') as configfile:
                    config.write(configfile)

                self.log_status(f"Configuration saved to: {e.path}")
                self.page.add(ft.SnackBar(
                    content=ft.Text(f"Configuration saved successfully to {Path(e.path).name}"),
                    bgcolor=Colors.GREEN_600
                ))

            except Exception as ex:
                self.show_error(f"Error saving configuration: {str(ex)}")
        else:
            self.log_status("Configuration save cancelled")

    async def load_config(self, e):
        """Load configuration from INI file"""
        try:
            file_picker = ft.FilePicker(on_result=self.on_config_load_result)
            self.page.overlay.append(file_picker)
            self.page.update()

            await file_picker.pick_files(
                allow_multiple=False,
                file_type=ft.FilePickerFileType.CUSTOM,
                allowed_extensions=["ini"],
                dialog_title="Load Configuration"
            )
        except Exception as ex:
            self.show_error(f"Error loading configuration: {str(ex)}")

    def on_config_load_result(self, e: ft.FilePickerResultEvent):
        """Handle configuration load result"""
        if e.files:
            try:
                config = configparser.ConfigParser()
                config.read(e.files[0].path)

                if 'ExcelSplitter' in config:
                    # Load configuration values
                    section = config['ExcelSplitter']

                    # Update form fields with loaded values
                    if 'source_path' in section and section['source_path']:
                        self.source_path = section['source_path']
                        self.source_field.value = self.source_path
                        self.source_field.update()

                    if 'sheet_name' in section and section['sheet_name']:
                        self.sheet_name = section['sheet_name']
                        self.sheet_dropdown.value = self.sheet_name
                        self.sheet_dropdown.update()

                    if 'key_column' in section and section['key_column']:
                        self.key_column = section['key_column']
                        self.key_dropdown.value = self.key_column
                        self.key_dropdown.update()

                    if 'template_path' in section and section['template_path']:
                        self.template_path = section['template_path']
                        self.template_field.value = self.template_path
                        self.template_field.update()

                    if 'output_dir' in section and section['output_dir']:
                        self.output_dir = section['output_dir']
                        self.output_field.value = self.output_dir
                        self.output_field.update()

                    if 'header_rows' in section:
                        self.header_rows = int(section['header_rows'])
                        self.header_rows_field.value = str(self.header_rows)
                        self.header_rows_field.update()

                    if 'pdf_engine' in section:
                        self.pdf_engine = section['pdf_engine']
                        self.pdf_engine_dropdown.value = self.pdf_engine
                        self.pdf_engine_dropdown.update()

                    if 'libreoffice_path' in section:
                        self.libreoffice_path = section['libreoffice_path']
                        self.libreoffice_field.value = self.libreoffice_path
                        self.libreoffice_field.update()

                    if 'prefix' in section:
                        self.prefix = section['prefix']
                        self.prefix_field.value = self.prefix
                        self.prefix_field.update()

                    if 'suffix' in section:
                        self.suffix = section['suffix']
                        self.suffix_field.value = self.suffix
                        self.suffix_field.update()

                    # Update configuration summary
                    self.update_config_summary()

                    self.log_status(f"Configuration loaded from: {e.files[0].path}")
                    self.page.add(ft.SnackBar(
                        content=ft.Text(f"Configuration loaded successfully from {Path(e.files[0].path).name}"),
                        bgcolor=Colors.GREEN_600
                    ))

                    # Refresh sheet names if source file is loaded
                    if self.source_path:
                        try:
                            xls = pd.ExcelFile(self.source_path)
                            sheets = xls.sheet_names
                            self.sheet_dropdown.options = [ft.dropdown.Option(sheet, sheet) for sheet in sheets]
                            if self.sheet_name not in sheets:
                                self.sheet_name = sheets[0] if sheets else ""
                                self.sheet_dropdown.value = self.sheet_name
                            self.sheet_dropdown.update()
                        except Exception as ex:
                            self.log_status(f"Warning: Could not refresh sheets: {str(ex)}")

                    # Refresh headers if source file and sheet are loaded
                    if self.source_path and self.sheet_name:
                        try:
                            df = pd.read_excel(self.source_path, sheet_name=self.sheet_name, nrows=0)
                            headers = list(df.columns.astype(str))
                            all_options = []
                            for i, header in enumerate(headers):
                                all_options.append(ft.dropdown.Option(header, f"{header} (Column {i+1})"))
                            for i in range(len(headers)):
                                all_options.append(ft.dropdown.Option(f"{i+1}", f"Column {i+1}"))

                            self.key_dropdown.options = all_options
                            if self.key_column not in [opt.key for opt in all_options]:
                                self.key_column = headers[0] if headers else ""
                                self.key_dropdown.value = self.key_column
                            self.key_dropdown.update()
                        except Exception as ex:
                            self.log_status(f"Warning: Could not refresh headers: {str(ex)}")

                    self.page.update()

                else:
                    self.show_error("Invalid configuration file - missing ExcelSplitter section")

            except Exception as ex:
                self.show_error(f"Error loading configuration: {str(ex)}")
        else:
            self.log_status("Configuration load cancelled")


def main(page: ft.Page):
    """Main application entry point"""
    app = ExcelSplitterApp(page)


if __name__ == "__main__":
    ft.app(target=main, assets_dir="assets")